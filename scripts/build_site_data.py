from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from database import (
    connect,
    init_db,
    replace_sales_rows_for_date,
    taipei_now_text,
    upsert_sales_import,
)


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
DATA_JSON = DOCS / "data.json"
SALES_FILE = Path(os.environ["SALES_FILE_PATH"]) if os.environ.get("SALES_FILE_PATH") else None
SALES_DATE = os.environ.get("SALES_DATE", "")
IMPORT_SALES_ONLY = os.environ.get("IMPORT_SALES_ONLY") == "1"
INVENTORY_FILES = [
    ("A", ROOT / "115年04月庫存(A) 30.xlsx"),
    ("B", ROOT / "115年04月庫存(B) 30.xlsx"),
]
HARVEST_FILE = ROOT / "每日採收規劃表0429.xlsx"
AVAILABILITY_FILE = ROOT / "115每日菜量表0228.xlsx"
OPENING_STOCK_MONTHS = {"2026-04"}
EXCLUDED_PRODUCT_NAMES = {"\u6cb9\u83dc", "\u5343\u5bf6\u83dc", "\u7d05\u5377", "\u7d05\u6372"}


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def is_excluded_product_name(value: Any) -> bool:
    text = clean(value)
    return bool(text and text in EXCLUDED_PRODUCT_NAMES)


def normalize_store_name(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    return re.sub(r"\s+", "", text)


def normalize_store_code(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    digits = re.sub(r"\D+", "", text)
    if digits:
        return digits
    return text


def formula_display_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    match = re.fullmatch(r'=HYPERLINK\([^,]+,\s*"([^"]+)"\)', value)
    if match:
        return match.group(1)
    return value


def num(value: Any) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0


def excel_date_label(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)) and 30000 < value < 70000:
        # Excel serial date. The current workbooks already mostly expose dates
        # as datetime objects, but keep this for older monthly files.
        base = datetime(1899, 12, 30)
        return (base + __import__("datetime").timedelta(days=float(value))).date().isoformat()
    text = clean(value)
    if not text:
        return None
    if "00:00:00" in text:
        return text.split(" ")[0]
    return text


def parse_sales() -> dict[str, Any]:
    if SALES_FILE is None:
        return {
            "date": SALES_DATE,
            "rows": [],
            "pivot": [],
            "totals": {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0},
            "sourceFiles": [],
        }
    parsed = parse_sales_file(SALES_FILE, SALES_DATE)
    parsed["sourceFiles"] = [SALES_FILE.name]
    return parsed


def parse_sales_file(path: Path, date: str) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [clean(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {name: idx for idx, name in enumerate(header) if name}
    required = ["門市代號", "門市名稱", "呼出碼", "商品名稱", "銷售量", "丟棄量", "期末庫存"]
    missing = [name for name in required if name not in col]
    if missing:
        raise ValueError(f"SalesData missing columns: {missing}")

    rows: list[dict[str, Any]] = []
    pivot: dict[tuple[str, str], dict[str, Any]] = {}
    totals = {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0}

    for values in ws.iter_rows(min_row=2, values_only=True):
        store_code = clean(values[col["門市代號"]])
        product_code = clean(values[col["呼出碼"]])
        if not store_code or not product_code:
            continue
        product_name = clean(values[col["商品名稱"]])
        if is_excluded_product_name(product_name):
            continue
        row = {
            "date": date,
            "storeCode": store_code,
            "storeName": clean(values[col["門市名稱"]]),
            "productCode": product_code,
            "productName": product_name,
            "salesQty": num(values[col["銷售量"]]),
            "discardQty": num(values[col["丟棄量"]]),
            "endingStock": num(values[col["期末庫存"]]),
        }
        rows.append(row)
        key = (store_code, product_code)
        if key not in pivot:
            pivot[key] = {
                "date": date,
                "storeCode": store_code,
                "storeName": row["storeName"],
                "productCode": product_code,
                "productName": row["productName"],
                "salesQty": 0.0,
                "discardQty": 0.0,
                "endingStock": 0.0,
            }
        pivot[key]["salesQty"] += row["salesQty"]
        pivot[key]["discardQty"] += row["discardQty"]
        pivot[key]["endingStock"] += row["endingStock"]
        totals["salesQty"] += row["salesQty"]
        totals["discardQty"] += row["discardQty"]
        totals["endingStock"] += row["endingStock"]

    return {"date": date, "rows": rows, "pivot": list(pivot.values()), "totals": totals}


def persist_sales_import(sales_date: str, source_file: Path, rows: list[dict[str, Any]]) -> None:
    conn = connect()
    try:
        init_db(conn)
        replace_sales_rows_for_date(
            conn,
            sales_date,
            [
                row
                for row in rows
                if row.get("storeCode") and row.get("productCode") and not is_excluded_product_name(row.get("productName"))
            ],
        )
        upsert_sales_import(conn, sales_date, source_file.name, "")
    finally:
        conn.close()


def read_sales_rows_from_db() -> tuple[list[dict[str, Any]], list[str]]:
    conn = connect()
    try:
        init_db(conn)
        rows = conn.execute(
            """
            SELECT
                sales_date,
                store_code,
                store_name,
                product_code,
                product_name,
                sales_qty,
                discard_qty,
                ending_stock
            FROM sales_rows
            ORDER BY sales_date, id
            """
        ).fetchall()
        sources = conn.execute(
            "SELECT file_name FROM sales_imports ORDER BY sales_date"
        ).fetchall()
    finally:
        conn.close()
    sales_rows = [
        {
            "date": row["sales_date"],
            "storeCode": row["store_code"],
            "storeName": row["store_name"],
            "productCode": row["product_code"],
            "productName": row["product_name"],
            "salesQty": num(row["sales_qty"]),
            "discardQty": num(row["discard_qty"]),
            "endingStock": num(row["ending_stock"]),
        }
        for row in rows
        if not is_excluded_product_name(row["product_name"])
    ]
    return sales_rows, [row["file_name"] for row in sources]


def read_sales_import_metadata() -> dict[str, Any]:
    conn = connect()
    try:
        init_db(conn)
        rows = conn.execute("SELECT sales_date, file_name FROM sales_imports ORDER BY sales_date").fetchall()
    finally:
        conn.close()
    return {
        "date": rows[-1]["sales_date"] if rows else SALES_DATE,
        "rows": [],
        "pivot": [],
        "totals": {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0},
        "sourceFiles": [row["file_name"] for row in rows],
    }


def parse_all_sales_imports() -> dict[str, Any]:
    if SALES_FILE is not None and SALES_DATE:
        parsed = parse_sales_file(SALES_FILE, SALES_DATE)
        persist_sales_import(SALES_DATE, SALES_FILE, parsed["rows"])

    return read_sales_import_metadata()


def import_sales_only() -> None:
    if SALES_FILE is None:
        raise ValueError("SALES_FILE_PATH is required")
    if not SALES_DATE:
        raise ValueError("SALES_DATE is required")
    parsed = parse_sales_file(SALES_FILE, SALES_DATE)
    persist_sales_import(SALES_DATE, SALES_FILE, parsed["rows"])
    print(
        json.dumps(
            {
                "salesDate": SALES_DATE,
                "salesRows": len(parsed["rows"]),
                "salesPivot": len(parsed["pivot"]),
                "salesTotals": parsed["totals"],
            },
            ensure_ascii=False,
        )
    )


def build_sales_snapshot_from_db() -> dict[str, Any]:
    rows, source_files = read_sales_rows_from_db()
    if not rows:
        return parse_sales()
    all_rows: list[dict[str, Any]] = []
    pivot: dict[tuple[str, str, str], dict[str, Any]] = {}
    totals = {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0}
    latest_date = ""
    for row in rows:
        latest_date = str(row.get("date") or latest_date)
        all_rows.append(row)
        key = (row["date"], row["storeCode"], row["productCode"])
        if key not in pivot:
            pivot[key] = {
                "date": row["date"],
                "storeCode": row["storeCode"],
                "storeName": row["storeName"],
                "productCode": row["productCode"],
                "productName": row["productName"],
                "salesQty": 0.0,
                "discardQty": 0.0,
                "endingStock": 0.0,
            }
        pivot[key]["salesQty"] += num(row.get("salesQty"))
        pivot[key]["discardQty"] += num(row.get("discardQty"))
        pivot[key]["endingStock"] += num(row.get("endingStock"))
        for key in totals:
            totals[key] += num(row.get(key))
    return {
        "date": latest_date or SALES_DATE,
        "rows": all_rows,
        "pivot": list(pivot.values()),
        "totals": totals,
        "sourceFiles": source_files,
    }


def parse_inventory_workbook(group: str, path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    wb_formula = load_workbook(path, read_only=True, data_only=False)
    stores: list[dict[str, Any]] = []
    inventory: list[dict[str, Any]] = []
    mail_sales: list[dict[str, Any]] = []
    mail_discards: list[dict[str, Any]] = []
    store_pages: list[dict[str, Any]] = []
    mail_stores: list[dict[str, Any]] = []
    mail_code_by_name: dict[str, str] = {}
    mail_order_by_code: dict[str, int] = {}
    mail_order_by_name: dict[str, int] = {}

    if "mail" in wb.sheetnames:
        ws = wb["mail"]
        ws_formula = wb_formula["mail"]
        mail_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
        formula_rows = list(
            ws_formula.iter_rows(min_row=1, max_row=ws_formula.max_row, max_col=ws_formula.max_column, values_only=True)
        )
        store_codes = [normalize_store_code(v) for v in mail_rows[1][2:]]
        store_name_values = [formula_display_text(v) for v in formula_rows[2][2:]]
        store_names = [normalize_store_name(v) for v in store_name_values]
        mail_stores = [
            {"mailOrder": idx, "storeCode": code, "storeName": name}
            for idx, (code, name) in enumerate(zip(store_codes, store_names))
            if code or name
        ]
        mail_code_by_name = {name: code for name, code in zip(store_names, store_codes) if name and code}
        mail_order_by_code = {code: idx for idx, code in enumerate(store_codes) if code}
        mail_order_by_name = {name: idx for idx, name in enumerate(store_names) if name}

        section = "sales"
        for row in mail_rows:
            label = clean(row[1] if len(row) > 1 else None)
            if label and "丟棄" in label:
                section = "discard"
            product_code = label
            if not product_code or not product_code.isdigit():
                continue
            for idx, value in enumerate(row[2:]):
                qty = num(value)
                if qty == 0:
                    continue
                item = {
                    "group": group,
                    "productCode": product_code,
                    "storeCode": normalize_store_code(store_codes[idx]),
                    "storeName": normalize_store_name(store_names[idx]),
                    "quantity": qty,
                }
                if section == "discard":
                    mail_discards.append(item)
                else:
                    mail_sales.append(item)

    for sheet_name in wb.sheetnames:
        if sheet_name in {"mail", "總銷量", "工作表1", "工作表2"}:
            continue
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 70), max_col=min(ws.max_column, 76), values_only=True))
        get = lambda r, c: sheet_rows[r - 1][c - 1] if len(sheet_rows) >= r and len(sheet_rows[r - 1]) >= c else None
        sheet_store_name = normalize_store_name(get(5, 2)) or normalize_store_name(sheet_name)
        mail_index_raw = get(3, 1)
        mail_index = int(mail_index_raw) - 1 if isinstance(mail_index_raw, (int, float)) and int(mail_index_raw) >= 1 else None
        mail_store = mail_stores[mail_index] if mail_index is not None and mail_index < len(mail_stores) else None
        name_match = mail_code_by_name.get(sheet_store_name)
        if name_match:
            store_name = sheet_store_name
            store_code = name_match
            store_order = mail_order_by_name.get(sheet_store_name, len(stores))
        elif mail_store:
            store_name = normalize_store_name(mail_store.get("storeName")) or sheet_store_name
            store_code = normalize_store_code(mail_store.get("storeCode"))
            store_order = mail_store.get("mailOrder", len(stores))
        else:
            store_name = sheet_store_name
            store_code = None
            store_order = len(stores)
        route = clean(get(4, 2))
        stores.append(
            {
                "group": group,
                "sheet": sheet_name,
                "storeName": store_name,
                "storeCode": store_code,
                "route": route,
                "mailOrder": store_order,
            }
        )

        headers = [clean(v) for v in (sheet_rows[3] if len(sheet_rows) > 3 else [])]
        dates = [excel_date_label(v) for v in (sheet_rows[4] if len(sheet_rows) > 4 else [])]
        sale_cols = [idx for idx, header in enumerate(headers) if header == "出售" and dates[idx]]
        first_sale_col = min(sale_cols) if sale_cols else len(headers)
        month_dates = [dates[idx] for idx in sale_cols]
        inventory_month = next(
            (
                date[:7]
                for date in month_dates
                if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date))
            ),
            "",
        )
        store_products: list[dict[str, Any]] = []
        for row in sheet_rows[5:]:
            product_code = clean(row[0] if len(row) > 0 else None)
            product_name = clean(row[1] if len(row) > 1 else None)
            if not product_code or not product_name:
                continue
            if is_excluded_product_name(product_name):
                continue
            base_stock = num(row[2] if len(row) > 2 else None)
            shipment = 0.0
            shipment_codes: list[str] = []
            shipment_events: list[dict[str, Any]] = []
            for idx, header in enumerate(headers):
                if idx < first_sale_col and header in {"到貨", "出貨"} and dates[idx]:
                    shipment_date = str(dates[idx])
                    value = num(row[idx] if len(row) > idx else None)
                    if not value:
                        continue
                    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", shipment_date) or shipment_date[:7] != inventory_month:
                        continue
                    shipment += value
                    shipment_codes.append(shipment_date)
                    shipment_events.append({"date": shipment_date, "quantity": value})
            opening_stocks_by_month = {inventory_month: base_stock} if inventory_month in OPENING_STOCK_MONTHS else {}
            current_stock = base_stock + shipment
            sale_values = [0 for _ in sale_cols]
            discard_values = [0 for _ in sale_cols]
            store_products.append(
                {
                    "productCode": product_code,
                    "productName": product_name,
                    "openingStocksByMonth": opening_stocks_by_month,
                    "sales": sale_values,
                    "discards": discard_values,
                    "monthlySales": 0,
                    "discardQty": 0,
                    "adjustmentQty": 0,
                    "existingShipmentQty": shipment,
                    "shipments": shipment_events,
                    "shipmentInputQty": 0,
                    "currentStock": current_stock,
                }
            )
            inventory.append(
                {
                    "group": group,
                    "storeSheet": sheet_name,
                    "storeName": store_name,
                    "storeCode": store_code,
                    "route": route,
                    "mailOrder": store_order,
                    "productCode": product_code,
                    "productName": product_name,
                    "baseStock": base_stock,
                    "salesQty": 0,
                    "discardQty": 0,
                    "shipmentQty": shipment,
                    "currentStock": current_stock,
                    "shipmentDates": shipment_codes,
                }
            )
        store_pages.append(
            {
                "key": f"{group}:{sheet_name}",
                "group": group,
                "sheet": sheet_name,
                "storeName": store_name,
                "storeCode": store_code,
                "route": route,
                "mailOrder": store_order,
                "dates": month_dates,
                "products": store_products,
            }
        )

    return {"stores": stores, "inventory": inventory, "mailSales": mail_sales, "mailDiscards": mail_discards, "storePages": store_pages}


def parse_inventory() -> dict[str, Any]:
    all_stores: list[dict[str, Any]] = []
    all_inventory: list[dict[str, Any]] = []
    all_mail_sales: list[dict[str, Any]] = []
    all_mail_discards: list[dict[str, Any]] = []
    all_store_pages: list[dict[str, Any]] = []
    for group, path in INVENTORY_FILES:
        parsed = parse_inventory_workbook(group, path)
        all_stores.extend(parsed["stores"])
        all_inventory.extend(parsed["inventory"])
        all_mail_sales.extend(parsed["mailSales"])
        all_mail_discards.extend(parsed["mailDiscards"])
        all_store_pages.extend(parsed["storePages"])
    return {
        "stores": all_stores,
        "inventory": all_inventory,
        "mailSales": all_mail_sales,
        "mailDiscards": all_mail_discards,
        "storePages": all_store_pages,
    }


def latest_inventory_date(inventory: dict[str, Any]) -> str:
    return max(
        (
            date
            for store in inventory.get("storePages", [])
            for date in store.get("dates", [])
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date))
        ),
        default="",
    )


def parse_availability() -> list[dict[str, Any]]:
    wb = load_workbook(AVAILABILITY_FILE, read_only=True, data_only=True)
    ws = wb["菜量表"]
    date_cols = [(c, clean(ws.cell(2, c).value)) for c in range(2, ws.max_column, 2) if clean(ws.cell(2, c).value)]
    records: list[dict[str, Any]] = []
    product: str | None = None
    for r in range(3, ws.max_row + 1):
        if clean(ws.cell(r, 1).value):
            product = clean(ws.cell(r, 1).value)
        if not product:
            continue
        if is_excluded_product_name(product):
            continue
        for c, label in date_cols:
            lot = clean(ws.cell(r, c).value)
            qty = num(ws.cell(r, c + 1).value)
            if lot or qty:
                records.append({"dateLabel": label, "productName": product, "lot": lot, "quantity": qty})
    return records


def parse_harvest_orders() -> list[dict[str, Any]]:
    wb = load_workbook(HARVEST_FILE, read_only=True, data_only=True)
    orders: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("採收"):
            continue
        ws = wb[sheet_name]
        store_codes = [clean(ws.cell(3, c).value) for c in range(3, ws.max_column + 1)]
        store_names = [clean(ws.cell(4, c).value) for c in range(3, ws.max_column + 1)]
        for r in range(5, ws.max_row + 1):
            product_code = clean(ws.cell(r, 1).value)
            product_name = clean(ws.cell(r, 2).value)
            if not product_name:
                continue
            if is_excluded_product_name(product_name):
                continue
            for idx, c in enumerate(range(3, ws.max_column + 1)):
                qty = num(ws.cell(r, c).value)
                if qty:
                    orders.append(
                        {
                            "sheet": sheet_name,
                            "storeCode": store_codes[idx],
                            "storeName": store_names[idx],
                            "productCode": product_code,
                            "productName": product_name,
                            "quantity": qty,
                        }
                    )
    return orders


def parse_harvest_planning_template() -> dict[str, Any]:
    sheet_name = "採收(五)"
    wb_values = load_workbook(HARVEST_FILE, read_only=False, data_only=True)
    wb_formulas = load_workbook(HARVEST_FILE, read_only=False, data_only=False)
    ws_values = wb_values[sheet_name]
    ws_formulas = wb_formulas[sheet_name]
    min_row = 3
    max_row = 48
    min_col = 2
    max_col = 16
    product_rows = list(range(5, 41)) + list(range(42, 49))
    hidden_rows = {
        row
        for row in product_rows
        if is_excluded_product_name(ws_values.cell(row, 2).value)
    }
    cells: list[dict[str, Any]] = []
    for row in range(min_row, max_row + 1):
        if row in hidden_rows:
            continue
        for col in range(min_col, max_col + 1):
            formula_value = ws_formulas.cell(row, col).value
            cached_value = ws_values.cell(row, col).value
            formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
            value = cached_value if formula else formula_value
            cells.append(
                {
                    "row": row,
                    "col": col,
                    "column": get_column_letter(col),
                    "address": f"{get_column_letter(col)}{row}",
                    "value": value.isoformat() if isinstance(value, datetime) else value,
                    "formula": formula,
                    "isFormula": bool(formula),
                }
            )

    merges = []
    for merged_range in ws_formulas.merged_cells.ranges:
        if merged_range.min_row <= max_row and merged_range.max_row >= 1 and merged_range.min_col <= max_col and merged_range.max_col >= 1:
            merges.append(
                {
                    "range": str(merged_range),
                    "row": merged_range.min_row,
                    "col": merged_range.min_col,
                    "rowspan": merged_range.max_row - merged_range.min_row + 1,
                    "colspan": merged_range.max_col - merged_range.min_col + 1,
                }
            )

    products = []
    for row in product_rows:
        product_name = clean(ws_values.cell(row, 2).value)
        if product_name and not is_excluded_product_name(product_name):
            products.append({"rowIndex": row, "productName": product_name})

    return {
        "source": HARVEST_FILE.name,
        "sheetName": sheet_name,
        "range": "B3:P48",
        "minRow": min_row,
        "maxRow": max_row,
        "minCol": min_col,
        "maxCol": max_col,
        "cells": cells,
        "merges": merges,
        "products": products,
        "hiddenRows": sorted(hidden_rows),
        "editableRows": [[5, 40], [42, 48]],
        "editableColumns": ["C:G", "I:L", "N:P"],
    }


def aggregate(rows: list[dict[str, Any]], key_fields: list[str], value_field: str) -> list[dict[str, Any]]:
    acc: dict[tuple[Any, ...], float] = defaultdict(float)
    for row in rows:
        key = tuple(row.get(field) for field in key_fields)
        acc[key] += num(row.get(value_field))
    return [{**dict(zip(key_fields, key)), value_field: value} for key, value in acc.items()]


def merge_sales_into_store_pages(data: dict[str, Any]) -> None:
    sales_dates = sorted({row.get("date") for row in data.get("salesPivot", []) if row.get("date")})
    if not sales_dates:
        return
    sales_by_date_store_product = {
        (row.get("date"), row.get("storeCode"), row.get("productCode")): num(row.get("salesQty"))
        for row in data.get("salesPivot", [])
    }
    discard_by_date_store_product = {
        (row.get("date"), row.get("storeCode"), row.get("productCode")): num(row.get("discardQty"))
        for row in data.get("salesPivot", [])
    }

    for store in data.get("storePages", []):
        store_code = store.get("storeCode")
        for sales_date in sales_dates:
            if sales_date not in store["dates"]:
                store["dates"].append(sales_date)
        sorted_dates = sorted(store["dates"])
        old_positions = {date: idx for idx, date in enumerate(store["dates"])}
        store["dates"] = sorted_dates
        for product in store.get("products", []):
            old_sales = product["sales"]
            old_discards = product.get("discards") or [0 for _ in old_sales]
            product["sales"] = [
                old_sales[old_positions[date]]
                if date in old_positions and old_positions[date] < len(old_sales)
                else 0
                for date in sorted_dates
            ]
            product["discards"] = [
                old_discards[old_positions[date]]
                if date in old_positions and old_positions[date] < len(old_discards)
                else 0
                for date in sorted_dates
            ]
            for sales_date in sales_dates:
                date_idx = store["dates"].index(sales_date)
                key = (sales_date, store_code, product.get("productCode"))
                if key in sales_by_date_store_product:
                    old_value = num(product["sales"][date_idx])
                    new_value = sales_by_date_store_product[key]
                    old_discard = num(product["discards"][date_idx])
                    new_discard = discard_by_date_store_product.get(key, 0)
                    product["sales"][date_idx] = new_value
                    product["discards"][date_idx] = new_discard
                    product["monthlySales"] = num(product.get("monthlySales")) - old_value + new_value
                    product["discardQty"] = num(product.get("discardQty")) - old_discard + new_discard
                    product["currentStock"] = (
                        num(product.get("currentStock"))
                        - (new_value - old_value)
                        - (new_discard - old_discard)
                    )


def main() -> None:
    if IMPORT_SALES_ONLY:
        import_sales_only()
        return

    sales = parse_all_sales_imports()
    inventory = parse_inventory()
    sales_date = sales["date"] or latest_inventory_date(inventory)
    data = {
        "generatedAt": taipei_now_text(),
        "salesDate": sales_date,
        "sources": sales.get("sourceFiles", []) + [p.name for _, p in INVENTORY_FILES] + [HARVEST_FILE.name, AVAILABILITY_FILE.name],
        "salesRows": sales["rows"],
        "salesPivot": sales["pivot"],
        "salesTotals": sales["totals"],
        "stores": inventory["stores"],
        "inventory": inventory["inventory"],
        "mailSales": inventory["mailSales"],
        "mailDiscards": inventory["mailDiscards"],
        "storePages": inventory["storePages"],
        "availability": parse_availability(),
        "harvestOrders": parse_harvest_orders(),
        "harvestPlanningTemplate": parse_harvest_planning_template(),
    }
    merge_sales_into_store_pages(data)

    product_names = {}
    for row in data["inventory"] + data["salesPivot"] + data["harvestOrders"]:
        code = row.get("productCode")
        name = row.get("productName")
        if code and name and not is_excluded_product_name(name):
            product_names.setdefault(code, name)
    data["products"] = [{"productCode": code, "productName": name} for code, name in sorted(product_names.items())]
    data["lowStock"] = [row for row in data["inventory"] if row["currentStock"] <= 3]
    data["salesByProduct"] = aggregate(data["salesPivot"], ["productCode", "productName"], "salesQty")
    data["discardByProduct"] = aggregate(data["salesPivot"], ["productCode", "productName"], "discardQty")

    DOCS.mkdir(parents=True, exist_ok=True)
    DATA_JSON.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    (DOCS / "data.js").write_text("window.APP_DATA=" + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(json.dumps({
        "salesRows": len(data["salesRows"]),
        "stores": len(data["stores"]),
        "inventoryRows": len(data["inventory"]),
        "lowStockRows": len(data["lowStock"]),
        "harvestOrders": len(data["harvestOrders"]),
        "harvestPlanningCells": len(data["harvestPlanningTemplate"]["cells"]),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
