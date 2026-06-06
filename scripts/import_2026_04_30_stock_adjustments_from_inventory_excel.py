from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from database import connect, get_current_snapshot, init_db, replace_stock_adjustments_for_date


ROOT = Path(__file__).resolve().parents[1]
TARGET_DATE = "2026-04-30"
EXCEL_FILES = {
    "A": ROOT / "115年04月庫存(A) 30.xlsx",
    "B": ROOT / "115年04月庫存(B) 30.xlsx",
}
SOURCE_RANGE = "BC6:BC48"


def product_code_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def adjustment_quantity(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if float(value) else None

    text = str(value).strip()
    if not text:
        return None
    is_negative = text.startswith("-") or (text.startswith("(") and text.endswith(")"))
    cleaned = (
        text.replace(",", "")
        .replace("+", "")
        .replace("-", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )
    if not cleaned:
        return None
    try:
        quantity = float(cleaned)
    except ValueError:
        return None
    if is_negative:
        quantity = -quantity
    return quantity if quantity else None


def store_lookup(snapshot: dict) -> dict[tuple[str, str], dict]:
    lookup: dict[tuple[str, str], dict] = {}
    for store in snapshot.get("storePages") or []:
        group = str(store.get("group") or "").strip().upper()
        sheet = str(store.get("sheet") or "").strip()
        if group and sheet:
            lookup[(group, sheet)] = store
    return lookup


def entries_from_sheet(ws) -> list[dict[str, object]]:
    entries: list[dict[str, object]] = []
    for row in range(6, 49):
        product_code = product_code_text(ws[f"A{row}"].value)
        quantity = adjustment_quantity(ws[f"BC{row}"].value)
        if not product_code or quantity is None:
            continue
        entries.append({"productCode": product_code, "quantity": quantity})
    return entries


def main() -> None:
    conn = connect()
    try:
        init_db(conn)
        snapshot = get_current_snapshot(conn)
        if not snapshot:
            raise RuntimeError("找不到目前資料快照，無法配對門市。")
        stores = store_lookup(snapshot)

        matched_stores = 0
        written_entries = 0
        unmatched_sheets: list[str] = []
        matched_store_keys: set[str] = set()

        for group, path in EXCEL_FILES.items():
            if not path.exists():
                raise FileNotFoundError(path)
            workbook = load_workbook(path, data_only=True, read_only=False)
            for sheet_name in workbook.sheetnames:
                normalized_sheet_name = sheet_name.strip()
                store = stores.get((group, normalized_sheet_name))
                if not store:
                    if normalized_sheet_name not in {"mail", "總銷量", "工作表1"}:
                        unmatched_sheets.append(f"{group}:{sheet_name}")
                    continue
                entries = entries_from_sheet(workbook[sheet_name])
                replace_stock_adjustments_for_date(conn, store["key"], TARGET_DATE, entries)
                matched_store_keys.add(str(store["key"]))
                matched_stores += 1
                written_entries += len(entries)

        missing_stores = [
            f"{store.get('group')}:{store.get('sheet')}"
            for store in snapshot.get("storePages") or []
            if str(store.get("group") or "").strip().upper() in EXCEL_FILES
            and str(store.get("key")) not in matched_store_keys
        ]
    finally:
        conn.close()

    print(f"來源範圍：{SOURCE_RANGE}")
    print(f"寫入日期：{TARGET_DATE}")
    print(f"已配對門市：{matched_stores}")
    print(f"寫入非零調整庫存：{written_entries}")
    print(f"未配對 Excel sheet：{len(unmatched_sheets)}")
    if unmatched_sheets:
        print("未配對 Excel sheet 範例：" + ", ".join(unmatched_sheets[:10]))
    print(f"系統中未在 Excel 找到的 A/B 門市：{len(missing_stores)}")
    if missing_stores:
        print("未找到門市範例：" + ", ".join(missing_stores[:10]))


if __name__ == "__main__":
    main()
