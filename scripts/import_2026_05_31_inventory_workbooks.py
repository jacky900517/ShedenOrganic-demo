from __future__ import annotations

import argparse
import gzip
import json
import os
import re
import subprocess
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook
from psycopg.conninfo import conninfo_to_dict
from psycopg.rows import dict_row

import psycopg


ROOT = Path(__file__).resolve().parents[1]
TAIPEI_TZ = timezone(timedelta(hours=8))
TARGET_DATE = "2026-05-31"
TARGET_MONTH = TARGET_DATE[:7]
MONTH_START = f"{TARGET_MONTH}-01"
EXCEL_FILES = {
    "A": ROOT / "115年05月庫存(A) 31.xlsx",
    "B": ROOT / "115年05月庫存(B) 31.xlsx",
}
EXCLUDED_PRODUCT_NAMES = {"油菜", "千寶菜", "紅卷", "紅捲"}
INVENTORY_SYNC_VERSION_KEY = "inventorySyncVersion"
GCLOUD_PYTHON = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "python"
    / "bin"
    / "python3"
)


def taipei_now() -> datetime:
    return datetime.now(TAIPEI_TZ)


def taipei_now_text() -> str:
    return taipei_now().isoformat(timespec="seconds")


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def product_code_text(value: Any) -> str:
    text = clean(value)
    return text or ""


def normalize_store_name(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    return re.sub(r"\s+", "", text)


def normalize_store_code(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    digits = re.sub(r"\D+", "", text)
    return digits or text


def formula_display_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    match = re.fullmatch(r'=HYPERLINK\([^,]+,\s*"([^"]+)"\)', value)
    if match:
        return match.group(1)
    return value


def num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    is_negative = text.startswith("-") or (text.startswith("(") and text.endswith(")"))
    cleaned = (
        text.replace(",", "")
        .replace("+", "")
        .replace("-", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )
    if not cleaned:
        return 0.0
    try:
        value_float = float(cleaned)
    except ValueError:
        return 0.0
    return -value_float if is_negative else value_float


def excel_date_label(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)) and 30000 < value < 70000:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(value))).date().isoformat()
    text = clean(value)
    if not text:
        return None
    if "00:00:00" in text:
        return text.split(" ")[0]
    return text


def is_excluded_product_name(value: Any) -> bool:
    text = clean(value)
    return bool(text and text in EXCLUDED_PRODUCT_NAMES)


def compact_number(value: float) -> int | float:
    return int(value) if float(value).is_integer() else value


def source_file_label() -> str:
    return ";".join(path.name for path in EXCEL_FILES.values())


def parse_inventory_workbook(group: str, path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    wb_formula = load_workbook(path, read_only=True, data_only=False)
    stores: list[dict[str, Any]] = []
    inventory: list[dict[str, Any]] = []
    mail_sales: list[dict[str, Any]] = []
    mail_discards: list[dict[str, Any]] = []
    store_pages: list[dict[str, Any]] = []
    sales_rows: list[dict[str, Any]] = []
    shipment_entries: list[dict[str, Any]] = []
    stock_adjustment_entries: list[dict[str, Any]] = []
    validation_warnings: list[str] = []
    mail_stores: list[dict[str, Any]] = []
    mail_code_by_name: dict[str, str] = {}
    mail_order_by_name: dict[str, int] = {}

    if "mail" in wb.sheetnames:
        ws = wb["mail"]
        ws_formula = wb_formula["mail"]
        mail_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
        formula_rows = list(
            ws_formula.iter_rows(min_row=1, max_row=ws_formula.max_row, max_col=ws_formula.max_column, values_only=True)
        )
        store_codes = [normalize_store_code(v) for v in mail_rows[1][2:]]
        store_name_values = [formula_display_text(v) for v in formula_rows[2][2:]]
        store_names = [normalize_store_name(v) for v in store_name_values]
        mail_stores = [
            {"mailOrder": idx, "storeCode": code, "storeName": name}
            for idx, (code, name) in enumerate(zip(store_codes, store_names))
            if code or name
        ]
        mail_code_by_name = {name: code for name, code in zip(store_names, store_codes) if name and code}
        mail_order_by_name = {name: idx for idx, name in enumerate(store_names) if name}

        section = "sales"
        for row in mail_rows:
            label = clean(row[1] if len(row) > 1 else None)
            if label and "丟棄" in label:
                section = "discard"
            product_code = label
            if not product_code or not product_code.isdigit():
                continue
            for idx, value in enumerate(row[2:]):
                qty = num(value)
                if qty == 0:
                    continue
                item = {
                    "group": group,
                    "productCode": product_code,
                    "storeCode": normalize_store_code(store_codes[idx]),
                    "storeName": normalize_store_name(store_names[idx]),
                    "quantity": compact_number(qty),
                }
                if section == "discard":
                    mail_discards.append(item)
                else:
                    mail_sales.append(item)

    for sheet_name in wb.sheetnames:
        if sheet_name in {"mail", "總銷量", "工作表1", "工作表2"}:
            continue
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), max_col=min(ws.max_column, 76), values_only=True))

        def get(row: int, col: int) -> Any:
            if len(sheet_rows) < row or len(sheet_rows[row - 1]) < col:
                return None
            return sheet_rows[row - 1][col - 1]

        sheet_store_name = normalize_store_name(get(5, 2)) or normalize_store_name(sheet_name)
        mail_index_raw = get(3, 1)
        mail_index = int(mail_index_raw) - 1 if isinstance(mail_index_raw, (int, float)) and int(mail_index_raw) >= 1 else None
        mail_store = mail_stores[mail_index] if mail_index is not None and mail_index < len(mail_stores) else None
        matched_store_name = sheet_store_name if mail_code_by_name.get(sheet_store_name) else normalize_store_name(sheet_name)
        name_match = mail_code_by_name.get(matched_store_name)
        if name_match:
            store_name = sheet_store_name
            store_code = name_match
            store_order = mail_order_by_name.get(matched_store_name, len(stores))
        elif mail_store:
            store_name = normalize_store_name(mail_store.get("storeName")) or sheet_store_name
            store_code = normalize_store_code(mail_store.get("storeCode"))
            store_order = mail_store.get("mailOrder", len(stores))
        else:
            store_name = sheet_store_name
            store_code = None
            store_order = len(stores)
        route = clean(get(4, 2))
        store_key = f"{group}:{sheet_name}"
        stores.append(
            {
                "group": group,
                "sheet": sheet_name,
                "storeName": store_name,
                "storeCode": store_code,
                "route": route,
                "mailOrder": store_order,
            }
        )

        headers = [clean(v) for v in (sheet_rows[3] if len(sheet_rows) > 3 else [])]
        dates = [excel_date_label(v) for v in (sheet_rows[4] if len(sheet_rows) > 4 else [])]
        sale_cols = [idx for idx, header in enumerate(headers) if header == "出售" and dates[idx]]
        if not sale_cols:
            validation_warnings.append(f"{store_key}: 找不到每日出售欄")
            continue
        first_sale_col = min(sale_cols)
        month_dates = [str(dates[idx]) for idx in sale_cols if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(dates[idx]))]
        inventory_month = next((date[:7] for date in month_dates if date.startswith(TARGET_MONTH)), TARGET_MONTH)
        if inventory_month != TARGET_MONTH:
            validation_warnings.append(f"{store_key}: 出售日期月份不是 {TARGET_MONTH}")

        all_dated_shipment_cols = [
            idx
            for idx, header in enumerate(headers)
            if 3 <= idx < first_sale_col
            and re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(dates[idx] or ""))
        ]
        shipment_cols = [
            idx
            for idx in all_dated_shipment_cols
            if str(dates[idx]).startswith(TARGET_MONTH)
        ]
        sale_total_col = next((idx for idx, value in enumerate(dates) if clean(value) == "出售\n小計"), None)
        discard_col = 53  # BB, zero-based
        adjustment_col = 54  # BC, zero-based
        stock_col = 55  # BD, zero-based

        store_products: list[dict[str, Any]] = []
        for row_number, row in enumerate(sheet_rows[5:], start=6):
            product_code = product_code_text(row[0] if len(row) > 0 else None)
            product_name = clean(row[1] if len(row) > 1 else None)
            if not product_code or not product_name or is_excluded_product_name(product_name):
                continue

            base_stock = num(row[2] if len(row) > 2 else None)
            sale_values = [num(row[idx] if len(row) > idx else None) for idx in sale_cols]
            monthly_sales = sum(sale_values)
            # Excel BB is a stock movement formula (報廢小計 * -1). The app stores
            # discard_qty as a positive quantity and subtracts it in stock math.
            discard_qty = abs(num(row[discard_col] if len(row) > discard_col else None))
            adjustment_qty = num(row[adjustment_col] if len(row) > adjustment_col else None)
            shipment_events: list[dict[str, Any]] = []
            shipment_total = 0.0
            non_target_month_shipment_total = 0.0
            shipment_dates: list[str] = []
            for idx in all_dated_shipment_cols:
                if idx in shipment_cols:
                    continue
                non_target_month_shipment_total += num(row[idx] if len(row) > idx else None)
            for idx in shipment_cols:
                value = num(row[idx] if len(row) > idx else None)
                if not value:
                    continue
                shipment_date = str(dates[idx])
                shipment_total += value
                shipment_dates.append(shipment_date)
                event = {
                    "storeKey": store_key,
                    "shipmentDate": shipment_date,
                    "productCode": product_code,
                    "quantity": compact_number(value),
                }
                shipment_entries.append(event)
                shipment_events.append({"date": shipment_date, "quantity": compact_number(value)})

            discards = [0.0 for _ in sale_cols]
            if TARGET_DATE in month_dates:
                discards[month_dates.index(TARGET_DATE)] = discard_qty
            for date_value, sales_qty, discard_value in zip(month_dates, sale_values, discards):
                if sales_qty or discard_value:
                    sales_rows.append(
                        {
                            "date": date_value,
                            "storeCode": store_code,
                            "storeName": store_name,
                            "productCode": product_code,
                            "productName": product_name,
                            "salesQty": compact_number(sales_qty),
                            "discardQty": compact_number(discard_value),
                            "endingStock": 0,
                        }
                    )

            stock_adjustment_entries.append(
                {
                    "storeKey": store_key,
                    "shipmentDate": TARGET_DATE,
                    "productCode": product_code,
                    "quantity": compact_number(adjustment_qty),
                }
            )
            current_stock = base_stock + shipment_total - monthly_sales - discard_qty + adjustment_qty
            workbook_stock = num(row[stock_col] if len(row) > stock_col else None)
            workbook_comparable_stock = current_stock + non_target_month_shipment_total
            if abs(workbook_comparable_stock - workbook_stock) > 0.0001:
                validation_warnings.append(
                    f"{store_key} row {row_number} {product_code}: 庫存計算 {workbook_comparable_stock} != Excel BD {workbook_stock}"
                )
            if sale_total_col is not None:
                workbook_sale_total = num(row[sale_total_col] if len(row) > sale_total_col else None)
                if abs(monthly_sales - workbook_sale_total) > 0.0001:
                    validation_warnings.append(
                        f"{store_key} row {row_number} {product_code}: 出售小計 {monthly_sales} != Excel BA {workbook_sale_total}"
                    )

            store_products.append(
                {
                    "productCode": product_code,
                    "productName": product_name,
                    "openingStocksByMonth": {TARGET_MONTH: compact_number(base_stock)},
                    "sales": [compact_number(value) for value in sale_values],
                    "discards": [compact_number(value) for value in discards],
                    "monthlySales": compact_number(monthly_sales),
                    "discardQty": compact_number(discard_qty),
                    "adjustmentQty": compact_number(adjustment_qty),
                    "existingShipmentQty": compact_number(shipment_total),
                    "shipments": shipment_events,
                    "shipmentInputQty": 0,
                    "currentStock": compact_number(current_stock),
                }
            )
            inventory.append(
                {
                    "group": group,
                    "storeSheet": sheet_name,
                    "storeName": store_name,
                    "storeCode": store_code,
                    "route": route,
                    "mailOrder": store_order,
                    "productCode": product_code,
                    "productName": product_name,
                    "baseStock": compact_number(base_stock),
                    "salesQty": compact_number(monthly_sales),
                    "discardQty": compact_number(discard_qty),
                    "shipmentQty": compact_number(shipment_total),
                    "adjustmentQty": compact_number(adjustment_qty),
                    "currentStock": compact_number(current_stock),
                    "shipmentDates": shipment_dates,
                }
            )
        store_pages.append(
            {
                "key": store_key,
                "group": group,
                "sheet": sheet_name,
                "storeName": store_name,
                "storeCode": store_code,
                "route": route,
                "mailOrder": store_order,
                "dates": month_dates,
                "products": store_products,
            }
        )

    return {
        "stores": stores,
        "inventory": inventory,
        "mailSales": mail_sales,
        "mailDiscards": mail_discards,
        "storePages": store_pages,
        "salesRows": sales_rows,
        "shipmentEntries": shipment_entries,
        "stockAdjustmentEntries": stock_adjustment_entries,
        "validationWarnings": validation_warnings,
    }


def parse_inventory_files() -> dict[str, Any]:
    combined: dict[str, Any] = {
        "stores": [],
        "inventory": [],
        "mailSales": [],
        "mailDiscards": [],
        "storePages": [],
        "salesRows": [],
        "shipmentEntries": [],
        "stockAdjustmentEntries": [],
        "validationWarnings": [],
    }
    for group, path in EXCEL_FILES.items():
        if not path.exists():
            raise FileNotFoundError(path)
        parsed = parse_inventory_workbook(group, path)
        for key in combined:
            combined[key].extend(parsed[key])
    combined["shipmentEntries"] = coalesce_quantity_entries(
        combined["shipmentEntries"],
        ["storeKey", "shipmentDate", "productCode"],
    )
    combined["stockAdjustmentEntries"] = coalesce_quantity_entries(
        combined["stockAdjustmentEntries"],
        ["storeKey", "shipmentDate", "productCode"],
    )
    return combined


def coalesce_quantity_entries(
    entries: list[dict[str, Any]],
    key_fields: list[str],
    keep_zero: bool = False,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], dict[str, Any]] = {}
    for entry in entries:
        key = tuple(str(entry.get(field) or "") for field in key_fields)
        if key not in grouped:
            grouped[key] = {field: entry.get(field) for field in key_fields}
            grouped[key]["quantity"] = 0.0
        grouped[key]["quantity"] += num(entry.get("quantity"))
    return [
        {**entry, "quantity": compact_number(num(entry.get("quantity")))}
        for entry in grouped.values()
        if keep_zero or num(entry.get("quantity"))
    ]


def aggregate(rows: list[dict[str, Any]], key_fields: list[str], value_field: str) -> list[dict[str, Any]]:
    acc: dict[tuple[Any, ...], float] = defaultdict(float)
    for row in rows:
        key = tuple(row.get(field) for field in key_fields)
        acc[key] += num(row.get(value_field))
    return [{**dict(zip(key_fields, key)), value_field: compact_number(value)} for key, value in acc.items()]


def build_sales_pivot(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, float]]:
    pivot: dict[tuple[str, str, str], dict[str, Any]] = {}
    totals = {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0}
    for row in rows:
        key = (str(row["date"]), str(row["storeCode"]), str(row["productCode"]))
        if key not in pivot:
            pivot[key] = {
                "date": row["date"],
                "storeCode": row["storeCode"],
                "storeName": row["storeName"],
                "productCode": row["productCode"],
                "productName": row["productName"],
                "salesQty": 0.0,
                "discardQty": 0.0,
                "endingStock": 0.0,
            }
        pivot[key]["salesQty"] += num(row.get("salesQty"))
        pivot[key]["discardQty"] += num(row.get("discardQty"))
        pivot[key]["endingStock"] += num(row.get("endingStock"))
        totals["salesQty"] += num(row.get("salesQty"))
        totals["discardQty"] += num(row.get("discardQty"))
        totals["endingStock"] += num(row.get("endingStock"))
    for row in pivot.values():
        for field in ("salesQty", "discardQty", "endingStock"):
            row[field] = compact_number(row[field])
    return list(pivot.values()), {key: compact_number(value) for key, value in totals.items()}


def product_list(parsed: dict[str, Any]) -> list[dict[str, str]]:
    products: dict[str, str] = {}
    for row in parsed["inventory"]:
        code = str(row.get("productCode") or "")
        name = str(row.get("productName") or "")
        if code and name:
            products.setdefault(code, name)
    return [{"productCode": code, "productName": name} for code, name in sorted(products.items())]


def month_sources(old_sources: list[Any]) -> list[str]:
    keep = [
        str(source)
        for source in old_sources
        if source
        and "SalesData_" not in str(source)
        and "115年04月庫存" not in str(source)
        and "115年05月庫存" not in str(source)
    ]
    return [path.name for path in EXCEL_FILES.values()] + keep


def build_next_snapshot(current: dict[str, Any], parsed: dict[str, Any]) -> dict[str, Any]:
    sales_pivot, sales_totals = build_sales_pivot(parsed["salesRows"])
    next_payload = dict(current)
    next_payload.update(
        {
            "generatedAt": taipei_now_text(),
            "salesDate": TARGET_DATE,
            "sources": month_sources(list(current.get("sources") or [])),
            "salesRows": [],
            "salesPivot": [],
            "salesTotals": {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0},
            "stores": parsed["stores"],
            "inventory": parsed["inventory"],
            "mailSales": [],
            "mailDiscards": [],
            "storePages": parsed["storePages"],
            "products": product_list(parsed),
            "lowStock": [],
            "salesByProduct": aggregate(sales_pivot, ["productCode", "productName"], "salesQty"),
            "discardByProduct": aggregate(sales_pivot, ["productCode", "productName"], "discardQty"),
            "importSummary": {
                "source": source_file_label(),
                "importedAt": taipei_now_text(),
                "salesTotals": sales_totals,
                "targetDate": TARGET_DATE,
            },
        }
    )
    compact_bootstrap_snapshot(next_payload)
    return next_payload


def compact_bootstrap_snapshot(payload: dict[str, Any]) -> None:
    """Drop values that are recomputed or loaded by API after bootstrap."""
    payload["lowStock"] = []
    payload["mailSales"] = []
    payload["mailDiscards"] = []
    for row in payload.get("inventory") or []:
        row.pop("shipmentDates", None)
    for store in payload.get("storePages") or []:
        for product in store.get("products") or []:
            for key in (
                "sales",
                "discards",
                "monthlySales",
                "discardQty",
                "adjustmentQty",
                "shipments",
                "shipmentInputQty",
                "currentStock",
            ):
                product.pop(key, None)


def gcloud_secret_value(secret_name: str, project: str) -> str:
    env = os.environ.copy()
    if GCLOUD_PYTHON.exists():
        env["CLOUDSDK_PYTHON"] = str(GCLOUD_PYTHON)
    return subprocess.check_output(
        ["gcloud", "secrets", "versions", "access", "latest", f"--secret={secret_name}", f"--project={project}"],
        text=True,
        env=env,
    ).strip()


def connection_kwargs(args: argparse.Namespace) -> dict[str, Any]:
    if args.use_gcp_secret:
        database_url = gcloud_secret_value(args.secret_name, args.project)
        info = conninfo_to_dict(database_url)
        return {
            "host": args.proxy_host,
            "port": args.proxy_port,
            "dbname": info.get("dbname") or "sheden_organic",
            "user": info.get("user"),
            "password": info.get("password"),
            "row_factory": dict_row,
        }
    database_url = os.environ.get("DATABASE_URL", "").strip()
    if not database_url:
        raise ValueError("DATABASE_URL is required unless --use-gcp-secret is passed")
    return {"conninfo": database_url, "row_factory": dict_row}


def connect_database(args: argparse.Namespace) -> psycopg.Connection:
    kwargs = connection_kwargs(args)
    conninfo = kwargs.pop("conninfo", None)
    if conninfo:
        return psycopg.connect(conninfo, **kwargs)
    return psycopg.connect(**kwargs)


def dump_rows(cur: psycopg.Cursor, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cur.execute(sql, params)
    return [dict(row) for row in cur.fetchall()]


def backup_production_state(conn: psycopg.Connection, parsed: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    backup_path = output_dir / f"production_inventory_import_backup_{taipei_now().strftime('%Y%m%d_%H%M%S')}.json.gz"
    store_keys = sorted({store["key"] for store in parsed["storePages"]})
    with conn.cursor() as cur:
        backup = {
            "createdAt": taipei_now_text(),
            "targetMonth": TARGET_MONTH,
            "targetDate": TARGET_DATE,
            "sourceFiles": [path.name for path in EXCEL_FILES.values()],
            "currentSnapshot": dump_rows(
                cur,
                """
                SELECT id, sales_date, generated_at, is_current, payload_json
                FROM app_snapshots
                WHERE is_current = 1
                ORDER BY id DESC
                LIMIT 1
                """,
            ),
            "salesImports": dump_rows(
                cur,
                "SELECT * FROM sales_imports WHERE sales_date >= %s AND sales_date <= %s ORDER BY sales_date",
                (MONTH_START, TARGET_DATE),
            ),
            "salesRows": dump_rows(
                cur,
                "SELECT * FROM sales_rows WHERE sales_date >= %s AND sales_date <= %s ORDER BY sales_date, id",
                (MONTH_START, TARGET_DATE),
            ),
            "shipmentEntries": dump_rows(
                cur,
                """
                SELECT * FROM shipment_entries
                WHERE shipment_date >= %s AND shipment_date <= %s AND store_key = ANY(%s)
                ORDER BY store_key, shipment_date, product_code
                """,
                (MONTH_START, TARGET_DATE, store_keys),
            ),
            "stockAdjustmentEntries": dump_rows(
                cur,
                """
                SELECT * FROM stock_adjustment_entries
                WHERE shipment_date >= %s AND shipment_date <= %s AND store_key = ANY(%s)
                ORDER BY store_key, shipment_date, product_code
                """,
                (MONTH_START, TARGET_DATE, store_keys),
            ),
            "inventorySyncVersion": dump_rows(
                cur,
                "SELECT * FROM app_settings WHERE setting_key = %s",
                (INVENTORY_SYNC_VERSION_KEY,),
            ),
        }
    with gzip.open(backup_path, "wt", encoding="utf-8") as handle:
        json.dump(backup, handle, ensure_ascii=False, separators=(",", ":"))
    return backup_path


def upsert_inventory_sync_version(cur: psycopg.Cursor, version: str) -> None:
    cur.execute(
        """
        INSERT INTO app_settings (setting_key, setting_value, updated_at)
        VALUES (%s, %s, %s)
        ON CONFLICT(setting_key) DO UPDATE SET
            setting_value = excluded.setting_value,
            updated_at = excluded.updated_at
        """,
        (INVENTORY_SYNC_VERSION_KEY, json.dumps(version, ensure_ascii=False, separators=(",", ":")), version),
    )


def write_import(conn: psycopg.Connection, parsed: dict[str, Any], next_snapshot: dict[str, Any]) -> dict[str, Any]:
    store_keys = sorted({store["key"] for store in parsed["storePages"]})
    timestamp = taipei_now_text()
    with conn.transaction():
        with conn.cursor() as cur:
            cur.execute("DELETE FROM sales_rows WHERE sales_date >= %s AND sales_date <= %s", (MONTH_START, TARGET_DATE))
            cur.execute("DELETE FROM sales_imports WHERE sales_date >= %s AND sales_date <= %s", (MONTH_START, TARGET_DATE))
            cur.executemany(
                """
                INSERT INTO sales_rows (
                    sales_date,
                    store_code,
                    store_name,
                    product_code,
                    product_name,
                    sales_qty,
                    discard_qty,
                    ending_stock
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        row["date"],
                        str(row.get("storeCode") or ""),
                        row.get("storeName"),
                        str(row.get("productCode") or ""),
                        row.get("productName"),
                        float(row.get("salesQty") or 0),
                        float(row.get("discardQty") or 0),
                        float(row.get("endingStock") or 0),
                    )
                    for row in parsed["salesRows"]
                    if row.get("storeCode") and row.get("productCode")
                ],
            )
            cur.executemany(
                """
                INSERT INTO sales_imports (sales_date, file_name, file_path, imported_at)
                VALUES (%s, %s, %s, %s)
                """,
                [(f"{TARGET_MONTH}-{day:02d}", source_file_label(), "", timestamp) for day in range(1, 32)],
            )
            cur.execute(
                """
                DELETE FROM shipment_entries
                WHERE shipment_date >= %s AND shipment_date <= %s AND store_key = ANY(%s)
                """,
                (MONTH_START, TARGET_DATE, store_keys),
            )
            cur.executemany(
                """
                INSERT INTO shipment_entries (store_key, shipment_date, product_code, quantity, updated_at)
                VALUES (%s, %s, %s, %s, %s)
                """,
                [
                    (
                        entry["storeKey"],
                        entry["shipmentDate"],
                        str(entry["productCode"]),
                        float(entry["quantity"]),
                        timestamp,
                    )
                    for entry in parsed["shipmentEntries"]
                    if float(entry["quantity"])
                ],
            )
            cur.execute(
                """
                DELETE FROM stock_adjustment_entries
                WHERE shipment_date >= %s AND shipment_date <= %s AND store_key = ANY(%s)
                """,
                (MONTH_START, TARGET_DATE, store_keys),
            )
            cur.executemany(
                """
                INSERT INTO stock_adjustment_entries (store_key, shipment_date, product_code, quantity, updated_at)
                VALUES (%s, %s, %s, %s, %s)
                """,
                [
                    (
                        entry["storeKey"],
                        entry["shipmentDate"],
                        str(entry["productCode"]),
                        float(entry["quantity"]),
                        timestamp,
                    )
                    for entry in parsed["stockAdjustmentEntries"]
                ],
            )
            cur.execute("UPDATE app_snapshots SET is_current = 0")
            cur.execute(
                """
                INSERT INTO app_snapshots (sales_date, generated_at, payload_json, is_current)
                VALUES (%s, %s, %s, 1)
                """,
                (
                    TARGET_DATE,
                    next_snapshot["generatedAt"],
                    json.dumps(next_snapshot, ensure_ascii=False, separators=(",", ":")),
                ),
            )
            upsert_inventory_sync_version(cur, timestamp)
    return {
        "storeKeys": len(store_keys),
        "salesRows": len(parsed["salesRows"]),
        "shipmentEntries": len(parsed["shipmentEntries"]),
        "stockAdjustmentEntries": len(parsed["stockAdjustmentEntries"]),
        "snapshotStores": len(parsed["storePages"]),
    }


def production_summary(conn: psycopg.Connection) -> dict[str, Any]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT sales_date, COUNT(*) AS row_count, SUM(sales_qty) AS sales_qty, SUM(discard_qty) AS discard_qty
            FROM sales_rows
            WHERE sales_date >= %s AND sales_date <= %s
            GROUP BY sales_date
            ORDER BY sales_date
            """,
            (MONTH_START, TARGET_DATE),
        )
        sales_by_date = [dict(row) for row in cur.fetchall()]
        cur.execute(
            """
            SELECT sales_date, generated_at, LENGTH(payload_json) AS payload_bytes
            FROM app_snapshots
            WHERE is_current = 1
            ORDER BY id DESC
            LIMIT 1
            """
        )
        snapshot = dict(cur.fetchone() or {})
        cur.execute(
            """
            SELECT COUNT(*) AS count
            FROM shipment_entries
            WHERE shipment_date >= %s AND shipment_date <= %s
            """,
            (MONTH_START, TARGET_DATE),
        )
        shipment_count = cur.fetchone()["count"]
        cur.execute(
            """
            SELECT COUNT(*) AS count
            FROM stock_adjustment_entries
            WHERE shipment_date = %s
            """,
            (TARGET_DATE,),
        )
        adjustment_count = cur.fetchone()["count"]
    return {
        "currentSnapshot": snapshot,
        "salesByDate": sales_by_date,
        "shipmentEntriesInMonth": shipment_count,
        "stockAdjustmentEntriesOnTargetDate": adjustment_count,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import 2026-05-31 inventory workbooks into the production database.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only; do not write the database.")
    parser.add_argument("--use-gcp-secret", action="store_true", help="Read DATABASE_URL from Secret Manager.")
    parser.add_argument("--secret-name", default="sheden-database-url")
    parser.add_argument("--project", default=os.environ.get("GCP_PROJECT", ""))
    parser.add_argument("--proxy-host", default="127.0.0.1")
    parser.add_argument("--proxy-port", type=int, default=5433)
    parser.add_argument("--backup-dir", type=Path, default=ROOT / "backups")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    parsed = parse_inventory_files()
    next_snapshot: dict[str, Any] | None = None
    summary = {
        "stores": len(parsed["stores"]),
        "storePages": len(parsed["storePages"]),
        "inventoryRows": len(parsed["inventory"]),
        "salesRowsToInsert": len(parsed["salesRows"]),
        "shipmentEntriesToInsert": len(parsed["shipmentEntries"]),
        "stockAdjustmentEntriesToInsert": len(parsed["stockAdjustmentEntries"]),
        "validationWarnings": len(parsed["validationWarnings"]),
    }
    if parsed["validationWarnings"]:
        summary["validationWarningSamples"] = parsed["validationWarnings"][:20]

    if args.dry_run:
        print(json.dumps({"ok": True, "dryRun": True, **summary}, ensure_ascii=False, indent=2))
        return

    conn = connect_database(args)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT payload_json FROM app_snapshots WHERE is_current = 1 ORDER BY id DESC LIMIT 1")
            row = cur.fetchone()
            if not row:
                raise RuntimeError("Production database has no current app snapshot.")
            current = json.loads(row["payload_json"])
        next_snapshot = build_next_snapshot(current, parsed)
        backup_path = backup_production_state(conn, parsed, args.backup_dir)
        write_summary = write_import(conn, parsed, next_snapshot)
        conn.commit()
        final_summary = production_summary(conn)
    finally:
        conn.close()

    print(
        json.dumps(
            {
                "ok": True,
                **summary,
                "backupPath": str(backup_path),
                "writeSummary": write_summary,
                "productionSummary": final_summary,
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        )
    )


if __name__ == "__main__":
    main()
