from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from database import connect, init_db, replace_field_net_houses


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "115每棟蔬菜狀況.xlsx"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def date_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value)


def numeric_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def column_index(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    return None


def cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def load_zones(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    zones: list[dict[str, Any]] = []
    for sheet_index, worksheet in enumerate(workbook.worksheets):
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {clean_text(value): index for index, value in enumerate(header_row) if clean_text(value)}
        code_index = column_index(headers, "網室編號")
        grid_index = column_index(headers, "格數")
        status_index = column_index(headers, "狀態")
        crop1_index = column_index(headers, "作物1", "作物")
        crop2_index = column_index(headers, "作物2")
        planted_index = column_index(headers, "種植日")
        harvest_index = column_index(headers, "採收日")
        note_index = column_index(headers, "備註")
        if code_index is None:
            raise ValueError(f"{worksheet.title} 缺少「網室編號」欄位")

        net_houses = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            net_house_code = clean_text(cell(row, code_index))
            if not net_house_code:
                continue
            net_houses.append(
                {
                    "netHouseCode": net_house_code,
                    "gridCount": numeric_value(cell(row, grid_index)),
                    "status": clean_text(cell(row, status_index)),
                    "crop1": clean_text(cell(row, crop1_index)),
                    "crop2": clean_text(cell(row, crop2_index)),
                    "plantedDate": date_text(cell(row, planted_index)),
                    "harvestDate": date_text(cell(row, harvest_index)),
                    "note": clean_text(cell(row, note_index)),
                    "sourceRow": row_number,
                }
            )
        zones.append({"name": worksheet.title, "sortOrder": sheet_index, "netHouses": net_houses})
    return zones


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)
    zones = load_zones(WORKBOOK_PATH)
    conn = connect()
    try:
        init_db(conn)
        replace_field_net_houses(conn, zones)
    finally:
        conn.close()
    total = sum(len(zone["netHouses"]) for zone in zones)
    print(f"Imported {total} net houses from {len(zones)} zones.")


if __name__ == "__main__":
    main()
