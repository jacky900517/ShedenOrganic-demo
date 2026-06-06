from __future__ import annotations

import argparse
import gzip
import json
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from import_2026_05_31_inventory_workbooks import (
    clean,
    compact_number,
    connect_database,
    excel_date_label,
    is_excluded_product_name,
    num,
    product_code_text,
    taipei_now,
    taipei_now_text,
)
from shift_shipment_entry_date import INVENTORY_SYNC_VERSION_KEY, mark_inventory_sync_changed


ROOT = Path(__file__).resolve().parents[1]


def parse_arrival_entries(workbook_path: Path, group: str, arrival_date: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    entries: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name in {"mail", "總銷量", "工作表1", "工作表2"}:
            continue
        sheet = workbook[sheet_name]
        rows = list(
            sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row, 80),
                max_col=min(sheet.max_column, 76),
                values_only=True,
            )
        )
        headers = [clean(value) for value in (rows[3] if len(rows) > 3 else [])]
        dates = [excel_date_label(value) for value in (rows[4] if len(rows) > 4 else [])]
        sale_cols = [idx for idx, header in enumerate(headers) if header == "出售" and dates[idx]]
        if not sale_cols:
            continue
        first_sale_col = min(sale_cols)
        arrival_cols = [
            idx
            for idx, _header in enumerate(headers)
            if 3 <= idx < first_sale_col and str(dates[idx] or "") == arrival_date
        ]
        if not arrival_cols:
            continue

        store_key = f"{group}:{sheet_name}"
        for row in rows[5:]:
            product_code = product_code_text(row[0] if len(row) > 0 else None)
            product_name = clean(row[1] if len(row) > 1 else None)
            if not product_code or not product_name or is_excluded_product_name(product_name):
                continue
            quantity = sum(num(row[idx] if len(row) > idx else None) for idx in arrival_cols)
            if not quantity:
                continue
            entries.append(
                {
                    "storeKey": store_key,
                    "shipmentDate": arrival_date,
                    "productCode": product_code,
                    "quantity": compact_number(quantity),
                }
            )
    return entries


def dump_rows(cursor, sql: str, params: tuple[Any, ...]) -> list[dict[str, Any]]:
    cursor.execute(sql, params)
    return [dict(row) for row in cursor.fetchall()]


def existing_summary(cursor, arrival_date: str, group: str) -> dict[str, Any]:
    group_prefix = f"{group}:%"
    cursor.execute(
        """
        SELECT
          COUNT(*) AS row_count,
          COUNT(DISTINCT store_key) AS store_count,
          COALESCE(SUM(quantity), 0) AS quantity_sum
        FROM shipment_entries
        WHERE shipment_date = %s AND store_key LIKE %s
        """,
        (arrival_date, group_prefix),
    )
    return dict(cursor.fetchone() or {})


def overlap_summary(cursor, entries: list[dict[str, Any]], arrival_date: str) -> dict[str, Any]:
    cursor.execute("DROP TABLE IF EXISTS tmp_workbook_arrivals")
    cursor.execute(
        """
        CREATE TEMP TABLE tmp_workbook_arrivals (
          store_key TEXT,
          product_code TEXT,
          quantity DOUBLE PRECISION
        ) ON COMMIT DROP
        """
    )
    cursor.executemany(
        """
        INSERT INTO tmp_workbook_arrivals (store_key, product_code, quantity)
        VALUES (%s, %s, %s)
        """,
        [(entry["storeKey"], entry["productCode"], float(entry["quantity"])) for entry in entries],
    )
    cursor.execute(
        """
        SELECT
          COUNT(source.*) AS overlap_count,
          COALESCE(SUM(source.quantity), 0) AS existing_quantity_sum,
          COALESCE(SUM(incoming.quantity) FILTER (WHERE source.store_key IS NOT NULL), 0) AS incoming_quantity_sum,
          COUNT(*) FILTER (
            WHERE source.store_key IS NOT NULL
              AND ABS(source.quantity - incoming.quantity) < 0.0001
          ) AS exact_quantity_matches
        FROM tmp_workbook_arrivals incoming
        LEFT JOIN shipment_entries source
          ON source.shipment_date = %s
         AND source.store_key = incoming.store_key
         AND source.product_code = incoming.product_code
        """,
        (arrival_date,),
    )
    return dict(cursor.fetchone() or {})


def backup_state(cursor, args: argparse.Namespace, entries: list[dict[str, Any]], before: dict[str, Any]) -> Path:
    args.backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = args.backup_dir / (
        f"workbook_arrivals_{args.group}_{args.arrival_date}_{taipei_now().strftime('%Y%m%d_%H%M%S')}.json.gz"
    )
    group_prefix = f"{args.group}:%"
    backup = {
        "createdAt": taipei_now_text(),
        "operation": "import_workbook_arrivals",
        "workbook": str(args.workbook),
        "group": args.group,
        "arrivalDate": args.arrival_date,
        "summaryBefore": before,
        "incomingEntries": entries,
        "existingRows": dump_rows(
            cursor,
            """
            SELECT *
            FROM shipment_entries
            WHERE shipment_date = %s AND store_key LIKE %s
            ORDER BY store_key, product_code
            """,
            (args.arrival_date, group_prefix),
        ),
        "inventorySyncVersion": dump_rows(
            cursor,
            "SELECT * FROM app_settings WHERE setting_key = %s",
            (INVENTORY_SYNC_VERSION_KEY,),
        ),
    }
    with gzip.open(backup_path, "wt", encoding="utf-8") as handle:
        json.dump(backup, handle, ensure_ascii=False, separators=(",", ":"), default=str)
    return backup_path


def run(args: argparse.Namespace) -> dict[str, Any]:
    args.workbook = args.workbook.resolve()
    entries = parse_arrival_entries(args.workbook, args.group, args.arrival_date)
    if not entries:
        raise ValueError(f"No arrivals found for {args.arrival_date} in {args.workbook}")

    incoming_summary = {
        "rowCount": len(entries),
        "storeCount": len({entry["storeKey"] for entry in entries}),
        "quantitySum": sum(float(entry["quantity"]) for entry in entries),
    }
    with connect_database(args) as conn:
        with conn.transaction():
            with conn.cursor() as cursor:
                cursor.execute("LOCK TABLE shipment_entries IN SHARE ROW EXCLUSIVE MODE")
                before = {
                    "existing": existing_summary(cursor, args.arrival_date, args.group),
                    "overlap": overlap_summary(cursor, entries, args.arrival_date),
                }
                backup_path = backup_state(cursor, args, entries, before)
                inserted_rows = 0
                sync_version = ""
                if args.apply:
                    cursor.execute(
                        "DELETE FROM shipment_entries WHERE shipment_date = %s AND store_key LIKE %s",
                        (args.arrival_date, f"{args.group}:%"),
                    )
                    sync_version = taipei_now_text()
                    cursor.executemany(
                        """
                        INSERT INTO shipment_entries (store_key, shipment_date, product_code, quantity, updated_at)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        [
                            (
                                entry["storeKey"],
                                entry["shipmentDate"],
                                entry["productCode"],
                                float(entry["quantity"]),
                                sync_version,
                            )
                            for entry in entries
                        ],
                    )
                    inserted_rows = len(entries)
                    mark_inventory_sync_changed(cursor, sync_version)
                after = {
                    "existing": existing_summary(cursor, args.arrival_date, args.group),
                    "overlap": overlap_summary(cursor, entries, args.arrival_date),
                }

    return {
        "applied": bool(args.apply),
        "workbook": str(args.workbook),
        "group": args.group,
        "arrivalDate": args.arrival_date,
        "incoming": incoming_summary,
        "backupPath": str(backup_path),
        "insertedRows": inserted_rows,
        "syncVersion": sync_version,
        "before": before,
        "after": after,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import arrival columns from an inventory workbook into shipment_entries.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--group", required=True)
    parser.add_argument("--arrival-date", required=True)
    parser.add_argument("--apply", action="store_true", help="Write the import. Omit for dry-run.")
    parser.add_argument("--use-gcp-secret", action="store_true", help="Read DATABASE_URL from Secret Manager.")
    parser.add_argument("--secret-name", default="sheden-database-url")
    parser.add_argument("--project", default=os.environ.get("GCP_PROJECT", ""))
    parser.add_argument("--proxy-host", default="127.0.0.1")
    parser.add_argument("--proxy-port", type=int, default=5433)
    parser.add_argument("--backup-dir", type=Path, default=ROOT / "backups")
    return parser.parse_args()


def main() -> None:
    print(json.dumps(run(parse_args()), ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
