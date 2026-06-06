from __future__ import annotations

import json
import hashlib
import hmac
import mimetypes
import os
import re
import secrets
import subprocess
import sys
import uuid
from http.cookies import SimpleCookie
from io import BytesIO
from datetime import datetime, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, quote, unquote, urlencode, urlparse
from urllib.request import Request, urlopen

sys.path.insert(0, str(Path(__file__).resolve().parent))

from database import (
    add_field_work_message,
    add_daily_vegetable_manual_entry,
    add_harvest_message,
    delete_field_work_message,
    delete_daily_vegetable_manual_entry,
    connect,
    delete_field_work_task_for_zones,
    delete_harvest_message,
    ensure_field_work_task_for_zones,
    get_field_bt_area_keys_for_date,
    get_daily_vegetable_manual_entries,
    get_current_snapshot,
    get_field_work_message_photo_file_names,
    get_field_bt_records,
    get_field_net_house_data,
    get_field_work_messages,
    get_field_work_records,
    get_harvest_entries,
    get_harvest_messages,
    get_net_house_status_records,
    get_setting,
    get_shipment_entries,
    get_stock_adjustment_entries,
    get_store_notes,
    get_store_routes,
    init_db,
    is_field_work_blocked_by_planting,
    replace_harvest_entries_for_date,
    replace_field_bt_records,
    replace_field_work_records,
    replace_shipments_for_date,
    replace_stock_adjustments_for_date,
    save_net_house_status_record,
    save_store_note,
    save_store_route,
    save_setting,
    set_current_snapshot,
    sync_net_house_status_from_field_work,
    taipei_now,
    taipei_now_text,
    update_daily_vegetable_manual_entry,
)
from storage import storage_backend


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"


def runtime_app_version() -> str:
    configured_version = os.environ.get("APP_VERSION") or os.environ.get("K_REVISION") or os.environ.get("GAE_VERSION")
    if configured_version:
        return configured_version
    local_version_files = [
        DOCS / "index.html",
        DOCS / "app.js",
        DOCS / "styles.css",
        DOCS / "sw.js",
    ]
    latest_mtime = max((path.stat().st_mtime_ns for path in local_version_files if path.exists()), default=0)
    return f"local-dev-{latest_mtime}" if latest_mtime else "local-dev"


APP_VERSION = runtime_app_version()
APP_VERSION_QUERY = quote(APP_VERSION, safe="")
INVENTORY_SYNC_VERSION_KEY = "inventorySyncVersion"


def app_path_from_env(name: str, default: Path) -> Path:
    value = os.environ.get(name)
    if not value:
        return default
    path = Path(value).expanduser()
    return path if path.is_absolute() else ROOT / path


UPLOADS = app_path_from_env("APP_UPLOADS_DIR", ROOT / "uploads")
FIELD_WORK_MESSAGE_PHOTO_STORAGE_PREFIX = "field-work-message-photos"
FIELD_WORK_MESSAGE_PHOTO_URL_PREFIX = "/uploads/field-work-message-photos"
DATA_JSON = DOCS / "data.json"
DATA_JS = DOCS / "data.js"
OPENING_STOCK_MONTHS = {"2026-04"}
SNAPSHOT_CACHE_ID: int | None = None
SNAPSHOT_CACHE_PAYLOAD: dict | None = None
SALES_DOWNLOAD_URL = "http://b2b1.pxstore.com.tw/farmer/Query_Report.ashx"
SALES_FARMER_ID = "4523"
SALES_DOWNLOAD_TIMEOUT_SECONDS = 30
GENERAL_CHANNEL_PUQIAN_COLUMN_KEYS = {"zhongshan01", "tucheng02", "zhongping06", "minan07"}
GENERAL_CHANNEL_COLUMN_SOURCES = {"puqianTotal"}
FIELD_BT_SHARED_AREA_KEYS = {"three", "five", "threeFive", "four"}
FIELD_WORK_MESSAGE_PHOTO_TYPES = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
    "image/heic": ".heic",
    "image/heif": ".heif",
}
FIELD_WORK_MESSAGE_PHOTO_EXTENSIONS = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
    ".heic": "image/heic",
    ".heif": "image/heif",
}
MAX_FIELD_WORK_MESSAGE_PHOTOS = 6
MAX_FIELD_WORK_MESSAGE_PHOTO_BYTES = 8 * 1024 * 1024
FIELD_WORK_EXPORT_TASK_ORDER = {
    "soilPreparationTractorClean": 1,
    "burn": 2,
    "weed": 3,
    "directSow": 4,
    "seedlingTransplant": 5,
    "bacillusThuringiensis": 6,
    "destroyCrop": 7,
    "harvest": 7,
}
CODEX_PYTHON = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "python"
    / "bin"
    / "python3"
)
PYTHON = str(CODEX_PYTHON) if CODEX_PYTHON.exists() else sys.executable
CODEX_PYTHON_SITE_PACKAGES = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "python"
    / "lib"
    / "python3.12"
    / "site-packages"
)
AUTH_COOKIE_NAME = "app_session"
AUTH_SESSION_DAYS = 30
AUTH_COOKIE_SECURE = os.environ.get("AUTH_COOKIE_SECURE", "1" if os.environ.get("K_SERVICE") else "0") == "1"
PASSWORD_HASH_ITERATIONS = 260_000
USER_ROLES = {"root", "inside", "field"}
ROLE_LABELS = {
    "root": "管理員",
    "inside": "內場",
    "field": "外場",
}
DEFAULT_USERS = [
    {"username": "root", "displayName": "admin", "password": "root1234", "role": "root"},
    {"username": "inside", "displayName": "inside", "password": "inside1234", "role": "inside"},
    {"username": "field", "displayName": "field", "password": "field1234", "role": "field"},
]
ALLOW_DEFAULT_USERS = os.environ.get(
    "ALLOW_DEFAULT_USERS",
    "1" if os.environ.get("DATABASE_BACKEND", "sqlite").strip().lower() == "sqlite" and not os.environ.get("K_SERVICE") else "0",
) == "1"
INITIAL_ROOT_USERNAME = os.environ.get("INITIAL_ROOT_USERNAME", "root").strip() or "root"
INITIAL_ROOT_DISPLAY_NAME = os.environ.get("INITIAL_ROOT_DISPLAY_NAME", "admin").strip() or "admin"
INITIAL_ROOT_PASSWORD = os.environ.get("INITIAL_ROOT_PASSWORD", "")
INSIDE_API_ROLES = {"inside"}
FIELD_API_ROLES = {"field"}
ROOT_API_ROLES = {"root"}
ANY_AUTH_ROLES = {"root", "inside", "field"}


def read_json_body(handler: SimpleHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length) if length else b"{}"
    return json.loads(raw.decode("utf-8"))


def inventory_sync_version(conn) -> str:
    return str(get_setting(conn, INVENTORY_SYNC_VERSION_KEY) or "")


def mark_inventory_sync_changed(conn) -> str:
    version = taipei_now_text()
    save_setting(conn, INVENTORY_SYNC_VERSION_KEY, version)
    return version


def password_hash(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        PASSWORD_HASH_ITERATIONS,
    )
    return f"pbkdf2_sha256${PASSWORD_HASH_ITERATIONS}${salt.hex()}${digest.hex()}"


def password_matches(password: str, encoded: str) -> bool:
    try:
        algorithm, iterations, salt_hex, digest_hex = str(encoded or "").split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            bytes.fromhex(salt_hex),
            int(iterations),
        )
        return hmac.compare_digest(digest.hex(), digest_hex)
    except (TypeError, ValueError):
        return False


def user_payload(row: Any) -> dict[str, Any]:
    role = str(row["role"] or "")
    return {
        "id": int(row["id"]),
        "username": str(row["username"] or ""),
        "displayName": str(row["display_name"] or row["username"] or ""),
        "role": role,
        "roleLabel": ROLE_LABELS.get(role, role),
    }


def account_name_from_user(user: dict[str, Any] | None) -> str:
    if not user:
        return ""
    return str(user.get("displayName") or user.get("username") or "").strip()[:80]


def listed_user_payload(row: Any) -> dict[str, Any]:
    return {
        **user_payload(row),
        "isActive": bool(row["is_active"]),
        "createdAt": str(row["created_at"] or ""),
        "updatedAt": str(row["updated_at"] or ""),
    }


def validate_username(username: str) -> str:
    normalized = str(username or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9_.-]{2,32}", normalized):
        raise ValueError("帳號只能使用 2-32 個英數字、底線、點或連字號")
    return normalized


def validate_display_name(display_name: str) -> str:
    normalized = str(display_name or "").strip()
    if not normalized or len(normalized) > 40:
        raise ValueError("名稱需要 1-40 個字")
    return normalized


def validate_role(role: str) -> str:
    normalized = str(role or "").strip()
    if normalized not in USER_ROLES:
        raise ValueError("角色必須是 root、inside 或 field")
    return normalized


def validate_password(password: str) -> str:
    normalized = str(password or "")
    if len(normalized) < 6:
        raise ValueError("密碼至少需要 6 個字")
    return normalized


def seed_default_users(conn) -> None:
    row = conn.execute("SELECT COUNT(*) AS count FROM users").fetchone()
    if row and int(row["count"]) > 0:
        return
    if not ALLOW_DEFAULT_USERS:
        if not INITIAL_ROOT_PASSWORD:
            raise RuntimeError(
                "No users exist. Set INITIAL_ROOT_PASSWORD for first production startup, "
                "or set ALLOW_DEFAULT_USERS=1 for local development only."
            )
        timestamp = taipei_now_text()
        conn.execute(
            """
            INSERT INTO users (username, display_name, password_hash, role, is_active, created_at, updated_at)
            VALUES (?, ?, ?, 'root', 1, ?, ?)
            """,
            (
                validate_username(INITIAL_ROOT_USERNAME),
                validate_display_name(INITIAL_ROOT_DISPLAY_NAME),
                password_hash(validate_password(INITIAL_ROOT_PASSWORD)),
                timestamp,
                timestamp,
            ),
        )
        conn.commit()
        return
    timestamp = taipei_now_text()
    conn.executemany(
        """
        INSERT INTO users (username, display_name, password_hash, role, is_active, created_at, updated_at)
        VALUES (?, ?, ?, ?, 1, ?, ?)
        """,
        [
            (
                user["username"],
                user["displayName"],
                password_hash(user["password"]),
                user["role"],
                timestamp,
                timestamp,
            )
            for user in DEFAULT_USERS
        ],
    )
    conn.commit()


def read_local_snapshot() -> dict | None:
    if DATA_JSON.exists():
        return json.loads(DATA_JSON.read_text(encoding="utf-8"))
    if DATA_JS.exists():
        text = DATA_JS.read_text(encoding="utf-8").strip()
        prefix = "window.APP_DATA="
        suffix = ";"
        if text.startswith(prefix) and text.endswith(suffix):
            return json.loads(text[len(prefix):-1])
    return None


def current_snapshot_payload(conn) -> dict:
    global SNAPSHOT_CACHE_ID, SNAPSHOT_CACHE_PAYLOAD
    row = conn.execute(
        "SELECT id, payload_json FROM app_snapshots WHERE is_current = 1 ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if row:
        snapshot_id = int(row["id"])
        if SNAPSHOT_CACHE_ID == snapshot_id and SNAPSHOT_CACHE_PAYLOAD is not None:
            return SNAPSHOT_CACHE_PAYLOAD
        payload = snapshot_with_local_template(json.loads(row["payload_json"]))
        SNAPSHOT_CACHE_ID = snapshot_id
        SNAPSHOT_CACHE_PAYLOAD = payload
        return payload

    payload = snapshot_with_local_template(read_local_snapshot() or {})
    SNAPSHOT_CACHE_ID = None
    SNAPSHOT_CACHE_PAYLOAD = payload
    return payload


def validate_sales_date(sales_date: str | None) -> str:
    if not sales_date or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", sales_date):
        raise ValueError("請選擇有效的匯入日期")
    return sales_date


def sales_download_url(sales_date: str) -> str:
    query = urlencode(
        {
            "B_DATE": sales_date,
            "E_DATE": sales_date,
            "GROUP_TYPE": "0",
            "FARMER_ID": SALES_FARMER_ID,
            "CHECK": "true",
        }
    )
    return f"{SALES_DOWNLOAD_URL}?{query}"


def download_sales_workbook(sales_date: str) -> Path:
    UPLOADS.mkdir(exist_ok=True)
    filename = f"SalesData_{SALES_FARMER_ID}_{sales_date}.xlsx"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = UPLOADS / f"{timestamp}_{filename}"
    request = Request(sales_download_url(sales_date), headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request, timeout=SALES_DOWNLOAD_TIMEOUT_SECONDS) as response:
            status = response.getcode()
            content = response.read()
    except HTTPError as exc:
        raise ValueError(f"網站下載失敗，HTTP 回應碼：{exc.code}") from exc
    except URLError as exc:
        raise ValueError(f"網站下載失敗：{exc.reason}") from exc

    if status != 200:
        raise ValueError(f"網站下載失敗，HTTP 回應碼：{status}")
    if not content:
        raise ValueError("網站回傳空的 Excel 檔")
    if not content.startswith(b"PK"):
        raise ValueError("網站回傳內容不是 .xlsx 檔")
    path.write_bytes(content)
    return path


def to_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def sales_payload(start_date: str | None = None, end_date: str | None = None, store_code: str | None = None) -> dict:
    clauses = []
    params: list[str] = []
    if start_date:
        validate_sales_date(start_date)
        clauses.append("sales_date >= ?")
        params.append(start_date)
    if end_date:
        validate_sales_date(end_date)
        clauses.append("sales_date <= ?")
        params.append(end_date)
    if store_code:
        clauses.append("store_code = ?")
        params.append(store_code)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""

    conn = connect()
    try:
        init_db(conn)
        rows = conn.execute(
            f"""
            SELECT
                sales_date,
                store_code,
                store_name,
                product_code,
                product_name,
                sales_qty,
                discard_qty,
                ending_stock
            FROM sales_rows
            {where}
            ORDER BY sales_date, id
            """,
            params,
        ).fetchall()
    finally:
        conn.close()

    sales_rows = [
        {
            "date": row["sales_date"],
            "storeCode": row["store_code"],
            "storeName": row["store_name"],
            "productCode": row["product_code"],
            "productName": row["product_name"],
            "salesQty": to_float(row["sales_qty"]),
            "discardQty": to_float(row["discard_qty"]),
            "endingStock": to_float(row["ending_stock"]),
        }
        for row in rows
    ]
    pivot: dict[tuple[str, str, str], dict] = {}
    totals = {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0}
    for row in sales_rows:
        key = (row["date"], row["storeCode"], row["productCode"])
        if key not in pivot:
            pivot[key] = {
                "date": row["date"],
                "storeCode": row["storeCode"],
                "storeName": row["storeName"],
                "productCode": row["productCode"],
                "productName": row["productName"],
                "salesQty": 0.0,
                "discardQty": 0.0,
                "endingStock": 0.0,
            }
        pivot[key]["salesQty"] += to_float(row.get("salesQty"))
        pivot[key]["discardQty"] += to_float(row.get("discardQty"))
        pivot[key]["endingStock"] += to_float(row.get("endingStock"))
        for field in totals:
            totals[field] += to_float(row.get(field))
    return {
        "salesRows": sales_rows,
        "salesPivot": list(pivot.values()),
        "salesTotals": totals,
        "dates": sorted({row["date"] for row in sales_rows if row.get("date")}),
    }


def previous_month_range(month: str) -> tuple[str, str]:
    parsed = datetime.strptime(month + "-01", "%Y-%m-%d")
    previous_end = parsed.replace(day=1) - timedelta(days=1)
    previous_start = previous_end.replace(day=1)
    return previous_start.date().isoformat(), previous_end.date().isoformat()


def opening_stock_payload(month: str | None, store_code: str | None) -> dict:
    if not month or not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("month 格式錯誤")
    if not store_code:
        raise ValueError("缺少 storeCode")

    conn = connect()
    try:
        init_db(conn)
        snapshot = current_snapshot_payload(conn)
        store = next(
            (
                item
                for item in snapshot.get("storePages", [])
                if str(item.get("storeCode") or "") == str(store_code)
            ),
            None,
        )
        if not store:
            return {"openingStocks": {}, "source": "none"}

        products = store.get("products") or []
        product_codes = [
            str(product.get("productCode") or "")
            for product in products
            if str(product.get("productCode") or "")
        ]
        explicit = {
            str(product.get("productCode") or ""): to_float((product.get("openingStocksByMonth") or {}).get(month))
            for product in products
            if month in (product.get("openingStocksByMonth") or {})
        }
        if explicit:
            return {
                "openingStocks": {product_code: explicit.get(product_code, 0.0) for product_code in product_codes},
                "source": "explicit",
                "month": month,
            }

        first_month = min(
            (
                key
                for product in products
                for key in (product.get("openingStocksByMonth") or {}).keys()
                if re.fullmatch(r"\d{4}-\d{2}", str(key))
            ),
            default="",
        )
        if not first_month or month <= first_month:
            return {"openingStocks": {product_code: 0.0 for product_code in product_codes}, "source": "none", "month": month}

        baseline = {
            str(product.get("productCode") or ""): to_float((product.get("openingStocksByMonth") or {}).get(first_month))
            for product in products
        }
        start_date = f"{first_month}-01"
        _, end_date = previous_month_range(month)
        rows = conn.execute(
            """
            SELECT product_code, SUM(sales_qty) AS sales_qty, SUM(discard_qty) AS discard_qty
            FROM sales_rows
            WHERE store_code = ? AND sales_date >= ? AND sales_date <= ?
            GROUP BY product_code
            """,
            (store_code, start_date, end_date),
        ).fetchall()
        movement = {
            row["product_code"]: to_float(row["sales_qty"]) + to_float(row["discard_qty"])
            for row in rows
        }
        shipment_rows = conn.execute(
            """
            SELECT product_code, SUM(quantity) AS quantity
            FROM shipment_entries
            WHERE store_key = ? AND shipment_date >= ? AND shipment_date <= ?
            GROUP BY product_code
            """,
            (store.get("key"), start_date, end_date),
        ).fetchall()
        shipments = {row["product_code"]: to_float(row["quantity"]) for row in shipment_rows}
        adjustment_rows = conn.execute(
            """
            SELECT product_code, shipment_date, quantity
            FROM stock_adjustment_entries
            WHERE store_key = ? AND shipment_date >= ? AND shipment_date <= ?
            ORDER BY shipment_date
            """,
            (store.get("key"), start_date, end_date),
        ).fetchall()
        adjustments: dict[str, float] = {}
        for row in adjustment_rows:
            adjustment_month = str(row["shipment_date"] or "")[:7]
            product_code = str(row["product_code"] or "")
            adjustments[f"{product_code}|{adjustment_month}"] = to_float(row["quantity"])

        adjustment_totals: dict[str, float] = {}
        for key, quantity in adjustments.items():
            product_code = key.split("|", 1)[0]
            adjustment_totals[product_code] = adjustment_totals.get(product_code, 0.0) + quantity

        opening = {}
        for product_code, value in baseline.items():
            opening[product_code] = (
                value
                + shipments.get(product_code, 0.0)
                - movement.get(product_code, 0.0)
                + adjustment_totals.get(product_code, 0.0)
            )
        return {
            "openingStocks": opening,
            "source": "computed",
            "month": month,
            "fromMonth": first_month,
            "through": end_date,
        }
    finally:
        conn.close()


def inventory_sync_payload(conn, since: str | None = None) -> dict:
    version = inventory_sync_version(conn)
    if since and since == version:
        return {"version": version, "changed": False}
    return {
        "version": version,
        "changed": True,
        "shipmentEntries": get_shipment_entries(conn),
        "stockAdjustmentEntries": get_stock_adjustment_entries(conn),
        "harvestEntries": get_harvest_entries(conn),
        "harvestCellPriorities": get_setting(conn, "harvestCellPriorities") or {},
        "harvestCellFormulas": get_setting(conn, "harvestCellFormulas") or {},
        "harvestFieldExtraColumnsByDate": get_setting(conn, "harvestFieldExtraColumnsByDate") or {},
        "harvestConversionSettings": get_setting(conn, "harvestConversionSettings") or {},
        "cityTableEntries": get_setting(conn, "cityTableEntries") or {},
        "yfyTableEntries": get_setting(conn, "yfyTableEntries") or {},
        "yfyShipmentTimes": get_setting(conn, "yfyShipmentTimes") or {},
        "looseVegetableColumns": get_setting(conn, "looseVegetableColumns") or [],
        "looseVegetableColumnsByDate": get_setting(conn, "looseVegetableColumnsByDate") or {},
        "looseVegetableTableEntries": get_setting(conn, "looseVegetableTableEntries") or {},
        "generalChannelColumns": get_setting(conn, "generalChannelColumns") or [],
        "generalChannelColumnsByDate": get_setting(conn, "generalChannelColumnsByDate") or {},
        "generalChannelTableEntries": get_setting(conn, "generalChannelTableEntries") or {},
        "generalChannelPuqianEntries": get_setting(conn, "generalChannelPuqianEntries") or {},
    }


def lightweight_sales_snapshot(payload: dict, sales_date: str) -> dict:
    next_payload = dict(payload)
    next_payload["generatedAt"] = taipei_now_text()
    next_payload["salesDate"] = sales_date
    next_payload["salesRows"] = []
    next_payload["salesPivot"] = []
    next_payload["salesTotals"] = {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0}
    next_payload["salesByProduct"] = []
    next_payload["discardByProduct"] = []
    return next_payload


def snapshot_with_local_template(payload: dict) -> dict:
    if payload.get("harvestPlanningTemplate"):
        return payload
    local_snapshot = read_local_snapshot() or {}
    template = local_snapshot.get("harvestPlanningTemplate")
    if not template:
        return payload
    return {**payload, "harvestPlanningTemplate": template}


def snapshot_product_rows(payload: dict, dates: list[str]) -> list[dict]:
    products = payload.get("products") or []
    if not products:
        seen = {}
        for row in payload.get("inventory") or []:
            code = str(row.get("productCode") or "").strip()
            name = str(row.get("productName") or "").strip()
            key = code or name
            if key and key not in seen:
                seen[key] = {"productCode": code, "productName": name}
        products = list(seen.values())
    return [
        {
            "productCode": str(product.get("productCode") or "").strip(),
            "productName": str(product.get("productName") or "").strip(),
            "openingStocksByMonth": {month: 0 for month in OPENING_STOCK_MONTHS},
            "sales": [0 for _ in dates],
            "discards": [0 for _ in dates],
            "monthlySales": 0,
            "discardQty": 0,
            "adjustmentQty": 0,
            "existingShipmentQty": 0,
            "shipments": [],
            "shipmentInputQty": 0,
            "currentStock": 0,
        }
        for product in products
        if str(product.get("productCode") or product.get("productName") or "").strip()
    ]


def add_store_to_snapshot(payload: dict, store: dict) -> tuple[dict, dict]:
    payload = dict(payload)
    store_pages = list(payload.get("storePages") or [])
    stores = list(payload.get("stores") or [])
    inventory = list(payload.get("inventory") or [])
    group = str(store.get("group") or "").strip().upper()
    store_name = str(store.get("storeName") or "").strip()
    store_code = str(store.get("storeCode") or "").strip()
    route = str(store.get("route") or "").strip()
    if group not in {"A", "B"}:
        raise ValueError("group 必須是 A 或 B")
    if not store_name:
        raise ValueError("缺少門市名稱")
    if not store_code:
        raise ValueError("缺少門市代號")
    sheet = str(store.get("sheet") or store_name).strip() or store_name
    key = f"{group}:{sheet}"
    existing_keys = {str(item.get("key") or "") for item in store_pages}
    existing_codes = {
        str(item.get("storeCode") or "").strip()
        for item in store_pages
        if str(item.get("group") or "").strip().upper() == group
    }
    if key in existing_keys:
        raise ValueError("這個門市名稱已存在")
    if store_code in existing_codes:
        raise ValueError("這個門市代號已存在")
    dates = sorted({
        str(date)
        for item in store_pages
        for date in (item.get("dates") or [])
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date))
    })
    if not dates and re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(payload.get("salesDate") or "")):
        dates = [str(payload.get("salesDate"))]
    mail_orders = [
        int(item.get("mailOrder"))
        for item in store_pages
        if str(item.get("group") or "").strip().upper() == group and str(item.get("mailOrder") or "").isdigit()
    ]
    mail_order = max(mail_orders, default=-1) + 1
    product_rows = snapshot_product_rows(payload, dates)
    store_page = {
        "key": key,
        "group": group,
        "sheet": sheet,
        "storeName": store_name,
        "storeCode": store_code,
        "route": route,
        "mailOrder": mail_order,
        "dates": dates,
        "products": product_rows,
    }
    store_record = {
        "group": group,
        "sheet": sheet,
        "storeName": store_name,
        "storeCode": store_code,
        "route": route,
        "mailOrder": mail_order,
    }
    inventory_rows = [
        {
            "group": group,
            "storeSheet": sheet,
            "storeName": store_name,
            "storeCode": store_code,
            "route": route,
            "mailOrder": mail_order,
            "productCode": product.get("productCode") or "",
            "productName": product.get("productName") or "",
            "baseStock": 0,
            "salesQty": 0,
            "discardQty": 0,
            "shipmentQty": 0,
            "currentStock": 0,
            "shipmentDates": [],
        }
        for product in product_rows
    ]
    payload["storePages"] = [*store_pages, store_page]
    payload["stores"] = [*stores, store_record]
    payload["inventory"] = [*inventory, *inventory_rows]
    payload["lowStock"] = list(payload.get("lowStock") or []) + inventory_rows
    return payload, store_page


def update_store_name_in_snapshot(payload: dict, store_key: str, store_name: str) -> tuple[dict, dict]:
    payload = dict(payload)
    store_key = str(store_key or "").strip()
    store_name = str(store_name or "").strip()
    if not store_key:
        raise ValueError("缺少 storeKey")
    if not store_name:
        raise ValueError("缺少門市名稱")

    store_pages = list(payload.get("storePages") or [])
    current_store = next((store for store in store_pages if str(store.get("key") or "") == store_key), None)
    if not current_store:
        raise ValueError("找不到門市")

    group = str(current_store.get("group") or "").strip()
    sheet = str(current_store.get("sheet") or "").strip()
    store_code = str(current_store.get("storeCode") or "").strip()
    duplicate = next(
        (
            store
            for store in store_pages
            if str(store.get("key") or "") != store_key
            and str(store.get("group") or "").strip() == group
            and str(store.get("storeName") or "").strip() == store_name
        ),
        None,
    )
    if duplicate:
        raise ValueError("同一群組已有相同門市名稱")

    def is_same_store(record: dict) -> bool:
        if str(record.get("key") or "") == store_key:
            return True
        return str(record.get("group") or "").strip() == group and (
            (sheet and str(record.get("sheet") or record.get("storeSheet") or "").strip() == sheet)
            or (store_code and str(record.get("storeCode") or "").strip() == store_code)
        )

    def rename_records(records: list[dict]) -> list[dict]:
        renamed = []
        for record in records:
            if isinstance(record, dict) and is_same_store(record):
                renamed.append({**record, "storeName": store_name})
            else:
                renamed.append(record)
        return renamed

    payload["storePages"] = rename_records(store_pages)
    payload["stores"] = rename_records(list(payload.get("stores") or []))
    payload["inventory"] = rename_records(list(payload.get("inventory") or []))
    payload["lowStock"] = rename_records(list(payload.get("lowStock") or []))
    updated_store = next(store for store in payload["storePages"] if str(store.get("key") or "") == store_key)
    return payload, updated_store


def delete_store_from_snapshot(payload: dict, store_key: str) -> tuple[dict, dict]:
    payload = dict(payload)
    store_pages = list(payload.get("storePages") or [])
    removed = next((store for store in store_pages if str(store.get("key") or "") == store_key), None)
    if not removed:
        raise ValueError("找不到門市")
    group = str(removed.get("group") or "")
    sheet = str(removed.get("sheet") or "")
    store_code = str(removed.get("storeCode") or "")
    store_name = str(removed.get("storeName") or "")
    payload["storePages"] = [store for store in store_pages if str(store.get("key") or "") != store_key]
    payload["stores"] = [
        store
        for store in (payload.get("stores") or [])
        if not (
            str(store.get("group") or "") == group
            and (
                str(store.get("sheet") or "") == sheet
                or str(store.get("storeCode") or "") == store_code
            )
        )
    ]
    def keep_inventory_row(row: dict) -> bool:
        return not (
            str(row.get("group") or "") == group
            and (
                str(row.get("storeSheet") or "") == sheet
                or str(row.get("storeCode") or "") == store_code
                or str(row.get("storeName") or "") == store_name
            )
        )
    payload["inventory"] = [row for row in (payload.get("inventory") or []) if keep_inventory_row(row)]
    payload["lowStock"] = [row for row in (payload.get("lowStock") or []) if keep_inventory_row(row)]
    return payload, removed


def safe_sheet_title(value: str) -> str:
    text = re.sub(r"[\\/*?:\[\]]", "_", str(value or "").strip()) or "出貨單"
    return text[:31]


def manifest_store_name_parts(store: dict) -> tuple[str, str]:
    store_name = str(store.get("storeName") or "").strip()
    route = str(store.get("route") or "").strip()
    group = str(store.get("group") or "").strip().upper()
    if not group:
        key = str(store.get("key") or "")
        group = key.split(":", 1)[0].strip().upper() if ":" in key else ""
    if group == "A" and route:
        return store_name, f" {route}"
    return store_name, ""


def manifest_store_name_cell_value(store: dict) -> str:
    store_name, route_suffix = manifest_store_name_parts(store)
    return f"{store_name}{route_suffix}".strip()


def compact_shipment_date(shipment_date: str) -> str:
    match = re.fullmatch(r"\d{4}-(\d{2})-(\d{2})", shipment_date or "")
    if not match:
        return "0000"
    return f"{match.group(1)}{match.group(2)}"


def field_work_month_bounds(month: str) -> tuple[str, str]:
    if not re.fullmatch(r"\d{4}-\d{2}", month or ""):
        raise ValueError("month 格式錯誤")
    start = datetime.strptime(f"{month}-01", "%Y-%m-%d")
    end = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def natural_code_sort_key(value: object) -> list[tuple[int, object]]:
    parts = re.split(r"(\d+)", str(value or ""))
    key: list[tuple[int, object]] = []
    for part in parts:
        if part == "":
            continue
        key.append((0, int(part)) if part.isdigit() else (1, part.lower()))
    return key


def normalize_city_table_entries(value: object) -> dict:
    if not isinstance(value, dict):
        raise ValueError("cityTableEntries 必須是物件")
    allowed_columns = {
        "stock",
        "farEastern",
        "tianmu",
        "fuxing",
        "banqiao",
        "zhubei",
        "hsinchu",
        "reserve",
    }
    normalized: dict[str, dict[str, dict[str, float]]] = {}
    for raw_date, rows in value.items():
        date = str(raw_date)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date) or not isinstance(rows, dict):
            continue
        for raw_row_key, columns in rows.items():
            row_key = str(raw_row_key)
            if not re.fullmatch(r"row:\d+", row_key) or not isinstance(columns, dict):
                continue
            for raw_column_key, raw_quantity in columns.items():
                column_key = str(raw_column_key)
                if column_key not in allowed_columns:
                    continue
                try:
                    quantity = float(raw_quantity)
                except (TypeError, ValueError):
                    continue
                if quantity <= 0:
                    continue
                normalized.setdefault(date, {}).setdefault(row_key, {})[column_key] = quantity
    return normalized


def normalize_yfy_table_entries(value: object) -> dict:
    if not isinstance(value, dict):
        raise ValueError("yfyTableEntries 必須是物件")
    allowed_columns = {"greenSafe", "greenSafeBaibao"}
    normalized: dict[str, dict[str, dict[str, float]]] = {}
    for raw_date, rows in value.items():
        date = str(raw_date)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date) or not isinstance(rows, dict):
            continue
        for raw_row_key, columns in rows.items():
            row_key = str(raw_row_key)
            if not re.fullmatch(r"row:\d+", row_key) or not isinstance(columns, dict):
                continue
            for raw_column_key, raw_quantity in columns.items():
                column_key = str(raw_column_key)
                if column_key not in allowed_columns:
                    continue
                try:
                    quantity = float(raw_quantity)
                except (TypeError, ValueError):
                    continue
                if quantity <= 0:
                    continue
                normalized.setdefault(date, {}).setdefault(row_key, {})[column_key] = quantity
    return normalized


def normalize_yfy_shipment_times(value: object) -> dict:
    if value is None:
        return {}
    if not isinstance(value, dict):
        raise ValueError("yfyShipmentTimes 必須是物件")
    normalized: dict[str, str] = {}
    for raw_date, raw_time in value.items():
        date = str(raw_date)
        shipment_time = str(raw_time or "").strip()
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            continue
        if not re.fullmatch(r"([01]\d|2[0-3]):[0-5]\d", shipment_time):
            continue
        normalized[date] = shipment_time
    return normalized


def normalize_loose_vegetable_columns(value: object) -> list[dict[str, str]]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("looseVegetableColumns 必須是陣列")
    normalized: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        key = str(item.get("key") or "").strip()
        label = str(item.get("label") or "").strip()
        if not re.fullmatch(r"[A-Za-z0-9_-]{1,48}", key):
            continue
        if not label or key in seen:
            continue
        normalized.append({"key": key, "label": label[:24]})
        seen.add(key)
    return normalized


def normalize_loose_vegetable_columns_by_date(value: object) -> dict:
    if value is None:
        return {}
    if not isinstance(value, dict):
        raise ValueError("looseVegetableColumnsByDate 必須是物件")
    normalized: dict[str, list[dict[str, str]]] = {}
    for raw_date, columns in value.items():
        date = str(raw_date)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            continue
        normalized_columns = normalize_loose_vegetable_columns(columns)
        if normalized_columns:
            normalized[date] = normalized_columns
    return normalized


def loose_vegetable_columns_for_date(columns: object, date: str) -> list[dict[str, str]]:
    if isinstance(columns, dict):
        return normalize_loose_vegetable_columns(columns.get(date))
    return normalize_loose_vegetable_columns(columns)


def normalize_loose_vegetable_table_entries(value: object, columns: object) -> dict:
    if not isinstance(value, dict):
        raise ValueError("looseVegetableTableEntries 必須是物件")
    normalized: dict[str, dict[str, dict[str, float]]] = {}
    for raw_date, rows in value.items():
        date = str(raw_date)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date) or not isinstance(rows, dict):
            continue
        allowed_columns = {
            str(column.get("key") or "")
            for column in loose_vegetable_columns_for_date(columns, date)
        }
        for raw_row_key, row_columns in rows.items():
            row_key = str(raw_row_key)
            if not re.fullmatch(r"row:\d+", row_key) or not isinstance(row_columns, dict):
                continue
            for raw_column_key, raw_quantity in row_columns.items():
                column_key = str(raw_column_key)
                if column_key not in allowed_columns:
                    continue
                try:
                    quantity = float(raw_quantity)
                except (TypeError, ValueError):
                    continue
                if quantity <= 0:
                    continue
                normalized.setdefault(date, {}).setdefault(row_key, {})[column_key] = quantity
    return normalized


def normalize_general_channel_columns(value: object) -> list[dict[str, str]]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("generalChannelColumns 必須是陣列")
    normalized: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        key = str(item.get("key") or "").strip()
        label = str(item.get("label") or "").strip()
        source = str(item.get("source") or "").strip()
        if not re.fullmatch(r"[A-Za-z0-9_-]{1,48}", key):
            continue
        if not label or key in seen:
            continue
        normalized_item = {"key": key, "label": label[:24]}
        if source in GENERAL_CHANNEL_COLUMN_SOURCES or label[:24] == "埔墘統倉":
            normalized_item["source"] = "puqianTotal"
        normalized.append(normalized_item)
        seen.add(key)
    return normalized


def normalize_general_channel_columns_by_date(value: object) -> dict:
    if value is None:
        return {}
    if not isinstance(value, dict):
        raise ValueError("generalChannelColumnsByDate 必須是物件")
    normalized: dict[str, list[dict[str, str]]] = {}
    for raw_date, columns in value.items():
        date = str(raw_date)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            continue
        normalized_columns = normalize_general_channel_columns(columns)
        if normalized_columns:
            normalized[date] = normalized_columns
    return normalized


def general_channel_columns_for_date(columns: object, date: str) -> list[dict[str, str]]:
    if isinstance(columns, dict):
        return normalize_general_channel_columns(columns.get(date))
    return normalize_general_channel_columns(columns)


def normalize_general_channel_table_entries(value: object, columns: object) -> dict:
    if not isinstance(value, dict):
        raise ValueError("generalChannelTableEntries 必須是物件")
    normalized: dict[str, dict[str, dict[str, float]]] = {}
    for raw_date, rows in value.items():
        date = str(raw_date)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date) or not isinstance(rows, dict):
            continue
        allowed_columns = {
            str(column.get("key") or "")
            for column in general_channel_columns_for_date(columns, date)
            if str(column.get("source") or "") != "puqianTotal"
        }
        for raw_row_key, row_columns in rows.items():
            row_key = str(raw_row_key)
            if not re.fullmatch(r"row:\d+", row_key) or not isinstance(row_columns, dict):
                continue
            for raw_column_key, raw_quantity in row_columns.items():
                column_key = str(raw_column_key)
                if column_key not in allowed_columns:
                    continue
                try:
                    quantity = float(raw_quantity)
                except (TypeError, ValueError):
                    continue
                if quantity <= 0:
                    continue
                normalized.setdefault(date, {}).setdefault(row_key, {})[column_key] = quantity
    return normalized


def normalize_general_channel_puqian_entries(value: object) -> dict:
    if value is None:
        return {}
    if not isinstance(value, dict):
        raise ValueError("generalChannelPuqianEntries 必須是物件")
    normalized: dict[str, dict[str, dict[str, float]]] = {}
    for raw_date, rows in value.items():
        date = str(raw_date)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date) or not isinstance(rows, dict):
            continue
        for raw_row_key, row_columns in rows.items():
            row_key = str(raw_row_key)
            if not re.fullmatch(r"row:\d+", row_key) or not isinstance(row_columns, dict):
                continue
            for raw_column_key, raw_quantity in row_columns.items():
                column_key = str(raw_column_key)
                if column_key not in GENERAL_CHANNEL_PUQIAN_COLUMN_KEYS:
                    continue
                try:
                    quantity = float(raw_quantity)
                except (TypeError, ValueError):
                    continue
                if quantity <= 0:
                    continue
                normalized.setdefault(date, {}).setdefault(row_key, {})[column_key] = quantity
    return normalized


HARVEST_PRIORITY_EXCLUDED_COLUMNS = {"C", "SI", "SJ"}


def normalize_harvest_priority_entries(value: object) -> list[dict[str, int | str]]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("priorities 必須是陣列")
    normalized: list[dict[str, int | str]] = []
    seen_priorities: set[tuple[int, int]] = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        try:
            row_index = int(item.get("rowIndex") or 0)
            priority = int(item.get("priority") or 0)
        except (TypeError, ValueError):
            continue
        column_letter = str(item.get("columnLetter") or "").strip().upper()
        if row_index < 1 or not re.fullmatch(r"[A-Z]+", column_letter):
            continue
        if column_letter in HARVEST_PRIORITY_EXCLUDED_COLUMNS:
            continue
        if priority < 1 or priority > 7:
            continue
        priority_key = (row_index, priority)
        if priority_key in seen_priorities:
            continue
        seen_priorities.add(priority_key)
        normalized.append(
            {
                "rowIndex": row_index,
                "columnLetter": column_letter,
                "priority": priority,
            }
        )
    return normalized


def normalize_harvest_formula_entries(value: object) -> list[dict[str, int | str]]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("formulas 必須是陣列")
    normalized: list[dict[str, int | str]] = []
    seen_cells: set[tuple[int, str]] = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        try:
            row_index = int(item.get("rowIndex") or 0)
        except (TypeError, ValueError):
            continue
        column_letter = str(item.get("columnLetter") or "").strip().upper()
        formula = str(item.get("formula") or "").strip()
        if row_index < 1 or not re.fullmatch(r"[A-Z]+", column_letter):
            continue
        if not formula:
            continue
        if len(formula) > 160:
            formula = formula[:160]
        cell_key = (row_index, column_letter)
        if cell_key in seen_cells:
            continue
        seen_cells.add(cell_key)
        normalized.append(
            {
                "rowIndex": row_index,
                "columnLetter": column_letter,
                "formula": formula,
            }
        )
    return normalized


def field_work_month_workbook_bytes(month: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError:
        if CODEX_PYTHON_SITE_PACKAGES.exists():
            sys.path.insert(0, str(CODEX_PYTHON_SITE_PACKAGES))
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        else:
            raise

    start_date, end_date = field_work_month_bounds(month)
    conn = connect()
    try:
        zone_rows = conn.execute(
            """
            SELECT zone_name, sort_order
            FROM field_zones
            ORDER BY sort_order, zone_name
            """
        ).fetchall()
        if not zone_rows:
            zone_rows = conn.execute(
                """
                SELECT DISTINCT zone_name, 0 AS sort_order
                FROM (
                    SELECT zone_name
                    FROM field_work_records
                    WHERE work_date >= ? AND work_date < ?
                    UNION
                    SELECT r.zone_name
                    FROM net_house_status_records r
                    WHERE r.status = '種植'
                        AND (
                            (r.harvest_date >= ? AND r.harvest_date < ?)
                            OR EXISTS (
                                SELECT 1
                                FROM net_house_status_record_crops c
                                WHERE c.record_date = r.record_date
                                    AND c.zone_name = r.zone_name
                                    AND c.net_house_code = r.net_house_code
                                    AND c.harvest_date >= ?
                                    AND c.harvest_date < ?
                            )
                        )
                ) export_zones
                ORDER BY zone_name
                """,
                (start_date, end_date, start_date, end_date, start_date, end_date),
            ).fetchall()
        record_rows = conn.execute(
            """
            SELECT
                r.work_date,
                r.zone_name,
                r.net_house_code,
                r.task_key,
                r.task_label,
                r.fertilizer_bag_count,
                r.crop_key,
                r.crop_name,
                r.account_name,
                h.grid_count
            FROM field_work_records r
            LEFT JOIN field_net_houses h
                ON h.zone_name = r.zone_name
                AND h.net_house_code = r.net_house_code
            WHERE r.work_date >= ? AND r.work_date < ?
            """,
            (start_date, end_date),
        ).fetchall()
        crop_rows = conn.execute(
            """
            SELECT
                work_date,
                zone_name,
                net_house_code,
                task_key,
                crop_key,
                crop_name,
                tray_count
            FROM field_work_record_crops
            WHERE work_date >= ? AND work_date < ?
            ORDER BY crop_index
            """,
            (start_date, end_date),
        ).fetchall()
        harvest_crop_rows = conn.execute(
            """
            SELECT
                r.record_date,
                r.zone_name,
                r.net_house_code,
                r.planted_date,
                c.crop_index,
                c.crop_name,
                c.harvest_date,
                c.estimated_quantity,
                c.harvest_quantity,
                c.destroyed,
                h.grid_count
            FROM net_house_status_records r
            JOIN net_house_status_record_crops c
                ON c.record_date = r.record_date
                AND c.zone_name = r.zone_name
                AND c.net_house_code = r.net_house_code
            LEFT JOIN field_net_houses h
                ON h.zone_name = r.zone_name
                AND h.net_house_code = r.net_house_code
            WHERE r.status = '種植'
                AND c.harvest_date >= ?
                AND c.harvest_date < ?
            ORDER BY c.harvest_date, r.zone_name, r.net_house_code, c.crop_index
            """,
            (start_date, end_date),
        ).fetchall()
        harvest_fallback_rows = conn.execute(
            """
            SELECT
                r.record_date,
                r.zone_name,
                r.net_house_code,
                r.planted_date,
                r.crop_name,
                r.harvest_date,
                r.estimated_quantity,
                r.harvest_quantity,
                h.grid_count,
                0 AS destroyed
            FROM net_house_status_records r
            LEFT JOIN field_net_houses h
                ON h.zone_name = r.zone_name
                AND h.net_house_code = r.net_house_code
            WHERE r.status = '種植'
                AND r.harvest_date >= ?
                AND r.harvest_date < ?
                AND NOT EXISTS (
                    SELECT 1
                    FROM net_house_status_record_crops c
                    WHERE c.record_date = r.record_date
                        AND c.zone_name = r.zone_name
                        AND c.net_house_code = r.net_house_code
                        AND NULLIF(c.harvest_date, '') IS NOT NULL
                )
            ORDER BY r.harvest_date, r.zone_name, r.net_house_code
            """,
            (start_date, end_date),
        ).fetchall()
    finally:
        conn.close()

    crop_items_by_record: dict[tuple[str, str, str, str], list[dict[str, Any]]] = {}
    for row in crop_rows:
        key = (row["work_date"], row["zone_name"], row["net_house_code"], row["task_key"])
        crop_items_by_record.setdefault(key, []).append(
            {
                "cropName": row["crop_name"],
                "trayCount": row["tray_count"],
            }
        )

    def split_crop_names_for_export(crop_name: object) -> list[str]:
        return [part.strip() for part in str(crop_name or "").split("、") if part.strip()]

    records_by_zone: dict[str, list[dict[str, Any]]] = {}
    for row in record_rows:
        row_payload = dict(row)
        row_payload["estimated_quantity"] = None
        row_payload["harvest_quantity"] = None
        records_by_zone.setdefault(row["zone_name"], []).append(row_payload)

    seen_harvest_rows: set[tuple[str, str, str, str, str]] = set()

    def add_harvest_export_row(
        harvest_date: str,
        zone_name: str,
        net_house_code: str,
        crop_name: str,
        destroyed: bool = False,
        grid_count: object = None,
        estimated_quantity: object = None,
        harvest_quantity: object = None,
    ) -> None:
        task_key = "destroyCrop" if destroyed else "harvest"
        task_label = "打掉" if destroyed else "採收"
        row_key = (harvest_date, zone_name, net_house_code, crop_name, task_key)
        if row_key in seen_harvest_rows:
            return
        seen_harvest_rows.add(row_key)
        records_by_zone.setdefault(zone_name, []).append(
            {
                "work_date": harvest_date,
                "zone_name": zone_name,
                "net_house_code": net_house_code,
                "task_key": task_key,
                "task_label": task_label,
                "fertilizer_bag_count": None,
                "crop_key": "",
                "crop_name": crop_name,
                "account_name": "",
                "grid_count": grid_count,
                "estimated_quantity": 0 if destroyed else estimated_quantity,
                "harvest_quantity": 0 if destroyed else harvest_quantity,
            }
        )

    for row in harvest_crop_rows:
        add_harvest_export_row(
            str(row["harvest_date"] or ""),
            str(row["zone_name"] or ""),
            str(row["net_house_code"] or ""),
            str(row["crop_name"] or ""),
            bool(row["destroyed"]),
            row["grid_count"],
            row["estimated_quantity"],
            row["harvest_quantity"],
        )

    for row in harvest_fallback_rows:
        crop_names = split_crop_names_for_export(row["crop_name"]) or [""]
        for crop_name in crop_names:
            add_harvest_export_row(
                str(row["harvest_date"] or ""),
                str(row["zone_name"] or ""),
                str(row["net_house_code"] or ""),
                crop_name,
                bool(row["destroyed"]),
                row["grid_count"],
                row["estimated_quantity"],
                row["harvest_quantity"],
            )

    def number_or_blank(value: object) -> int | float | None:
        if value in (None, ""):
            return None
        number = float(value)
        return int(number) if number.is_integer() else number

    def package_count_for_record(row: Any) -> int | float | None:
        if row["task_key"] == "soilPreparationTractorClean":
            return number_or_blank(row["fertilizer_bag_count"])
        return None

    def tray_count_for_export(row: Any, item: dict[str, Any]) -> int | float | str | None:
        if row["task_key"] == "directSow":
            return None
        return number_or_blank(item.get("trayCount"))

    wb = Workbook()
    header = ["日期", "網室編號", "格數", "工作項目", "使用包數", "作物", "盤數", "預估量", "採收量", "填寫帳號"]
    font = Font(name="Microsoft JhengHei", size=12)
    title_font = Font(name="Microsoft JhengHei", size=14, bold=True)
    header_font = Font(name="Microsoft JhengHei", size=12, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DCE2DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E9F3EE")

    zone_names = [row["zone_name"] for row in zone_rows] or ["工作紀錄"]
    for zone_index, zone_name in enumerate(zone_names):
        ws = wb.active if zone_index == 0 else wb.create_sheet()
        ws.title = safe_sheet_title(zone_name)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        title_cell = ws.cell(row=1, column=1, value=f"{month} 田間工作紀錄 - {zone_name}")
        title_cell.font = title_font
        title_cell.alignment = center
        for column_index, value in enumerate(header, start=1):
            cell = ws.cell(row=2, column=column_index, value=value)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border

        rows = sorted(
            records_by_zone.get(zone_name, []),
            key=lambda row: (
                row["work_date"],
                natural_code_sort_key(row["net_house_code"]),
                FIELD_WORK_EXPORT_TASK_ORDER.get(row["task_key"], 99),
                row["task_key"],
            ),
        )
        current_row = 3
        if not rows:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(header))
            cell = ws.cell(row=current_row, column=1, value="本月沒有工作紀錄")
            cell.font = font
            cell.alignment = center
            cell.border = border
        for row in rows:
            key = (row["work_date"], row["zone_name"], row["net_house_code"], row["task_key"])
            crop_items = crop_items_by_record.get(key, [])
            if not crop_items and (row["crop_key"] or row["crop_name"]):
                crop_items = [{"cropName": row["crop_name"], "trayCount": None}]
            export_items = crop_items if crop_items else [{"cropName": "", "trayCount": None}]
            for item in export_items:
                values = [
                    row["work_date"],
                    row["net_house_code"],
                    number_or_blank(row["grid_count"]),
                    row["task_label"],
                    package_count_for_record(row),
                    item.get("cropName") or "",
                    tray_count_for_export(row, item),
                    number_or_blank(row["estimated_quantity"]),
                    number_or_blank(row["harvest_quantity"]),
                    row["account_name"],
                ]
                for column_index, value in enumerate(values, start=1):
                    cell = ws.cell(row=current_row, column=column_index, value=value)
                    cell.font = font
                    cell.alignment = left if column_index in {2, 4, 6, 10} else center
                    cell.border = border
                current_row += 1

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:J{max(ws.max_row, 2)}"
        widths = [13, 16, 10, 24, 12, 18, 10, 12, 12, 16]
        for column_index, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + column_index)].width = width
        ws.row_dimensions[1].height = 24

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def field_bt_month_summary_workbook_bytes(month: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError:
        if CODEX_PYTHON_SITE_PACKAGES.exists():
            sys.path.insert(0, str(CODEX_PYTHON_SITE_PACKAGES))
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        else:
            raise

    start_date, end_date = field_work_month_bounds(month)
    conn = connect()
    try:
        rows = conn.execute(
            """
            SELECT work_date, area_key, package_count
            FROM field_bt_records
            WHERE work_date >= ? AND work_date < ?
            ORDER BY work_date, area_key
            """,
            (start_date, end_date),
        ).fetchall()
    finally:
        conn.close()

    one_two_total = 0
    shared_package_by_date: dict[str, int] = {}
    for row in rows:
        package_count = row["package_count"]
        if package_count in (None, ""):
            continue
        package_count = int(package_count)
        area_key = row["area_key"]
        if area_key == "oneTwo":
            one_two_total += package_count
        elif area_key in FIELD_BT_SHARED_AREA_KEYS:
            work_date = row["work_date"]
            shared_package_by_date[work_date] = package_count

    wb = Workbook()
    ws = wb.active
    ws.title = "蘇力菌統計"
    font = Font(name="Microsoft JhengHei", size=12)
    title_font = Font(name="Microsoft JhengHei", size=14, bold=True)
    header_font = Font(name="Microsoft JhengHei", size=12, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DCE2DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E9F3EE")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_cell = ws.cell(row=1, column=1, value=f"{month} 蘇力菌使用包數統計")
    title_cell.font = title_font
    title_cell.alignment = center

    headers = ["月份", "場區", "使用包數總計"]
    for column_index, value in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=column_index, value=value)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    summary_rows = [
        [month, "1、2場", one_two_total],
        [month, "3、3F、4、5場", sum(shared_package_by_date.values())],
    ]
    for row_index, values in enumerate(summary_rows, start=3):
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.font = font
            cell.alignment = left if column_index == 2 else center
            cell.border = border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def daily_vegetable_quantity_text(value: object) -> str:
    text = str(value if value is not None else "").strip()
    if not text:
        return ""
    try:
        number = float(text)
    except (TypeError, ValueError):
        return text
    return str(int(number)) if number.is_integer() else f"{number:g}"


DAILY_VEGETABLE_FIELD_FILL_COLORS = {
    "三場": "C6EFCE",
    "三場F": "C6EFCE",
    "四場": "F2CEEF",
    "五場": "CAEDFB",
}


def daily_vegetable_workbook_bytes(payload: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        if CODEX_PYTHON_SITE_PACKAGES.exists():
            sys.path.insert(0, str(CODEX_PYTHON_SITE_PACKAGES))
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        else:
            raise

    start_date = str(payload.get("startDate") or "").strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", start_date):
        raise ValueError("startDate 格式錯誤")

    raw_dates = payload.get("dates") or []
    raw_rows = payload.get("rows") or []
    raw_totals = payload.get("totals") or []
    if not isinstance(raw_dates, list) or not isinstance(raw_rows, list) or not isinstance(raw_totals, list):
        raise ValueError("dates、rows 與 totals 必須是陣列")

    dates: list[dict[str, str]] = []
    for item in raw_dates:
        if not isinstance(item, dict):
            continue
        date = str(item.get("date") or "").strip()
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            continue
        label = str(item.get("label") or date).strip() or date
        dates.append({"date": date, "label": label})
    if not dates:
        raise ValueError("缺少可匯出的日期")

    totals_by_date: dict[str, dict[str, Any]] = {}
    for item in raw_totals:
        if not isinstance(item, dict):
            continue
        date = str(item.get("date") or "").strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            totals_by_date[date] = item

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title("每日菜量表")

    column_count = 1 + len(dates) * 2
    title_font = Font(name="Microsoft JhengHei", size=14, bold=True)
    header_font = Font(name="Microsoft JhengHei", size=12, bold=True)
    body_font = Font(name="Microsoft JhengHei", size=12)
    total_font = Font(name="Microsoft JhengHei", size=12, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DCE2DC")
    red = Side(style="thick", color="D00000")
    product_separator = Side(style="medium", color="FF000000")
    inner_product_separator = Side(style="dotted", color="FF000000")
    total_top = Side(style="medium", color="9AA69E")
    header_fill = PatternFill("solid", fgColor="E9F3EE")
    total_fill = PatternFill("solid", fgColor="FFF8EA")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = ws.cell(row=1, column=1, value=f"{start_date} 起 每日菜量表")
    title_cell.font = title_font
    title_cell.alignment = center

    ws.cell(row=2, column=1, value="作物")
    for index, date_item in enumerate(dates):
        start_col = 2 + index * 2
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 1)
        ws.cell(row=2, column=start_col, value=date_item["label"])

    current_row = 3
    field_cell_fill_colors: dict[tuple[int, int], str] = {}
    product_boundary_rows: set[int] = set()
    product_boundary_previous_rows: set[int] = set()
    product_inner_bottom_rows: set[int] = set()
    product_inner_top_rows: set[int] = set()
    rendered_product_index = 0
    for row_item in raw_rows:
        if not isinstance(row_item, dict):
            continue
        product_name = str(row_item.get("productName") or "").strip()
        cells = row_item.get("cells") if isinstance(row_item.get("cells"), list) else []
        cells_by_date: dict[str, dict[str, Any]] = {}
        for cell_item in cells:
            if not isinstance(cell_item, dict):
                continue
            date = str(cell_item.get("date") or "").strip()
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
                cells_by_date[date] = cell_item

        items_by_date: dict[str, list[dict[str, str]]] = {}
        max_item_rows = 1
        for date_item in dates:
            cell_item = cells_by_date.get(date_item["date"], {})
            raw_items = cell_item.get("items") if isinstance(cell_item, dict) else []
            items = raw_items if isinstance(raw_items, list) else []
            normalized_items: list[dict[str, str]] = []
            for entry in items:
                if not isinstance(entry, dict):
                    continue
                lot = str(entry.get("lot") or "").strip()
                quantity = daily_vegetable_quantity_text(entry.get("quantity"))
                field_name = str(entry.get("fieldName") or "").strip()
                field_color = DAILY_VEGETABLE_FIELD_FILL_COLORS.get(field_name)
                if lot or quantity:
                    normalized_items.append({"lot": lot, "quantity": quantity, "fieldColor": field_color or ""})
            items_by_date[date_item["date"]] = normalized_items
            max_item_rows = max(max_item_rows, len(normalized_items))

        if rendered_product_index > 0:
            product_boundary_rows.add(current_row)
            product_boundary_previous_rows.add(current_row - 1)
        if max_item_rows > 1:
            product_inner_bottom_rows.update(range(current_row, current_row + max_item_rows - 1))
            product_inner_top_rows.update(range(current_row + 1, current_row + max_item_rows))

        for item_row_index in range(max_item_rows):
            row_number = current_row + item_row_index
            ws.cell(row=row_number, column=1, value=product_name if item_row_index == 0 else None)
            for index, date_item in enumerate(dates):
                start_col = 2 + index * 2
                item = items_by_date.get(date_item["date"], [])[item_row_index:item_row_index + 1]
                entry = item[0] if item else {}
                ws.cell(row=row_number, column=start_col, value=entry.get("lot") or None)
                ws.cell(row=row_number, column=start_col + 1, value=entry.get("quantity") or None)
                if entry.get("fieldColor"):
                    field_cell_fill_colors[(row_number, start_col)] = entry["fieldColor"]
            ws.row_dimensions[row_number].height = 20
        if max_item_rows > 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + max_item_rows - 1, end_column=1)
        current_row += max_item_rows
        rendered_product_index += 1

    total_row = current_row
    ws.cell(row=total_row, column=1, value="小計")
    for index, date_item in enumerate(dates):
        start_col = 2 + index * 2
        total = totals_by_date.get(date_item["date"], {})
        try:
            count = float(total.get("count") or 0)
            quantity = float(total.get("quantity") or 0)
        except (TypeError, ValueError):
            count = 0
            quantity = 0
        ws.cell(row=total_row, column=start_col, value=None)
        if count:
            total_value = int(quantity) if float(quantity).is_integer() else quantity
            ws.cell(row=total_row, column=start_col + 1, value=total_value)
        else:
            ws.cell(row=total_row, column=start_col + 1, value=None)

    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=column_count):
        for cell in row:
            is_header = cell.row == 2
            is_total = cell.row == total_row
            is_boundary = cell.column > 2 and cell.column % 2 == 0
            field_fill_color = field_cell_fill_colors.get((cell.row, cell.column))
            cell.font = header_font if is_header else total_font if is_total else body_font
            cell.fill = (
                header_fill
                if is_header
                else total_fill
                if is_total
                else PatternFill("solid", fgColor=field_fill_color)
                if field_fill_color
                else PatternFill(fill_type=None)
            )
            cell.alignment = left if cell.column == 1 and not is_header else center
            cell.border = Border(
                left=red if is_boundary else thin,
                right=thin,
                top=(
                    total_top
                    if is_total
                    else product_separator
                    if cell.row in product_boundary_rows
                    else inner_product_separator
                    if cell.row in product_inner_top_rows
                    else thin
                ),
                bottom=(
                    product_separator
                    if cell.row in product_boundary_previous_rows
                    else inner_product_separator
                    if cell.row in product_inner_bottom_rows
                    else thin
                ),
            )

    ws.freeze_panes = "B3"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:2"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.column_dimensions["A"].width = 14
    for index in range(len(dates)):
        lot_col = 2 + index * 2
        quantity_col = lot_col + 1
        ws.column_dimensions[get_column_letter(lot_col)].width = 11
        ws.column_dimensions[get_column_letter(quantity_col)].width = 8
    for row in ws.iter_rows(min_row=total_row, max_row=total_row, min_col=1, max_col=column_count):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0" if float(cell.value).is_integer() else "#,##0.##"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def manifest_workbook_bytes(payload: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Alignment, Border, Font, Side
    except ModuleNotFoundError:
        if CODEX_PYTHON_SITE_PACKAGES.exists():
            sys.path.insert(0, str(CODEX_PYTHON_SITE_PACKAGES))
            from openpyxl import Workbook
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            from openpyxl.styles import Alignment, Border, Font, Side
        else:
            raise

    shipment_date = str(payload.get("shipmentDate") or "").strip()
    round_label = str(payload.get("roundLabel") or "").strip() or "未命名輪次"
    stores = payload.get("stores") or []
    rows = payload.get("rows") or []
    box_rows = payload.get("boxRows") or []
    grand_total = payload.get("grandTotal") or 0
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", shipment_date):
        raise ValueError("shipmentDate 格式錯誤")
    if not isinstance(stores, list) or not isinstance(rows, list) or not isinstance(box_rows, list):
        raise ValueError("stores、rows 與 boxRows 必須是陣列")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(f"{shipment_date}_{round_label}")

    font = Font(name="Microsoft JhengHei", size=12)
    bold_font = Font(name="Microsoft JhengHei", size=12, bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DCE2DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row_1 = 1
    header_row_2 = 2
    header_row_3 = 3
    data_start_row = 4

    ws.cell(row=header_row_1, column=1, value="呼出碼")
    ws.cell(row=header_row_1, column=2, value="品項")
    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_3, end_column=1)
    ws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_3, end_column=2)

    for idx, store in enumerate(stores, start=3):
        ws.cell(row=header_row_1, column=idx, value=str(store.get("route") or ""))
        ws.cell(row=header_row_2, column=idx, value=str(store.get("storeCode") or store.get("sheet") or ""))
        store_name, route_suffix = manifest_store_name_parts(store)
        if route_suffix:
            ws.cell(
                row=header_row_3,
                column=idx,
                value=CellRichText(
                    TextBlock(InlineFont(rFont="Microsoft JhengHei", sz=12, b=True), store_name),
                    TextBlock(InlineFont(rFont="Microsoft JhengHei", sz=2, b=True), route_suffix),
                ),
            )
        else:
            ws.cell(row=header_row_3, column=idx, value=store_name)

    total_col = len(stores) + 3
    ws.cell(row=header_row_1, column=total_col, value="出貨量加總")
    ws.merge_cells(start_row=header_row_1, start_column=total_col, end_row=header_row_3, end_column=total_col)

    current_row = data_start_row
    for row in rows:
        ws.cell(row=current_row, column=1, value=str(row.get("productCode") or ""))
        ws.cell(row=current_row, column=2, value=str(row.get("productName") or ""))
        quantities = row.get("quantities") or {}
        for idx, store in enumerate(stores, start=3):
            value = quantities.get(store.get("key"))
            ws.cell(row=current_row, column=idx, value=float(value) if value not in (None, "", 0) else None)
        total_value = row.get("totalQty")
        ws.cell(row=current_row, column=total_col, value=float(total_value) if total_value not in (None, "", 0) else None)
        current_row += 1

    subtotal_row = current_row
    ws.cell(row=subtotal_row, column=1, value="小計")
    for idx, store in enumerate(stores, start=3):
        total_qty = store.get("totalQty")
        ws.cell(row=subtotal_row, column=idx, value=float(total_qty) if total_qty not in (None, "", 0) else None)
    ws.cell(row=subtotal_row, column=total_col, value=float(grand_total) if grand_total not in (None, "", 0) else None)
    current_row = subtotal_row + 2

    box_value_rows: list[int] = []
    for box_row in box_rows:
        box_value_rows.append(current_row)
        ws.cell(row=current_row, column=1, value=str(box_row.get("label") or ""))
        quantities = box_row.get("quantities") or {}
        for idx, store in enumerate(stores, start=3):
            value = quantities.get(store.get("key"))
            ws.cell(row=current_row, column=idx, value=float(value) if value not in (None, "", 0) else None)
        total_value = box_row.get("totalQty")
        ws.cell(row=current_row, column=total_col, value=float(total_value) if total_value not in (None, "", 0) else None)
        current_row += 1
    last_row = current_row - 1 if box_value_rows else subtotal_row

    emphasized_rows = {subtotal_row, *box_value_rows}
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=total_col):
        for cell in row:
            cell.font = bold_font if cell.row <= 3 or cell.row in emphasized_rows else font
            cell.alignment = align
            cell.border = border

    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = True
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    for idx in range(3, total_col):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 10
    ws.column_dimensions[ws.cell(row=1, column=total_col).column_letter].width = 12

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def manifest_upload_workbook_bytes(payload: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
    except ModuleNotFoundError:
        if CODEX_PYTHON_SITE_PACKAGES.exists():
            sys.path.insert(0, str(CODEX_PYTHON_SITE_PACKAGES))
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, Side
        else:
            raise

    stores = payload.get("stores") or []
    rows = payload.get("rows") or []
    if not isinstance(stores, list) or not isinstance(rows, list):
        raise ValueError("stores 與 rows 必須是陣列")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title("匯出上傳")

    font = Font(name="Microsoft JhengHei", size=12)
    bold_font = Font(name="Microsoft JhengHei", size=12, bold=True)
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DCE2DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["門市代碼", "呼出碼", "數量"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
        cell.alignment = align
        cell.border = border

    current_row = 2
    for row in rows:
        product_code = str(row.get("productCode") or "").strip()
        quantities = row.get("quantities") or {}
        for store in stores:
            store_code = str(store.get("storeCode") or store.get("sheet") or "").strip()
            quantity = quantities.get(store.get("key"))
            try:
                numeric_quantity = float(quantity or 0)
            except (TypeError, ValueError):
                numeric_quantity = 0
            if not store_code or not product_code or not numeric_quantity:
                continue
            ws.cell(row=current_row, column=1, value=store_code)
            ws.cell(row=current_row, column=2, value=product_code)
            ws.cell(row=current_row, column=3, value=numeric_quantity)
            current_row += 1

    last_row = max(1, current_row - 1)
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=3):
        for cell in row:
            cell.font = font
            cell.alignment = align
            cell.border = border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def harvest_shipment_workbook_bytes(payload: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError:
        if CODEX_PYTHON_SITE_PACKAGES.exists():
            sys.path.insert(0, str(CODEX_PYTHON_SITE_PACKAGES))
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        else:
            raise

    shipment_date = str(payload.get("shipmentDate") or "").strip()
    columns = payload.get("columns") or []
    rows = payload.get("rows") or []
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", shipment_date):
        raise ValueError("shipmentDate 格式錯誤")
    if not isinstance(columns, list) or not isinstance(rows, list):
        raise ValueError("columns 與 rows 必須是陣列")

    headers = [str(column.get("label") or "").strip() for column in columns if isinstance(column, dict)]
    if not headers or headers[0] != "品項":
        raise ValueError("匯出欄位必須從品項開始")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(f"{shipment_date}_到貨量")

    font = Font(name="Microsoft JhengHei", size=12)
    bold_font = Font(name="Microsoft JhengHei", size=12, bold=True)
    header_fill = PatternFill("solid", fgColor="F7F7EF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = align
        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        values = row.get("values") if isinstance(row, dict) else []
        if not isinstance(values, list):
            values = []
        for col, value in enumerate(values[: len(headers)], start=1):
            cell_value = None if value in (None, "") else value
            if isinstance(cell_value, (int, float)):
                ws.cell(row=row_index, column=col, value=float(cell_value) if cell_value % 1 else int(cell_value))
            else:
                ws.cell(row=row_index, column=col, value=str(cell_value) if cell_value is not None else None)
            cell = ws.cell(row=row_index, column=col)
            cell.font = bold_font if col == 1 else font
            cell.alignment = align
            cell.border = border

    last_row = max(1, len(rows) + 1)
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = align
            cell.border = border

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = True
    for col in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letter].width = 14 if col == 1 else 11

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def excel_fill_color(value: object) -> str:
    color = str(value or "").strip().lstrip("#")
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", color):
        return ""
    return f"FF{color.upper()}"


def large_label_cell_value(value: object) -> object:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) if float(value) % 1 else int(value)
    return str(value)


def harvest_large_label_workbook_bytes(payload: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError:
        if CODEX_PYTHON_SITE_PACKAGES.exists():
            sys.path.insert(0, str(CODEX_PYTHON_SITE_PACKAGES))
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        else:
            raise

    harvest_date = str(payload.get("harvestDate") or "").strip()
    rows = payload.get("rows") or []
    merges = payload.get("merges") or []
    column_widths = payload.get("columnWidths") or []
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
        raise ValueError("harvestDate 格式錯誤")
    if not isinstance(rows, list) or not rows:
        raise ValueError("rows 必須是非空陣列")
    if not isinstance(merges, list) or not isinstance(column_widths, list):
        raise ValueError("merges 與 columnWidths 必須是陣列")

    max_cols = max((len(row) for row in rows if isinstance(row, list)), default=0)
    if max_cols < 1:
        raise ValueError("rows 沒有可匯出的欄位")
    if len(rows) > 500 or max_cols > 120:
        raise ValueError("大標表格超出可匯出範圍")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(f"{harvest_date}_大標")

    number_format = '#,##0;[Red](#,##0)'
    font = Font(name="Microsoft JhengHei", size=11)
    bold_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    red_font = Font(name="Microsoft JhengHei", size=11, color="FFC84C31")
    bold_red_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFC84C31")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_index, row in enumerate(rows, start=1):
        if not isinstance(row, list):
            continue
        for col_index, item in enumerate(row[:max_cols], start=1):
            item = item if isinstance(item, dict) else {"value": item}
            value = large_label_cell_value(item.get("value"))
            is_number = isinstance(value, (int, float)) and not isinstance(value, bool)
            is_negative_number = is_number and value < 0
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if is_negative_number:
                cell.font = bold_red_font if item.get("bold") else red_font
            else:
                cell.font = bold_font if item.get("bold") else font
            if is_number:
                cell.number_format = number_format
            cell.alignment = align
            cell.border = border
            color = excel_fill_color(item.get("backgroundColor"))
            if color:
                cell.fill = PatternFill("solid", fgColor=color)

    for merge in merges:
        if not isinstance(merge, dict):
            continue
        try:
            start_row = int(merge.get("startRow") or 0)
            start_col = int(merge.get("startCol") or 0)
            end_row = int(merge.get("endRow") or 0)
            end_col = int(merge.get("endCol") or 0)
        except (TypeError, ValueError):
            continue
        if start_row < 1 or start_col < 1 or end_row < start_row or end_col < start_col:
            continue
        if end_row > len(rows) or end_col > max_cols:
            continue
        if start_row == end_row and start_col == end_col:
            continue
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=max_cols):
        for cell in row:
            cell.alignment = align
            cell.border = border

    for col_index in range(1, max_cols + 1):
        width = column_widths[col_index - 1] if col_index <= len(column_widths) else None
        try:
            width = float(width)
        except (TypeError, ValueError):
            width = 11
        width = min(max(width, 8), 24)
        ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = width

    for row_index in range(1, len(rows) + 1):
        ws.row_dimensions[row_index].height = 26 if row_index <= 2 else 22
    ws.freeze_panes = str(payload.get("freezePane") or "B3")
    ws.sheet_view.showGridLines = True

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def bootstrap_db() -> None:
    conn = connect()
    try:
        init_db(conn)
        if not inventory_sync_version(conn):
            mark_inventory_sync_changed(conn)
        seed_default_users(conn)
        if get_current_snapshot(conn) is None:
            snapshot = read_local_snapshot()
            if snapshot:
                set_current_snapshot(conn, snapshot)
    finally:
        conn.close()


class AppHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(DOCS), **kwargs)

    def cache_control_for_path(self, request_path: str) -> str:
        path = urlparse(request_path).path
        if path.startswith("/api/") or path in {"/data.json", "/data.js"}:
            return "no-store"
        if path in {"/", "/index.html", "/manifest.json", "/sw.js"}:
            return "no-cache, no-store, must-revalidate"
        if path.startswith(f"{FIELD_WORK_MESSAGE_PHOTO_URL_PREFIX}/"):
            return ""
        suffix = Path(path).suffix.lower()
        if suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".ico"}:
            return "public, max-age=31536000, immutable"
        if suffix in {".js", ".css"} and urlparse(request_path).query:
            if APP_VERSION.startswith("local-dev"):
                return "no-cache, no-store, must-revalidate"
            return "public, max-age=31536000, immutable"
        return "no-cache"

    def end_headers(self) -> None:
        cache_control = self.cache_control_for_path(self.path)
        if cache_control:
            self.send_header("Cache-Control", cache_control)
            if "no-cache" in cache_control or "no-store" in cache_control:
                self.send_header("Pragma", "no-cache")
                self.send_header("Expires", "0")
        super().end_headers()

    def send_versioned_text_file(self, file_path: Path, content_type: str, include_body: bool = True) -> None:
        try:
            content = file_path.read_text(encoding="utf-8")
        except OSError:
            self.send_error(404, "File not found")
            return
        body = content.replace("__APP_VERSION__", APP_VERSION_QUERY).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        if include_body:
            self.wfile.write(body)

    def serve_index(self, include_body: bool = True) -> None:
        self.send_versioned_text_file(DOCS / "index.html", "text/html; charset=utf-8", include_body=include_body)

    def serve_service_worker(self, include_body: bool = True) -> None:
        self.send_versioned_text_file(DOCS / "sw.js", "text/javascript; charset=utf-8", include_body=include_body)

    def session_token(self) -> str:
        cookie_header = self.headers.get("Cookie", "")
        if not cookie_header:
            return ""
        cookie = SimpleCookie()
        try:
            cookie.load(cookie_header)
        except Exception:
            return ""
        morsel = cookie.get(AUTH_COOKIE_NAME)
        return morsel.value if morsel else ""

    def current_user(self) -> dict[str, Any] | None:
        if hasattr(self, "_current_user"):
            return self._current_user
        token = self.session_token()
        if not token:
            self._current_user = None
            return None
        conn = connect()
        try:
            row = conn.execute(
                """
                SELECT users.id, users.username, users.display_name, users.role
                FROM auth_sessions
                JOIN users ON users.id = auth_sessions.user_id
                WHERE auth_sessions.token = ?
                    AND auth_sessions.expires_at > ?
                    AND users.is_active = 1
                """,
                (token, taipei_now_text()),
            ).fetchone()
            if not row:
                conn.execute("DELETE FROM auth_sessions WHERE token = ? OR expires_at <= ?", (token, taipei_now_text()))
                conn.commit()
                self._current_user = None
                return None
            self._current_user = user_payload(row)
            return self._current_user
        finally:
            conn.close()

    def require_roles(self, roles: set[str]) -> dict[str, Any] | None:
        user = self.current_user()
        if not user:
            self.send_json({"ok": False, "error": "請先登入"}, status=401)
            return None
        if user["role"] == "root" or user["role"] in roles:
            return user
        self.send_json({"ok": False, "error": "此帳號沒有使用這個功能的權限"}, status=403)
        return None

    def allowed_api_roles(self, method: str, path: str) -> set[str] | None:
        if path in {"/api/login", "/api/logout"}:
            return None
        if path == "/api/bootstrap":
            return ANY_AUTH_ROLES
        if path == "/api/users" or path.startswith("/api/users/"):
            return ROOT_API_ROLES
        if method == "GET":
            if path in {"/api/sales", "/api/opening-stock", "/api/inventory-sync", "/api/harvest-messages"}:
                return INSIDE_API_ROLES
            if path == "/api/field-work-messages":
                return FIELD_API_ROLES
        if method == "POST":
            if path in {
                "/api/harvest-messages",
                "/api/import-sales",
                "/api/import-sales-from-website",
                "/api/stores",
                "/api/export-shipment-manifest",
                "/api/export-shipment-upload",
                "/api/export-harvest-shipment",
                "/api/export-harvest-large-label",
            }:
                return INSIDE_API_ROLES
            if path in {
                "/api/field-work-messages",
            }:
                return FIELD_API_ROLES
            if path in {
                "/api/daily-vegetable-manual-entries",
                "/api/export-daily-vegetable",
                "/api/export-field-work-month",
                "/api/export-field-bt-month-summary",
            }:
                return ROOT_API_ROLES
        if method == "PUT":
            if path in {
                "/api/shipments",
                "/api/stock-adjustments",
                "/api/product-filter",
                "/api/harvest-product-filter",
                "/api/harvest-crew-columns",
                "/api/harvest-field-extra-columns",
                "/api/harvest-field-3f-expanded",
                "/api/harvest-conversion-settings",
                "/api/shipment-manifest-selection",
                "/api/harvest-entries",
                "/api/city-table-entries",
                "/api/yfy-table-entries",
                "/api/loose-vegetable-table",
                "/api/general-channel-table",
            } or path.startswith("/api/store-notes/") or path.startswith("/api/store-routes/") or path.startswith("/api/stores/"):
                return INSIDE_API_ROLES
            if path in {"/api/field-bt-records", "/api/field-work-records"}:
                return FIELD_API_ROLES
            if path in {"/api/daily-vegetable-product-filter", "/api/net-house-status-records"} or path.startswith("/api/daily-vegetable-manual-entries/"):
                return ROOT_API_ROLES
        if method == "DELETE":
            if path.startswith("/api/stores/") or path.startswith("/api/harvest-messages/"):
                return INSIDE_API_ROLES
            if path.startswith("/api/field-work-messages/"):
                return FIELD_API_ROLES
            if path.startswith("/api/daily-vegetable-manual-entries/"):
                return ROOT_API_ROLES
        return ANY_AUTH_ROLES

    def check_api_access(self, method: str, path: str) -> bool:
        roles = self.allowed_api_roles(method, path)
        if roles is None:
            return True
        return self.require_roles(roles) is not None

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path in {"/", "/index.html"}:
            self.serve_index()
            return
        if parsed.path == "/sw.js":
            self.serve_service_worker()
            return
        if parsed.path.startswith(f"{FIELD_WORK_MESSAGE_PHOTO_URL_PREFIX}/"):
            if not self.require_roles(FIELD_API_ROLES):
                return
            self.serve_field_work_message_photo(parsed.path)
            return
        if parsed.path in {"/data.json", "/data.js"}:
            if not self.require_roles(ROOT_API_ROLES):
                return
        if parsed.path.startswith("/api/") and not self.check_api_access("GET", parsed.path):
            return
        if parsed.path == "/api/bootstrap":
            self.handle_bootstrap()
            return
        if parsed.path == "/api/users":
            self.handle_list_users()
            return
        if parsed.path == "/api/sales":
            self.handle_sales(parsed)
            return
        if parsed.path == "/api/opening-stock":
            self.handle_opening_stock(parsed)
            return
        if parsed.path == "/api/inventory-sync":
            self.handle_inventory_sync(parsed)
            return
        if parsed.path == "/api/harvest-messages":
            self.handle_harvest_messages(parsed)
            return
        if parsed.path == "/api/field-work-messages":
            self.handle_field_work_messages(parsed)
            return
        super().do_GET()

    def do_HEAD(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path in {"/", "/index.html"}:
            self.serve_index(include_body=False)
            return
        if parsed.path == "/sw.js":
            self.serve_service_worker(include_body=False)
            return
        if parsed.path.startswith(f"{FIELD_WORK_MESSAGE_PHOTO_URL_PREFIX}/"):
            if not self.require_roles(FIELD_API_ROLES):
                return
            self.serve_field_work_message_photo(parsed.path, include_body=False)
            return
        if parsed.path in {"/data.json", "/data.js"} and not self.require_roles(ROOT_API_ROLES):
            return
        super().do_HEAD()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/") and not self.check_api_access("POST", parsed.path):
            return
        if self.path == "/api/login":
            self.handle_login()
            return
        if self.path == "/api/logout":
            self.handle_logout()
            return
        if self.path == "/api/users":
            self.handle_create_user()
            return
        if self.path == "/api/harvest-messages":
            self.handle_add_harvest_message()
            return
        if self.path == "/api/field-work-messages":
            self.handle_add_field_work_message()
            return
        if self.path == "/api/import-sales":
            self.handle_import_sales()
            return
        if self.path == "/api/import-sales-from-website":
            self.handle_import_sales_from_website()
            return
        if self.path == "/api/stores":
            self.handle_add_store()
            return
        if self.path == "/api/export-shipment-manifest":
            self.handle_export_shipment_manifest()
            return
        if self.path == "/api/export-shipment-upload":
            self.handle_export_shipment_upload()
            return
        if self.path == "/api/export-harvest-shipment":
            self.handle_export_harvest_shipment()
            return
        if self.path == "/api/export-harvest-large-label":
            self.handle_export_harvest_large_label()
            return
        if self.path == "/api/export-field-work-month":
            self.handle_export_field_work_month()
            return
        if self.path == "/api/export-field-bt-month-summary":
            self.handle_export_field_bt_month_summary()
            return
        if self.path == "/api/export-daily-vegetable":
            self.handle_export_daily_vegetable()
            return
        if self.path == "/api/daily-vegetable-manual-entries":
            self.handle_add_daily_vegetable_manual_entry()
            return
        self.send_error(404, "Not found")

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/") and not self.check_api_access("DELETE", parsed.path):
            return
        if parsed.path.startswith("/api/stores/"):
            store_key = unquote(parsed.path.split("/api/stores/", 1)[1])
            self.handle_delete_store(store_key)
            return
        if parsed.path.startswith("/api/users/"):
            user_id = unquote(parsed.path.split("/api/users/", 1)[1])
            self.handle_delete_user(user_id)
            return
        if parsed.path.startswith("/api/harvest-messages/"):
            message_id = unquote(parsed.path.split("/api/harvest-messages/", 1)[1])
            self.handle_delete_harvest_message(message_id, parsed)
            return
        if parsed.path.startswith("/api/field-work-messages/"):
            message_id = unquote(parsed.path.split("/api/field-work-messages/", 1)[1])
            self.handle_delete_field_work_message(message_id, parsed)
            return
        if parsed.path.startswith("/api/daily-vegetable-manual-entries/"):
            entry_id = unquote(parsed.path.split("/api/daily-vegetable-manual-entries/", 1)[1])
            self.handle_delete_daily_vegetable_manual_entry(entry_id)
            return
        self.send_error(404, "Not found")

    def do_PUT(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/") and not self.check_api_access("PUT", parsed.path):
            return
        if parsed.path.startswith("/api/users/"):
            user_id = unquote(parsed.path.split("/api/users/", 1)[1])
            self.handle_update_user(user_id)
            return
        if parsed.path == "/api/shipments":
            self.handle_save_shipments()
            return
        if parsed.path == "/api/stock-adjustments":
            self.handle_save_stock_adjustments()
            return
        if parsed.path == "/api/product-filter":
            self.handle_save_product_filter()
            return
        if parsed.path == "/api/harvest-product-filter":
            self.handle_save_harvest_product_filter()
            return
        if parsed.path == "/api/daily-vegetable-product-filter":
            self.handle_save_daily_vegetable_product_filter()
            return
        if parsed.path == "/api/harvest-crew-columns":
            self.handle_save_harvest_crew_columns()
            return
        if parsed.path == "/api/harvest-field-extra-columns":
            self.handle_save_harvest_field_extra_columns()
            return
        if parsed.path == "/api/harvest-field-3f-expanded":
            self.handle_save_harvest_field_3f_expanded()
            return
        if parsed.path == "/api/harvest-conversion-settings":
            self.handle_save_harvest_conversion_settings()
            return
        if parsed.path == "/api/shipment-manifest-selection":
            self.handle_save_shipment_manifest_selection()
            return
        if parsed.path == "/api/harvest-entries":
            self.handle_save_harvest_entries()
            return
        if parsed.path == "/api/city-table-entries":
            self.handle_save_city_table_entries()
            return
        if parsed.path == "/api/yfy-table-entries":
            self.handle_save_yfy_table_entries()
            return
        if parsed.path == "/api/loose-vegetable-table":
            self.handle_save_loose_vegetable_table()
            return
        if parsed.path == "/api/general-channel-table":
            self.handle_save_general_channel_table()
            return
        if parsed.path == "/api/field-bt-records":
            self.handle_save_field_bt_records()
            return
        if parsed.path == "/api/field-work-records":
            self.handle_save_field_work_records()
            return
        if parsed.path == "/api/net-house-status-records":
            self.handle_save_net_house_status_record()
            return
        if parsed.path.startswith("/api/daily-vegetable-manual-entries/"):
            entry_id = unquote(parsed.path.split("/api/daily-vegetable-manual-entries/", 1)[1])
            self.handle_update_daily_vegetable_manual_entry(entry_id)
            return
        if parsed.path.startswith("/api/stores/"):
            store_key = unquote(parsed.path.split("/api/stores/", 1)[1])
            self.handle_update_store(store_key)
            return
        if parsed.path.startswith("/api/store-notes/"):
            store_key = unquote(parsed.path.split("/api/store-notes/", 1)[1])
            self.handle_save_store_note(store_key)
            return
        if parsed.path.startswith("/api/store-routes/"):
            store_key = unquote(parsed.path.split("/api/store-routes/", 1)[1])
            self.handle_save_store_route(store_key)
            return
        self.send_error(404, "Not found")

    def auth_cookie_header(self, token: str, max_age: int) -> str:
        secure = "; Secure" if AUTH_COOKIE_SECURE else ""
        return (
            f"{AUTH_COOKIE_NAME}={token}; Path=/; HttpOnly; SameSite=Lax; "
            f"Max-Age={max_age}{secure}"
        )

    def clear_auth_cookie_header(self) -> str:
        return f"{AUTH_COOKIE_NAME}=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0"

    def handle_login(self) -> None:
        try:
            payload = read_json_body(self)
            username = str(payload.get("username") or "").strip()
            password = str(payload.get("password") or "")
            if not username or not password:
                raise ValueError("請輸入帳號與密碼")
            conn = connect()
            try:
                init_db(conn)
                user_row = conn.execute(
                    """
                    SELECT id, username, display_name, password_hash, role, is_active
                    FROM users
                    WHERE username = ?
                    """,
                    (username,),
                ).fetchone()
                if (
                    not user_row
                    or not int(user_row["is_active"])
                    or not password_matches(password, user_row["password_hash"])
                ):
                    self.send_json({"ok": False, "error": "帳號或密碼錯誤"}, status=401)
                    return
                token = secrets.token_urlsafe(32)
                created_at = taipei_now_text()
                expires_at = (taipei_now() + timedelta(days=AUTH_SESSION_DAYS)).isoformat(timespec="seconds")
                conn.execute(
                    """
                    INSERT INTO auth_sessions (token, user_id, created_at, expires_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    (token, int(user_row["id"]), created_at, expires_at),
                )
                conn.execute("DELETE FROM auth_sessions WHERE expires_at <= ?", (taipei_now_text(),))
                conn.commit()
                public_user = user_payload(user_row)
            finally:
                conn.close()
            self.send_json(
                {"ok": True, "user": public_user},
                headers=[("Set-Cookie", self.auth_cookie_header(token, AUTH_SESSION_DAYS * 24 * 60 * 60))],
            )
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def handle_logout(self) -> None:
        token = self.session_token()
        if token:
            conn = connect()
            try:
                init_db(conn)
                conn.execute("DELETE FROM auth_sessions WHERE token = ?", (token,))
                conn.commit()
            finally:
                conn.close()
        self.send_json(
            {"ok": True},
            headers=[("Set-Cookie", self.clear_auth_cookie_header())],
        )

    def handle_list_users(self) -> None:
        try:
            conn = connect()
            try:
                init_db(conn)
                rows = conn.execute(
                    """
                    SELECT id, username, display_name, role, is_active, created_at, updated_at
                    FROM users
                    ORDER BY username
                    """
                ).fetchall()
            finally:
                conn.close()
            self.send_json({"ok": True, "users": [listed_user_payload(row) for row in rows]})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_create_user(self) -> None:
        try:
            payload = read_json_body(self)
            username = validate_username(str(payload.get("username") or ""))
            display_name = validate_display_name(str(payload.get("displayName") or ""))
            password = validate_password(str(payload.get("password") or ""))
            role = validate_role(str(payload.get("role") or ""))
            timestamp = taipei_now_text()
            conn = connect()
            try:
                init_db(conn)
                cursor = conn.execute(
                    """
                    INSERT INTO users (username, display_name, password_hash, role, is_active, created_at, updated_at)
                    VALUES (?, ?, ?, ?, 1, ?, ?)
                    """,
                    (username, display_name, password_hash(password), role, timestamp, timestamp),
                )
                conn.commit()
                row = conn.execute(
                    """
                    SELECT id, username, display_name, role, is_active, created_at, updated_at
                    FROM users
                    WHERE id = ?
                    """,
                    (cursor.lastrowid,),
                ).fetchone()
            finally:
                conn.close()
            self.send_json({"ok": True, "user": listed_user_payload(row)})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def would_remove_last_active_root(self, conn, target_user_id: int, next_role: str, next_is_active: bool) -> bool:
        rows = conn.execute(
            "SELECT id, role, is_active FROM users"
        ).fetchall()
        active_roots = 0
        for row in rows:
            role = next_role if int(row["id"]) == target_user_id else str(row["role"] or "")
            is_active = next_is_active if int(row["id"]) == target_user_id else bool(row["is_active"])
            if role == "root" and is_active:
                active_roots += 1
        return active_roots < 1

    def handle_update_user(self, user_id: str) -> None:
        try:
            try:
                normalized_user_id = int(user_id)
            except ValueError as exc:
                raise ValueError("userId 格式錯誤") from exc
            payload = read_json_body(self)
            display_name = validate_display_name(str(payload.get("displayName") or ""))
            role = validate_role(str(payload.get("role") or ""))
            is_active = bool(payload.get("isActive"))
            password = str(payload.get("password") or "")
            requested_username = payload.get("username")
            current_token = self.session_token()
            conn = connect()
            try:
                init_db(conn)
                existing = conn.execute(
                    """
                    SELECT id, username, display_name, role, is_active, created_at, updated_at
                    FROM users
                    WHERE id = ?
                    """,
                    (normalized_user_id,),
                ).fetchone()
                if not existing:
                    raise ValueError("找不到帳號")
                username = validate_username(
                    str(requested_username if requested_username is not None else existing["username"])
                )
                duplicate = conn.execute(
                    "SELECT id FROM users WHERE username = ? AND id != ?",
                    (username, normalized_user_id),
                ).fetchone()
                if duplicate:
                    raise ValueError("帳號 ID 已存在")
                if self.would_remove_last_active_root(conn, normalized_user_id, role, is_active):
                    raise ValueError("至少需要保留一個啟用中的管理員帳號")
                timestamp = taipei_now_text()
                if password:
                    validate_password(password)
                    conn.execute(
                        """
                        UPDATE users
                        SET username = ?, display_name = ?, role = ?, is_active = ?, password_hash = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            username,
                            display_name,
                            role,
                            1 if is_active else 0,
                            password_hash(password),
                            timestamp,
                            normalized_user_id,
                        ),
                    )
                    conn.execute(
                        "DELETE FROM auth_sessions WHERE user_id = ? AND token != ?",
                        (normalized_user_id, current_token),
                    )
                else:
                    conn.execute(
                        """
                        UPDATE users
                        SET username = ?, display_name = ?, role = ?, is_active = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (username, display_name, role, 1 if is_active else 0, timestamp, normalized_user_id),
                    )
                    if not is_active:
                        conn.execute("DELETE FROM auth_sessions WHERE user_id = ?", (normalized_user_id,))
                conn.commit()
                row = conn.execute(
                    """
                    SELECT id, username, display_name, role, is_active, created_at, updated_at
                    FROM users
                    WHERE id = ?
                    """,
                    (normalized_user_id,),
                ).fetchone()
            finally:
                conn.close()
            self.send_json({"ok": True, "user": listed_user_payload(row)})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def handle_delete_user(self, user_id: str) -> None:
        try:
            try:
                normalized_user_id = int(user_id)
            except ValueError as exc:
                raise ValueError("userId 格式錯誤") from exc
            conn = connect()
            try:
                init_db(conn)
                existing = conn.execute(
                    """
                    SELECT id, username, display_name, role, is_active, created_at, updated_at
                    FROM users
                    WHERE id = ?
                    """,
                    (normalized_user_id,),
                ).fetchone()
                if not existing:
                    raise ValueError("找不到帳號")
                current_user = self.current_user()
                if current_user and int(current_user.get("id") or 0) == normalized_user_id:
                    raise ValueError("不能刪除目前登入中的帳號")
                if self.would_remove_last_active_root(conn, normalized_user_id, str(existing["role"] or ""), False):
                    raise ValueError("至少需要保留一個啟用中的管理員帳號")
                conn.execute("DELETE FROM auth_sessions WHERE user_id = ?", (normalized_user_id,))
                conn.execute("DELETE FROM users WHERE id = ?", (normalized_user_id,))
                conn.commit()
            finally:
                conn.close()
            self.send_json({"ok": True, "deleted": listed_user_payload(existing)})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def app_data_for_role(self, app_data: dict, role: str) -> dict:
        if role != "field":
            return app_data
        return {
            "generatedAt": app_data.get("generatedAt", ""),
            "salesDate": app_data.get("salesDate", ""),
            "harvestPlanningTemplate": app_data.get("harvestPlanningTemplate"),
            "products": [],
            "stores": [],
            "storePages": [],
            "inventory": [],
            "lowStock": [],
            "salesRows": [],
            "salesPivot": [],
            "salesTotals": {"salesQty": 0.0, "discardQty": 0.0, "endingStock": 0.0},
            "salesByProduct": [],
            "discardByProduct": [],
        }

    def bootstrap_payload_for_role(self, payload: dict, user: dict[str, Any]) -> dict:
        role = user["role"]
        payload = {**payload, "currentUser": user}
        payload["appData"] = self.app_data_for_role(payload.get("appData") or {}, role)
        if role == "root":
            return payload
        if role == "inside":
            payload["fieldNetHouses"] = {"zones": []}
            payload["fieldWorkRecords"] = []
            payload["fieldBtRecords"] = []
            payload["netHouseStatusRecords"] = []
            payload["visibleDailyVegetableProductKeys"] = []
            payload["dailyVegetableManualEntries"] = []
            return payload
        if role == "field":
            allowed = {
                "ok",
                "currentUser",
                "appData",
                "fieldNetHouses",
                "fieldWorkRecords",
                "fieldBtRecords",
                "netHouseStatusRecords",
            }
            filtered = {key: value for key, value in payload.items() if key in allowed}
            filtered.setdefault("fieldNetHouses", {"zones": []})
            filtered.setdefault("fieldWorkRecords", [])
            filtered.setdefault("fieldBtRecords", [])
            filtered.setdefault("netHouseStatusRecords", [])
            filtered.update(
                {
                    "storeNotes": {},
                    "storeRoutes": {},
                    "shipmentEntries": [],
                    "stockAdjustmentEntries": [],
                    "harvestEntries": [],
                    "harvestCellPriorities": {},
                    "harvestCellFormulas": {},
                    "visibleProductCodes": [],
                    "visibleProductCodesByDate": {},
                    "visibleHarvestProductKeys": [],
                    "visibleHarvestProductKeysByDate": {},
                    "visibleDailyVegetableProductKeys": [],
                    "dailyVegetableManualEntries": [],
                    "harvestCrewColumns": [],
                    "harvestCrewColumnsByDate": {},
                    "harvestFieldExtraColumnsByDate": {},
                    "harvestField3FExpandedByDate": {},
                    "shipmentManifestSelections": {},
                    "cityTableEntries": {},
                    "yfyTableEntries": {},
                    "yfyShipmentTimes": {},
                    "looseVegetableColumns": [],
                    "looseVegetableColumnsByDate": {},
                    "looseVegetableTableEntries": {},
                    "generalChannelColumns": [],
                    "generalChannelColumnsByDate": {},
                    "generalChannelTableEntries": {},
                    "generalChannelPuqianEntries": {},
                }
            )
            return filtered
        return payload

    def handle_bootstrap(self) -> None:
        user = self.current_user()
        if not user:
            self.send_json({"ok": False, "error": "請先登入"}, status=401)
            return
        conn = connect()
        try:
            payload = current_snapshot_payload(conn)
            response_payload = {
                    "ok": True,
                    "appData": payload,
                    "inventorySyncVersion": inventory_sync_version(conn),
                    "storeNotes": get_store_notes(conn),
                    "storeRoutes": get_store_routes(conn),
                    "shipmentEntries": get_shipment_entries(conn),
                    "stockAdjustmentEntries": get_stock_adjustment_entries(conn),
                    "harvestEntries": get_harvest_entries(conn),
                    "harvestCellPriorities": get_setting(conn, "harvestCellPriorities") or {},
                    "harvestCellFormulas": get_setting(conn, "harvestCellFormulas") or {},
                    "visibleProductCodes": get_setting(conn, "visibleProductCodes"),
                    "visibleProductCodesByDate": get_setting(conn, "visibleProductCodesByDate") or {},
                    "visibleHarvestProductKeys": get_setting(conn, "visibleHarvestProductKeys"),
                    "visibleHarvestProductKeysByDate": get_setting(conn, "visibleHarvestProductKeysByDate") or {},
                    "visibleDailyVegetableProductKeys": get_setting(conn, "visibleDailyVegetableProductKeys"),
                    "harvestCrewColumns": get_setting(conn, "harvestCrewColumns"),
                    "harvestCrewColumnsByDate": get_setting(conn, "harvestCrewColumnsByDate") or {},
                    "harvestFieldExtraColumnsByDate": get_setting(conn, "harvestFieldExtraColumnsByDate") or {},
                    "harvestField3FExpandedByDate": get_setting(conn, "harvestField3FExpandedByDate") or {},
                    "harvestConversionSettings": get_setting(conn, "harvestConversionSettings") or {},
                    "shipmentManifestSelections": get_setting(conn, "shipmentManifestSelections") or {},
                    "cityTableEntries": get_setting(conn, "cityTableEntries") or {},
                    "yfyTableEntries": get_setting(conn, "yfyTableEntries") or {},
                    "yfyShipmentTimes": get_setting(conn, "yfyShipmentTimes") or {},
                    "looseVegetableColumns": get_setting(conn, "looseVegetableColumns") or [],
                    "looseVegetableColumnsByDate": get_setting(conn, "looseVegetableColumnsByDate") or {},
                    "looseVegetableTableEntries": get_setting(conn, "looseVegetableTableEntries") or {},
                    "generalChannelColumns": get_setting(conn, "generalChannelColumns") or [],
                    "generalChannelColumnsByDate": get_setting(conn, "generalChannelColumnsByDate") or {},
                    "generalChannelTableEntries": get_setting(conn, "generalChannelTableEntries") or {},
                    "generalChannelPuqianEntries": get_setting(conn, "generalChannelPuqianEntries") or {},
                    "fieldNetHouses": get_field_net_house_data(conn),
                    "fieldWorkRecords": get_field_work_records(conn),
                    "fieldBtRecords": get_field_bt_records(conn),
                    "netHouseStatusRecords": get_net_house_status_records(conn),
                    "dailyVegetableManualEntries": get_daily_vegetable_manual_entries(conn),
                }
            self.send_json(self.bootstrap_payload_for_role(response_payload, user))
        finally:
            conn.close()

    def handle_sales(self, parsed) -> None:
        try:
            params = parse_qs(parsed.query)
            start_date = params.get("start", [""])[0] or None
            end_date = params.get("end", [""])[0] or None
            store_code = params.get("storeCode", [""])[0] or None
            payload = sales_payload(start_date, end_date, store_code)
            self.send_json({"ok": True, **payload})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_opening_stock(self, parsed) -> None:
        try:
            params = parse_qs(parsed.query)
            month = params.get("month", [""])[0] or None
            store_code = params.get("storeCode", [""])[0] or None
            payload = opening_stock_payload(month, store_code)
            self.send_json({"ok": True, **payload})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_inventory_sync(self, parsed) -> None:
        try:
            params = parse_qs(parsed.query)
            since = str((params.get("since") or [""])[0]).strip() or None
            conn = connect()
            try:
                payload = inventory_sync_payload(conn, since)
            finally:
                conn.close()
            self.send_json({"ok": True, **payload})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_import_sales(self) -> None:
        try:
            saved_path, sales_date = self.parse_import_form()
            result = self.import_sales_workbook(saved_path, sales_date)
            self.send_json({"ok": True, "file": saved_path.name, "salesDate": sales_date, "output": result.stdout})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_import_sales_from_website(self) -> None:
        try:
            payload = read_json_body(self)
            sales_date = validate_sales_date(str(payload.get("salesDate") or ""))
            saved_path = download_sales_workbook(sales_date)
            result = self.import_sales_workbook(saved_path, sales_date)
            self.send_json({"ok": True, "file": saved_path.name, "salesDate": sales_date, "output": result.stdout})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def import_sales_workbook(self, saved_path: Path, sales_date: str) -> subprocess.CompletedProcess[str]:
        env = {**os.environ, "SALES_FILE_PATH": str(saved_path), "SALES_DATE": sales_date, "IMPORT_SALES_ONLY": "1"}
        try:
            result = subprocess.run(
                [PYTHON, str(ROOT / "scripts" / "build_site_data.py")],
                cwd=str(ROOT),
                env=env,
                text=True,
                capture_output=True,
                check=True,
            )
        finally:
            saved_path.unlink(missing_ok=True)
        conn = connect()
        try:
            snapshot = current_snapshot_payload(conn)
            if not snapshot:
                raise ValueError("匯入後找不到資料快照")
            set_current_snapshot(conn, lightweight_sales_snapshot(snapshot, sales_date))
        finally:
            conn.close()
        return result

    def handle_add_store(self) -> None:
        try:
            payload = read_json_body(self)
            conn = connect()
            try:
                snapshot = current_snapshot_payload(conn)
                next_snapshot, store = add_store_to_snapshot(snapshot, payload)
                set_current_snapshot(conn, next_snapshot)
                save_store_route(conn, store["key"], str(payload.get("route") or ""))
            finally:
                conn.close()
            self.send_json({"ok": True, "appData": next_snapshot, "store": store})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_update_store(self, store_key: str) -> None:
        try:
            payload = read_json_body(self)
            conn = connect()
            try:
                snapshot = current_snapshot_payload(conn)
                next_snapshot, store = update_store_name_in_snapshot(
                    snapshot,
                    store_key,
                    str(payload.get("storeName") or ""),
                )
                set_current_snapshot(conn, next_snapshot)
            finally:
                conn.close()
            self.send_json({"ok": True, "appData": next_snapshot, "store": store})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_delete_store(self, store_key: str) -> None:
        try:
            if not store_key:
                raise ValueError("缺少 storeKey")
            conn = connect()
            try:
                snapshot = current_snapshot_payload(conn)
                next_snapshot, removed = delete_store_from_snapshot(snapshot, store_key)
                set_current_snapshot(conn, next_snapshot)
                conn.execute("DELETE FROM store_notes WHERE store_key = ?", (store_key,))
                conn.execute("DELETE FROM store_routes WHERE store_key = ?", (store_key,))
                conn.execute("DELETE FROM shipment_entries WHERE store_key = ?", (store_key,))
                conn.execute("DELETE FROM stock_adjustment_entries WHERE store_key = ?", (store_key,))
                selections = get_setting(conn, "shipmentManifestSelections") or {}
                if isinstance(selections, dict):
                    cleaned = {}
                    for key, value in selections.items():
                        if isinstance(value, list):
                            cleaned[key] = [item for item in value if str(item) != store_key]
                        elif isinstance(value, dict):
                            next_value = dict(value)
                            selected = next_value.get("selectedStoreKeys")
                            if isinstance(selected, list):
                                next_value["selectedStoreKeys"] = [item for item in selected if str(item) != store_key]
                            known_available = next_value.get("knownAvailableStoreKeys")
                            if isinstance(known_available, list):
                                next_value["knownAvailableStoreKeys"] = [
                                    item for item in known_available if str(item) != store_key
                                ]
                            rounds = next_value.get("rounds")
                            if isinstance(rounds, list):
                                cleaned_rounds = []
                                for round_record in rounds:
                                    if not isinstance(round_record, dict):
                                        cleaned_rounds.append(round_record)
                                        continue
                                    cleaned_round = dict(round_record)
                                    round_selected = cleaned_round.get("selectedStoreKeys")
                                    if isinstance(round_selected, list):
                                        cleaned_round["selectedStoreKeys"] = [
                                            item for item in round_selected if str(item) != store_key
                                        ]
                                    round_box_counts = cleaned_round.get("boxCounts")
                                    if isinstance(round_box_counts, dict):
                                        round_box_counts = dict(round_box_counts)
                                        round_box_counts.pop(store_key, None)
                                        cleaned_round["boxCounts"] = round_box_counts
                                    cleaned_rounds.append(cleaned_round)
                                next_value["rounds"] = cleaned_rounds
                            box_counts = next_value.get("boxCounts")
                            if isinstance(box_counts, dict):
                                box_counts.pop(store_key, None)
                            cleaned[key] = next_value
                        else:
                            cleaned[key] = value
                    save_setting(conn, "shipmentManifestSelections", cleaned)
                conn.commit()
            finally:
                conn.close()
            self.send_json({"ok": True, "appData": next_snapshot, "removed": removed})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_export_shipment_manifest(self) -> None:
        try:
            payload = read_json_body(self)
            workbook = manifest_workbook_bytes(payload)
            shipment_date = str(payload.get("shipmentDate") or "").strip()
            round_label = str(payload.get("roundLabel") or "").strip() or "未命名輪次"
            filename = f"{shipment_date}_{round_label}.xlsx"
            encoded_filename = quote(filename)
            body = workbook
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=shipment-manifest.xlsx; filename*=UTF-8''{encoded_filename}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_export_shipment_upload(self) -> None:
        try:
            payload = read_json_body(self)
            workbook = manifest_upload_workbook_bytes(payload)
            shipment_date = str(payload.get("shipmentDate") or "").strip()
            round_label = str(payload.get("roundLabel") or "").strip() or "未命名輪次"
            filename = f"全聯上傳{round_label}_{compact_shipment_date(shipment_date)}.xlsx"
            encoded_filename = quote(filename)
            body = workbook
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=shipment-upload.xlsx; filename*=UTF-8''{encoded_filename}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_export_harvest_shipment(self) -> None:
        try:
            payload = read_json_body(self)
            workbook = harvest_shipment_workbook_bytes(payload)
            shipment_date = str(payload.get("shipmentDate") or "").strip()
            filename = f"{shipment_date}_採收到貨量.xlsx"
            encoded_filename = quote(filename)
            body = workbook
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=harvest-shipment.xlsx; filename*=UTF-8''{encoded_filename}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_export_harvest_large_label(self) -> None:
        try:
            payload = read_json_body(self)
            workbook = harvest_large_label_workbook_bytes(payload)
            harvest_date = str(payload.get("harvestDate") or "").strip()
            filename = f"{harvest_date}_大標.xlsx"
            encoded_filename = quote(filename)
            body = workbook
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=harvest-large-label.xlsx; filename*=UTF-8''{encoded_filename}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_export_field_work_month(self) -> None:
        try:
            payload = read_json_body(self)
            month = str(payload.get("month") or "").strip()
            workbook = field_work_month_workbook_bytes(month)
            filename = f"{month}_田間工作紀錄.xlsx"
            encoded_filename = quote(filename)
            body = workbook
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=field-work-month.xlsx; filename*=UTF-8''{encoded_filename}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_export_field_bt_month_summary(self) -> None:
        try:
            payload = read_json_body(self)
            month = str(payload.get("month") or "").strip()
            workbook = field_bt_month_summary_workbook_bytes(month)
            filename = f"{month}_蘇力菌統計.xlsx"
            encoded_filename = quote(filename)
            body = workbook
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=field-bt-month-summary.xlsx; filename*=UTF-8''{encoded_filename}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_export_daily_vegetable(self) -> None:
        try:
            payload = read_json_body(self)
            workbook = daily_vegetable_workbook_bytes(payload)
            start_date = str(payload.get("startDate") or "").strip()
            filename = f"{start_date}_每日菜量表.xlsx"
            encoded_filename = quote(filename)
            body = workbook
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename=daily-vegetable.xlsx; filename*=UTF-8''{encoded_filename}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_add_daily_vegetable_manual_entry(self) -> None:
        try:
            payload = read_json_body(self)
            conn = connect()
            try:
                entry = add_daily_vegetable_manual_entry(
                    conn,
                    str(payload.get("harvestDate") or ""),
                    str(payload.get("cropName") or ""),
                    str(payload.get("netHouseCode") or ""),
                    payload.get("quantity"),
                )
            finally:
                conn.close()
            self.send_json({"ok": True, "entry": entry})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def handle_update_daily_vegetable_manual_entry(self, entry_id: str) -> None:
        try:
            try:
                normalized_entry_id = int(entry_id)
            except ValueError as exc:
                raise ValueError("entryId 格式錯誤") from exc
            payload = read_json_body(self)
            conn = connect()
            try:
                entry = update_daily_vegetable_manual_entry(
                    conn,
                    normalized_entry_id,
                    str(payload.get("harvestDate") or ""),
                    str(payload.get("cropName") or ""),
                    str(payload.get("netHouseCode") or ""),
                    payload.get("quantity"),
                )
            finally:
                conn.close()
            self.send_json({"ok": True, "entry": entry})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def handle_delete_daily_vegetable_manual_entry(self, entry_id: str) -> None:
        try:
            try:
                normalized_entry_id = int(entry_id)
            except ValueError as exc:
                raise ValueError("entryId 格式錯誤") from exc
            conn = connect()
            try:
                deleted = delete_daily_vegetable_manual_entry(conn, normalized_entry_id)
            finally:
                conn.close()
            if not deleted:
                raise ValueError("找不到手動菜量資料")
            self.send_json({"ok": True, "id": normalized_entry_id})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)

    def handle_save_shipments(self) -> None:
        try:
            payload = read_json_body(self)
            store_key = str(payload.get("storeKey") or "").strip()
            shipment_date = str(payload.get("shipmentDate") or "").strip()
            entries = payload.get("entries") or []
            if not store_key:
                raise ValueError("缺少 storeKey")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", shipment_date):
                raise ValueError("shipmentDate 格式錯誤")
            conn = connect()
            try:
                replace_shipments_for_date(conn, store_key, shipment_date, entries)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_stock_adjustments(self) -> None:
        try:
            payload = read_json_body(self)
            store_key = str(payload.get("storeKey") or "").strip()
            shipment_date = str(payload.get("shipmentDate") or "").strip()
            entries = payload.get("entries") or []
            if not store_key:
                raise ValueError("缺少 storeKey")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", shipment_date):
                raise ValueError("shipmentDate 格式錯誤")
            if not isinstance(entries, list):
                raise ValueError("entries 必須是陣列")
            conn = connect()
            try:
                replace_stock_adjustments_for_date(conn, store_key, shipment_date, entries)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_harvest_entries(self) -> None:
        try:
            payload = read_json_body(self)
            harvest_date = str(payload.get("harvestDate") or "").strip()
            sheet_name = str(payload.get("sheetName") or "").strip()
            entries = payload.get("entries") or []
            priority_entries_payload = payload.get("priorities")
            formula_entries_payload = payload.get("formulas")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("harvestDate 格式錯誤")
            if not sheet_name:
                raise ValueError("缺少 sheetName")
            if not isinstance(entries, list):
                raise ValueError("entries 必須是陣列")
            priority_entries = normalize_harvest_priority_entries(priority_entries_payload)
            formula_entries = normalize_harvest_formula_entries(formula_entries_payload)
            conn = connect()
            try:
                replace_harvest_entries_for_date(conn, harvest_date, sheet_name, entries)
                if priority_entries_payload is not None:
                    priorities = get_setting(conn, "harvestCellPriorities") or {}
                    if not isinstance(priorities, dict):
                        priorities = {}
                    prefix = f"{harvest_date}|{sheet_name}|"
                    priorities = {
                        str(key): value
                        for key, value in priorities.items()
                        if not str(key).startswith(prefix)
                    }
                    for entry in priority_entries:
                        priorities[
                            f"{harvest_date}|{sheet_name}|{entry['rowIndex']}|{entry['columnLetter']}"
                        ] = entry["priority"]
                    save_setting(conn, "harvestCellPriorities", priorities)
                if formula_entries_payload is not None:
                    formulas = get_setting(conn, "harvestCellFormulas") or {}
                    if not isinstance(formulas, dict):
                        formulas = {}
                    prefix = f"{harvest_date}|{sheet_name}|"
                    formulas = {
                        str(key): value
                        for key, value in formulas.items()
                        if not str(key).startswith(prefix)
                    }
                    for entry in formula_entries:
                        formulas[
                            f"{harvest_date}|{sheet_name}|{entry['rowIndex']}|{entry['columnLetter']}"
                        ] = entry["formula"]
                    save_setting(conn, "harvestCellFormulas", formulas)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_harvest_messages(self, parsed) -> None:
        try:
            query = parse_qs(parsed.query)
            harvest_date = str((query.get("date") or [""])[0]).strip()
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("harvestDate 格式錯誤")
            conn = connect()
            try:
                messages = get_harvest_messages(conn, harvest_date)
            finally:
                conn.close()
            self.send_json({"ok": True, "harvestDate": harvest_date, "messages": messages})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_add_harvest_message(self) -> None:
        try:
            payload = read_json_body(self)
            harvest_date = str(payload.get("harvestDate") or "").strip()
            message = str(payload.get("message") or "").strip()
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("harvestDate 格式錯誤")
            if not message:
                raise ValueError("留言不可空白")
            if len(message) > 1000:
                raise ValueError("留言最多 1000 字")
            conn = connect()
            try:
                saved_message = add_harvest_message(conn, harvest_date, message)
            finally:
                conn.close()
            self.send_json({"ok": True, "message": saved_message})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_delete_harvest_message(self, message_id: str, parsed) -> None:
        try:
            query = parse_qs(parsed.query)
            harvest_date = str((query.get("date") or [""])[0]).strip()
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("harvestDate 格式錯誤")
            try:
                normalized_message_id = int(message_id)
            except ValueError as exc:
                raise ValueError("messageId 格式錯誤") from exc
            conn = connect()
            try:
                deleted = delete_harvest_message(conn, normalized_message_id, harvest_date)
            finally:
                conn.close()
            self.send_json({"ok": True, "deleted": deleted})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_field_work_messages(self, parsed) -> None:
        try:
            query = parse_qs(parsed.query)
            work_date = str((query.get("date") or [""])[0]).strip()
            zone_name = str((query.get("zoneName") or [""])[0]).strip()
            net_house_code = str((query.get("netHouseCode") or [""])[0]).strip()
            if work_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", work_date):
                raise ValueError("workDate 格式錯誤")
            conn = connect()
            try:
                messages = get_field_work_messages(
                    conn,
                    work_date or None,
                    zone_name if zone_name else None,
                    net_house_code if net_house_code else None,
                )
            finally:
                conn.close()
            self.send_json({
                "ok": True,
                "workDate": work_date,
                "zoneName": zone_name,
                "netHouseCode": net_house_code,
                "messages": messages,
            })
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_add_field_work_message(self) -> None:
        saved_photo_keys: list[str] = []
        try:
            content_type = self.headers.get("Content-Type", "")
            photos: list[dict[str, str]] = []
            if content_type.startswith("multipart/form-data"):
                work_date, zone_name, net_house_code, message, photos, saved_photo_keys = self.parse_field_work_message_form(content_type)
            else:
                payload = read_json_body(self)
                work_date = str(payload.get("workDate") or "").strip()
                zone_name = str(payload.get("zoneName") or "").strip()
                net_house_code = str(payload.get("netHouseCode") or "").strip()
                message = str(payload.get("message") or "").strip()
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", work_date):
                raise ValueError("workDate 格式錯誤")
            if not zone_name or not net_house_code:
                raise ValueError("請選擇場區與網室")
            if not message and not photos:
                raise ValueError("請輸入留言或選擇照片")
            if len(message) > 1000:
                raise ValueError("留言最多 1000 字")
            conn = connect()
            try:
                saved_message = add_field_work_message(conn, work_date, zone_name, net_house_code, message, photos)
                conn.commit()
            finally:
                conn.close()
            self.send_json({"ok": True, "message": saved_message})
        except Exception as exc:
            self.delete_field_work_message_photo_files(saved_photo_keys)
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_delete_field_work_message(self, message_id: str, parsed) -> None:
        try:
            query = parse_qs(parsed.query)
            work_date = str((query.get("date") or [""])[0]).strip()
            zone_name = str((query.get("zoneName") or [""])[0]).strip()
            net_house_code = str((query.get("netHouseCode") or [""])[0]).strip()
            if work_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", work_date):
                raise ValueError("workDate 格式錯誤")
            try:
                normalized_message_id = int(message_id)
            except ValueError as exc:
                raise ValueError("messageId 格式錯誤") from exc
            conn = connect()
            try:
                message_row = conn.execute(
                    """
                    SELECT work_date, zone_name, net_house_code, message
                    FROM field_work_messages
                    WHERE id = ?
                    """,
                    (normalized_message_id,),
                ).fetchone()
                photo_file_names = get_field_work_message_photo_file_names(conn, normalized_message_id)
                deleted = delete_field_work_message(
                    conn,
                    normalized_message_id,
                    work_date or None,
                    zone_name if zone_name else None,
                    net_house_code if net_house_code else None,
                )
                if deleted:
                    conn.commit()
            finally:
                conn.close()
            if deleted:
                self.delete_field_work_message_photo_files(photo_file_names)
            self.send_json({"ok": True, "deleted": deleted})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_city_table_entries(self) -> None:
        try:
            payload = read_json_body(self)
            entries = normalize_city_table_entries(payload.get("cityTableEntries"))
            conn = connect()
            try:
                save_setting(conn, "cityTableEntries", entries)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_yfy_table_entries(self) -> None:
        try:
            payload = read_json_body(self)
            entries = normalize_yfy_table_entries(payload.get("yfyTableEntries"))
            shipment_times = normalize_yfy_shipment_times(payload.get("yfyShipmentTimes"))
            conn = connect()
            try:
                save_setting(conn, "yfyTableEntries", entries)
                save_setting(conn, "yfyShipmentTimes", shipment_times)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_loose_vegetable_table(self) -> None:
        try:
            payload = read_json_body(self)
            columns = normalize_loose_vegetable_columns(payload.get("looseVegetableColumns"))
            columns_by_date = normalize_loose_vegetable_columns_by_date(payload.get("looseVegetableColumnsByDate"))
            loose_vegetable_date = str(payload.get("looseVegetableDate") or "")
            if not columns_by_date and columns and re.fullmatch(r"\d{4}-\d{2}-\d{2}", loose_vegetable_date):
                columns_by_date[loose_vegetable_date] = columns
            entries = normalize_loose_vegetable_table_entries(
                payload.get("looseVegetableTableEntries"),
                columns_by_date or columns,
            )
            conn = connect()
            try:
                save_setting(conn, "looseVegetableColumns", columns)
                save_setting(conn, "looseVegetableColumnsByDate", columns_by_date)
                save_setting(conn, "looseVegetableTableEntries", entries)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_general_channel_table(self) -> None:
        try:
            payload = read_json_body(self)
            columns = normalize_general_channel_columns(payload.get("generalChannelColumns"))
            columns_by_date = normalize_general_channel_columns_by_date(payload.get("generalChannelColumnsByDate"))
            general_channel_date = str(payload.get("generalChannelDate") or "")
            if not columns_by_date and columns and re.fullmatch(r"\d{4}-\d{2}-\d{2}", general_channel_date):
                columns_by_date[general_channel_date] = columns
            entries = normalize_general_channel_table_entries(
                payload.get("generalChannelTableEntries"),
                columns_by_date or columns,
            )
            puqian_entries = normalize_general_channel_puqian_entries(payload.get("generalChannelPuqianEntries"))
            conn = connect()
            try:
                save_setting(conn, "generalChannelColumns", columns)
                save_setting(conn, "generalChannelColumnsByDate", columns_by_date)
                save_setting(conn, "generalChannelTableEntries", entries)
                save_setting(conn, "generalChannelPuqianEntries", puqian_entries)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_field_bt_records(self) -> None:
        try:
            payload = read_json_body(self)
            account_name = account_name_from_user(self.current_user())
            work_date = str(payload.get("workDate") or "").strip()
            area_keys = payload.get("areaKeys")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", work_date):
                raise ValueError("workDate 格式錯誤")
            if not isinstance(area_keys, list):
                raise ValueError("areaKeys 必須是陣列")

            def package_count_from_payload(field_name: str) -> int | None:
                value = payload.get(field_name)
                if value in (None, ""):
                    return None
                try:
                    package_count = int(value)
                except (TypeError, ValueError):
                    raise ValueError("使用包數必須是 1 到 30") from None
                if package_count < 1 or package_count > 30:
                    raise ValueError("使用包數必須是 1 到 30")
                return package_count

            one_two_package_count = package_count_from_payload("oneTwoPackageCount")
            shared_package_count = package_count_from_payload("sharedPackageCount")
            allowed_areas = {
                "oneTwo": "1、2場",
                "three": "3場",
                "five": "3F、5場",
                "four": "4場",
            }
            bt_area_zone_names = {
                "oneTwo": ["一場", "二場"],
                "three": ["三場"],
                "five": ["三場F", "五場"],
                "threeFive": ["三場", "三場F", "五場"],
                "four": ["四場"],
            }
            normalized_records = []
            seen = set()
            selected_zone_names = []
            for area_key in area_keys:
                area_key = str(area_key or "")
                if area_key not in allowed_areas or area_key in seen:
                    continue
                seen.add(area_key)
                selected_zone_names.extend(bt_area_zone_names[area_key])
                normalized_records.append(
                    {
                        "areaKey": area_key,
                        "areaLabel": allowed_areas[area_key],
                        "packageCount": one_two_package_count if area_key == "oneTwo" else shared_package_count,
                    }
                )
            selected_area_keys = set(seen)
            conn = connect()
            try:
                previous_area_keys = set(get_field_bt_area_keys_for_date(conn, work_date))
                selected_zone_name_set = set(selected_zone_names)
                previous_zone_names = {
                    zone_name
                    for area_key in previous_area_keys
                    for zone_name in bt_area_zone_names.get(area_key, [])
                }
                canceled_zone_names = [
                    zone_name
                    for zone_name in previous_zone_names
                    if zone_name not in selected_zone_name_set
                ]
                replace_field_bt_records(conn, work_date, normalized_records)
                synced_field_work_count = ensure_field_work_task_for_zones(
                    conn,
                    work_date,
                    selected_zone_names,
                    "bacillusThuringiensis",
                    "病蟲害防治",
                    account_name,
                )
                deleted_field_work_count = delete_field_work_task_for_zones(
                    conn,
                    work_date,
                    canceled_zone_names,
                    "bacillusThuringiensis",
                )
                conn.commit()
            finally:
                conn.close()
            self.send_json({
                "ok": True,
                "records": normalized_records,
                "syncedFieldWorkCount": synced_field_work_count,
                "deletedFieldWorkCount": deleted_field_work_count,
            })
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_field_work_records(self) -> None:
        try:
            payload = read_json_body(self)
            account_name = account_name_from_user(self.current_user())
            work_date = str(payload.get("workDate") or "").strip()
            zone_name = str(payload.get("zoneName") or "").strip()
            net_house_code = str(payload.get("netHouseCode") or "").strip()
            task_keys = payload.get("taskKeys")
            raw_fertilizer_bag_count = payload.get("fertilizerBagCount")
            direct_sow_crop_key = str(payload.get("directSowCropKey") or "").strip()
            direct_sow_crop_name = str(payload.get("directSowCropName") or "").strip()
            seedling_transplant_crop_key = str(payload.get("seedlingTransplantCropKey") or "").strip()
            seedling_transplant_crop_name = str(payload.get("seedlingTransplantCropName") or "").strip()
            raw_seedling_transplant_crop_items = payload.get("seedlingTransplantCropItems")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", work_date):
                raise ValueError("workDate 格式錯誤")
            if not zone_name:
                raise ValueError("缺少場區")
            if not net_house_code:
                raise ValueError("缺少網室")
            if not isinstance(task_keys, list):
                raise ValueError("taskKeys 必須是陣列")
            fertilizer_bag_count = None
            fertilizer_bag_count_text = "" if raw_fertilizer_bag_count is None else str(raw_fertilizer_bag_count).strip()
            if fertilizer_bag_count_text:
                try:
                    fertilizer_bag_count = float(fertilizer_bag_count_text)
                except (TypeError, ValueError):
                    raise ValueError("肥料使用包數必須是數字") from None
                if fertilizer_bag_count < 0:
                    raise ValueError("肥料使用包數不可小於 0")
            if any(str(task_key or "") == "soilPreparationTractorClean" for task_key in task_keys) and not fertilizer_bag_count_text:
                raise ValueError("勾選翻土整地,曳引機清潔後，請填寫肥料使用包數。")

            def normalize_crop_items(
                raw_items: object,
                fallback_key: str = "",
                fallback_name: str = "",
                fallback_tray_count: object = None,
                require_items: bool = False,
                require_integer_tray_count: bool = False,
            ) -> list[dict[str, Any]]:
                def normalize_tray_count(raw_tray_count: object) -> int | float | None:
                    tray_text = "" if raw_tray_count is None else str(raw_tray_count).strip()
                    if require_integer_tray_count:
                        if not tray_text:
                            raise ValueError("苗移植盤數不可空白")
                        if not re.fullmatch(r"\d+", tray_text):
                            raise ValueError("苗移植盤數必須是 0 以上整數")
                        return int(tray_text)
                    if raw_tray_count in (None, ""):
                        return None
                    try:
                        tray_count = float(raw_tray_count)
                    except (TypeError, ValueError):
                        raise ValueError("盤數必須是數字") from None
                    if tray_count < 0:
                        raise ValueError("盤數不可小於 0")
                    return tray_count

                crop_items = []
                if isinstance(raw_items, list):
                    for raw_item in raw_items:
                        if not isinstance(raw_item, dict):
                            if require_items:
                                raise ValueError("苗移植作物資料格式錯誤")
                            continue
                        crop_key = str(raw_item.get("cropKey") or "").strip()
                        crop_name = str(raw_item.get("cropName") or "").strip()
                        raw_tray_count = raw_item.get("trayCount")
                        if require_items and not (crop_key or crop_name):
                            raise ValueError("勾選苗移植後，請選擇作物")
                        tray_count = normalize_tray_count(raw_tray_count)
                        if crop_key or crop_name:
                            crop_items.append({"cropKey": crop_key, "cropName": crop_name[:80], "trayCount": tray_count})
                elif fallback_key or fallback_name:
                    tray_count = normalize_tray_count(fallback_tray_count)
                    crop_items.append({"cropKey": fallback_key, "cropName": fallback_name[:80], "trayCount": tray_count})
                if require_items and not crop_items:
                    raise ValueError("勾選苗移植後，請選擇作物")
                return crop_items

            direct_sow_crop_items = normalize_crop_items(
                None,
                direct_sow_crop_key,
                direct_sow_crop_name,
            )
            seedling_transplant_required = any(str(task_key or "") == "seedlingTransplant" for task_key in task_keys)
            seedling_transplant_crop_items = normalize_crop_items(
                raw_seedling_transplant_crop_items,
                seedling_transplant_crop_key,
                seedling_transplant_crop_name,
                require_items=seedling_transplant_required,
                require_integer_tray_count=seedling_transplant_required,
            )
            allowed_tasks = {
                "soilPreparationTractorClean": "翻土整地,曳引機清潔",
                "burn": "燒",
                "weed": "除草",
                "directSow": "直播",
                "seedlingTransplant": "苗移植",
                "bacillusThuringiensis": "病蟲害防治",
            }
            normalized_tasks = []
            seen = set()
            for task_key in task_keys:
                task_key = str(task_key or "")
                if task_key not in allowed_tasks or task_key in seen:
                    continue
                seen.add(task_key)
                normalized_task = {"taskKey": task_key, "taskLabel": allowed_tasks[task_key]}
                if task_key == "soilPreparationTractorClean":
                    normalized_task["fertilizerBagCount"] = fertilizer_bag_count
                if task_key == "directSow":
                    first_crop = direct_sow_crop_items[0] if direct_sow_crop_items else {"cropKey": "", "cropName": ""}
                    normalized_task["cropKey"] = first_crop["cropKey"]
                    normalized_task["cropName"] = first_crop["cropName"]
                    normalized_task["cropItems"] = direct_sow_crop_items
                if task_key == "seedlingTransplant":
                    first_crop = seedling_transplant_crop_items[0] if seedling_transplant_crop_items else {"cropKey": "", "cropName": ""}
                    normalized_task["cropKey"] = first_crop["cropKey"]
                    normalized_task["cropName"] = first_crop["cropName"]
                    normalized_task["cropItems"] = seedling_transplant_crop_items
                normalized_task["accountName"] = account_name
                normalized_tasks.append(normalized_task)
            conn = connect()
            try:
                if is_field_work_blocked_by_planting(conn, work_date, zone_name, net_house_code):
                    raise ValueError("此網室目前種植中，不能新增工作紀錄")
                field_work_audit_logged = replace_field_work_records(
                    conn,
                    work_date,
                    zone_name,
                    net_house_code,
                    normalized_tasks,
                    account_name,
                )
                net_house_status_record = sync_net_house_status_from_field_work(
                    conn,
                    work_date,
                    zone_name,
                    net_house_code,
                    normalized_tasks,
                )
                conn.commit()
            finally:
                conn.close()
            self.send_json({
                "ok": True,
                "tasks": normalized_tasks,
                "netHouseStatusRecord": net_house_status_record,
                "netHouseStatusRecordDate": work_date,
                "fieldWorkAuditLogged": field_work_audit_logged,
            })
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_net_house_status_record(self) -> None:
        try:
            payload = read_json_body(self)
            record_date = str(payload.get("recordDate") or "").strip()
            zone_name = str(payload.get("zoneName") or "").strip()
            net_house_code = str(payload.get("netHouseCode") or "").strip()
            status = str(payload.get("status") or "").strip()
            crop_name = str(payload.get("cropName") or "").strip()[:160]
            planted_date = str(payload.get("plantedDate") or "").strip()
            harvest_date = str(payload.get("harvestDate") or "").strip()
            raw_estimated_quantity = payload.get("estimatedQuantity")
            raw_harvest_quantity = payload.get("harvestQuantity")
            raw_crop_items = payload.get("cropItems")
            lunch_marked = bool(payload.get("lunchMarked"))
            allowed_statuses = {"空園", "苗室", "種植"}
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", record_date):
                raise ValueError("recordDate 格式錯誤")
            if not zone_name:
                raise ValueError("缺少場區")
            if not net_house_code:
                raise ValueError("缺少網室")
            if status not in allowed_statuses:
                raise ValueError("狀態必須是空園、苗室或種植")
            if planted_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", planted_date):
                raise ValueError("種植日期格式錯誤")
            if harvest_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("採收日期格式錯誤")
            def parse_optional_nonnegative_number(raw_value: object, label: str) -> float | None:
                if raw_value is None:
                    return None
                if isinstance(raw_value, str):
                    raw_value = raw_value.strip()
                    if raw_value == "":
                        return None
                try:
                    number = float(raw_value)
                except (TypeError, ValueError):
                    raise ValueError(f"{label}必須是數字") from None
                if number < 0:
                    raise ValueError(f"{label}不可小於 0")
                return number

            estimated_quantity = parse_optional_nonnegative_number(raw_estimated_quantity, "預估量")
            harvest_quantity = parse_optional_nonnegative_number(raw_harvest_quantity, "採收量")

            def normalize_status_crop_items(raw_items: object) -> list[dict[str, Any]]:
                crop_items = []
                if not isinstance(raw_items, list):
                    return crop_items
                for raw_item in raw_items:
                    if not isinstance(raw_item, dict):
                        continue
                    item_crop_name = str(raw_item.get("cropName") or "").strip()[:160]
                    if not item_crop_name:
                        continue
                    item_harvest_date = str(raw_item.get("harvestDate") or "").strip()
                    if item_harvest_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", item_harvest_date):
                        raise ValueError("採收日期格式錯誤")
                    item_estimated_quantity = parse_optional_nonnegative_number(
                        raw_item.get("estimatedQuantity"),
                        "預估量",
                    )
                    item_harvest_quantity = parse_optional_nonnegative_number(
                        raw_item.get("harvestQuantity"),
                        "採收量",
                    )
                    item_destroyed = bool(raw_item.get("destroyed"))
                    crop_items.append(
                        {
                            "cropName": item_crop_name,
                            "harvestDate": item_harvest_date,
                            "estimatedQuantity": 0 if item_destroyed else item_estimated_quantity,
                            "harvestQuantity": 0 if item_destroyed else item_harvest_quantity,
                            "destroyed": item_destroyed,
                        }
                    )
                return crop_items

            crop_items = normalize_status_crop_items(raw_crop_items)
            if status == "種植":
                effective_planted_date = planted_date or record_date
                if harvest_date and harvest_date < effective_planted_date:
                    raise ValueError("採收日期不可早於種植日期，可與種植日期同一天")
                if any(
                    str(item.get("harvestDate") or "").strip()
                    and str(item.get("harvestDate") or "").strip() < effective_planted_date
                    for item in crop_items
                ):
                    raise ValueError("採收日期不可早於種植日期，可與種植日期同一天")
            conn = connect()
            try:
                saved_record = save_net_house_status_record(
                    conn,
                    record_date,
                    zone_name,
                    net_house_code,
                    status,
                    crop_name,
                    planted_date,
                    harvest_date,
                    estimated_quantity,
                    harvest_quantity,
                    crop_items,
                    lunch_marked,
                )
            finally:
                conn.close()
            self.send_json({
                "ok": True,
                "record": saved_record,
            })
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_store_note(self, store_key: str) -> None:
        try:
            payload = read_json_body(self)
            conn = connect()
            try:
                save_store_note(conn, store_key, str(payload.get("note") or ""))
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_store_route(self, store_key: str) -> None:
        try:
            payload = read_json_body(self)
            conn = connect()
            try:
                save_store_route(conn, store_key, str(payload.get("route") or ""))
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_product_filter(self) -> None:
        try:
            payload = read_json_body(self)
            visible_product_codes_by_date = payload.get("visibleProductCodesByDate")
            if isinstance(visible_product_codes_by_date, dict):
                normalized = {
                    str(date): [str(code) for code in codes]
                    for date, codes in visible_product_codes_by_date.items()
                    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date)) and isinstance(codes, list)
                }
                conn = connect()
                try:
                    save_setting(conn, "visibleProductCodesByDate", normalized)
                finally:
                    conn.close()
                self.send_json({"ok": True})
                return
            visible_product_codes = payload.get("visibleProductCodes")
            if not isinstance(visible_product_codes, list):
                raise ValueError("visibleProductCodes 必須是陣列，或 visibleProductCodesByDate 必須是物件")
            conn = connect()
            try:
                save_setting(conn, "visibleProductCodes", [str(code) for code in visible_product_codes])
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_harvest_product_filter(self) -> None:
        try:
            payload = read_json_body(self)
            visible_harvest_product_keys_by_date = payload.get("visibleHarvestProductKeysByDate")
            if isinstance(visible_harvest_product_keys_by_date, dict):
                normalized = {
                    str(date): [str(key) for key in keys]
                    for date, keys in visible_harvest_product_keys_by_date.items()
                    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date)) and isinstance(keys, list)
                }
                conn = connect()
                try:
                    save_setting(conn, "visibleHarvestProductKeysByDate", normalized)
                finally:
                    conn.close()
                self.send_json({"ok": True})
                return
            visible_harvest_product_keys = payload.get("visibleHarvestProductKeys")
            if not isinstance(visible_harvest_product_keys, list):
                raise ValueError("visibleHarvestProductKeys 必須是陣列，或 visibleHarvestProductKeysByDate 必須是物件")
            conn = connect()
            try:
                save_setting(conn, "visibleHarvestProductKeys", [str(key) for key in visible_harvest_product_keys])
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_daily_vegetable_product_filter(self) -> None:
        try:
            payload = read_json_body(self)
            visible_daily_vegetable_product_keys = payload.get("visibleDailyVegetableProductKeys")
            if not isinstance(visible_daily_vegetable_product_keys, list):
                raise ValueError("visibleDailyVegetableProductKeys 必須是陣列")
            conn = connect()
            try:
                save_setting(
                    conn,
                    "visibleDailyVegetableProductKeys",
                    [str(key) for key in visible_daily_vegetable_product_keys],
                )
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_harvest_crew_columns(self) -> None:
        try:
            payload = read_json_body(self)
            harvest_date = str(payload.get("harvestDate") or "").strip()
            harvest_crew_columns = payload.get("harvestCrewColumns")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("harvestDate 格式錯誤")
            if not isinstance(harvest_crew_columns, list):
                raise ValueError("harvestCrewColumns 必須是陣列")
            normalized = []
            seen_cols = set()
            for item in harvest_crew_columns:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name") or "").strip()
                category = str(item.get("category") or "有機").strip()
                if category not in {"有機", "轉型"}:
                    category = "有機"
                col = int(item.get("col") or 0)
                if not name or col < 14 or col in seen_cols:
                    continue
                seen_cols.add(col)
                normalized.append(
                    {
                        "name": name[:12],
                        "category": category,
                        "col": col,
                        "column": str(item.get("column") or ""),
                        "source": "custom" if item.get("source") == "custom" else "default",
                    }
                )
            conn = connect()
            try:
                columns_by_date = get_setting(conn, "harvestCrewColumnsByDate") or {}
                if not isinstance(columns_by_date, dict):
                    columns_by_date = {}
                columns_by_date[harvest_date] = normalized
                save_setting(conn, "harvestCrewColumnsByDate", columns_by_date)
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_harvest_field_extra_columns(self) -> None:
        try:
            payload = read_json_body(self)
            harvest_date = str(payload.get("harvestDate") or "").strip()
            harvest_field_extra_columns = payload.get("harvestFieldExtraColumns")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("harvestDate 格式錯誤")
            if not isinstance(harvest_field_extra_columns, list):
                raise ValueError("harvestFieldExtraColumns 必須是陣列")

            def column_letter_from_index(index: int) -> str:
                text = ""
                value = index
                while value > 0:
                    value, remainder = divmod(value - 1, 26)
                    text = chr(65 + remainder) + text
                return text

            base_sections = {
                4: "quantity",
                5: "quantity",
                501: "quantity",
                6: "quantity",
                7: "quantity",
                9: "unfinished",
                10: "unfinished",
                502: "unfinished",
                11: "unfinished",
                12: "unfinished",
            }
            normalized = []
            seen_cols = set()
            for item in harvest_field_extra_columns:
                if not isinstance(item, dict):
                    continue
                col = int(item.get("col") or 0)
                base_col = int(item.get("baseCol") or 0)
                section = str(item.get("section") or "").strip()
                expected_section = base_sections.get(base_col)
                if not expected_section or section != expected_section or col < 600 or col in seen_cols:
                    continue
                seen_cols.add(col)
                normalized.append(
                    {
                        "section": section,
                        "baseCol": base_col,
                        "col": col,
                        "column": column_letter_from_index(col),
                    }
                )

            conn = connect()
            try:
                columns_by_date = get_setting(conn, "harvestFieldExtraColumnsByDate") or {}
                if not isinstance(columns_by_date, dict):
                    columns_by_date = {}
                columns_by_date[harvest_date] = normalized
                save_setting(conn, "harvestFieldExtraColumnsByDate", columns_by_date)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_harvest_field_3f_expanded(self) -> None:
        try:
            payload = read_json_body(self)
            harvest_date = str(payload.get("harvestDate") or "").strip()
            expanded = payload.get("harvestField3FExpanded")
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", harvest_date):
                raise ValueError("harvestDate 格式錯誤")
            if not isinstance(expanded, dict):
                raise ValueError("harvestField3FExpanded 必須是物件")
            normalized = {
                "unfinished": expanded.get("unfinished") if isinstance(expanded.get("unfinished"), bool) else True,
                "quantity": expanded.get("quantity") if isinstance(expanded.get("quantity"), bool) else True,
            }
            conn = connect()
            try:
                expanded_by_date = get_setting(conn, "harvestField3FExpandedByDate") or {}
                if not isinstance(expanded_by_date, dict):
                    expanded_by_date = {}
                expanded_by_date[harvest_date] = normalized
                save_setting(conn, "harvestField3FExpandedByDate", expanded_by_date)
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_harvest_conversion_settings(self) -> None:
        try:
            payload = read_json_body(self)
            settings = payload.get("harvestConversionSettings")
            if not isinstance(settings, dict):
                raise ValueError("harvestConversionSettings 必須是物件")
            package_keys = {"mostly", "leafy220", "crownDaisy", "kaleSesame", "basil", "cilantro"}
            packaging_channel_keys = {"city", "generalChannel", "looseVegetable", "yfy"}
            raw_precool = settings.get("hasPrecoolByDate") or {}
            raw_packaged_channels = settings.get("packagedChannelsByDate") or {}
            raw_packages = settings.get("productPackages") or {}
            if not isinstance(raw_precool, dict):
                raw_precool = {}
            if not isinstance(raw_packaged_channels, dict):
                raw_packaged_channels = {}
            if not isinstance(raw_packages, dict):
                raw_packages = {}
            normalized = {
                "hasPrecoolByDate": {
                    str(date): bool(checked)
                    for date, checked in raw_precool.items()
                    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date)) and isinstance(checked, bool)
                },
                "packagedChannelsByDate": {
                    str(date): {
                        str(channel_key): bool(checked)
                        for channel_key, checked in channels.items()
                        if str(channel_key) in packaging_channel_keys and isinstance(checked, bool)
                    }
                    for date, channels in raw_packaged_channels.items()
                    if (
                        re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date))
                        and isinstance(channels, dict)
                    )
                },
                "productPackages": {
                    str(product_key): str(package_key)
                    for product_key, package_key in raw_packages.items()
                    if str(product_key).strip() and str(package_key) in package_keys
                },
            }
            conn = connect()
            try:
                save_setting(conn, "harvestConversionSettings", normalized)
                sync_version = mark_inventory_sync_changed(conn)
            finally:
                conn.close()
            self.send_json({"ok": True, "syncVersion": sync_version})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def handle_save_shipment_manifest_selection(self) -> None:
        try:
            payload = read_json_body(self)
            shipment_manifest_selections = payload.get("shipmentManifestSelections")
            if not isinstance(shipment_manifest_selections, dict):
                raise ValueError("shipmentManifestSelections 必須是物件")
            normalized = {}
            for key, value in shipment_manifest_selections.items():
                date_key = str(key)
                if isinstance(value, list):
                    normalized[date_key] = [str(store_key) for store_key in value]
                elif isinstance(value, dict):
                    normalized[date_key] = value
            conn = connect()
            try:
                save_setting(conn, "shipmentManifestSelections", normalized)
            finally:
                conn.close()
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=500)

    def serve_field_work_message_photo(self, request_path: str, include_body: bool = True) -> None:
        filename = unquote(request_path.split(f"{FIELD_WORK_MESSAGE_PHOTO_URL_PREFIX}/", 1)[1])
        if not filename or Path(filename).name != filename:
            self.send_error(404, "Not found")
            return
        key = f"{FIELD_WORK_MESSAGE_PHOTO_STORAGE_PREFIX}/{filename}"
        store = storage_backend()
        if not store.exists(key):
            self.send_error(404, "Not found")
            return
        content, stored_content_type = store.read_bytes(key)
        content_type = stored_content_type or mimetypes.guess_type(filename)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Cache-Control", "public, max-age=31536000, immutable")
        self.end_headers()
        if include_body:
            self.wfile.write(content)

    def parse_field_work_message_form(
        self,
        content_type: str,
    ) -> tuple[str, str, str, str, list[dict[str, str]], list[str]]:
        marker = "boundary="
        if marker not in content_type:
            raise ValueError("Upload must be multipart/form-data")
        boundary = content_type.split(marker, 1)[1].split(";", 1)[0].strip().strip('"').encode()
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        parts = body.split(b"--" + boundary)
        work_date = ""
        zone_name = ""
        net_house_code = ""
        message = ""
        photos: list[dict[str, str]] = []
        saved_keys: list[str] = []
        store = storage_backend()
        for part in parts:
            if b"\r\n\r\n" not in part:
                continue
            header, _, payload = part.partition(b"\r\n\r\n")
            header_text = header.decode("utf-8", errors="ignore")
            if payload.endswith(b"\r\n"):
                payload = payload[:-2]
            field_name = self.multipart_name_from_header(header_text)
            if field_name == "workDate":
                work_date = payload.decode("utf-8", errors="ignore").strip()
                continue
            if field_name == "zoneName":
                zone_name = payload.decode("utf-8", errors="ignore").strip()
                continue
            if field_name == "netHouseCode":
                net_house_code = payload.decode("utf-8", errors="ignore").strip()
                continue
            if field_name == "message":
                message = payload.decode("utf-8", errors="ignore").strip()
                continue
            if field_name != "photos" or not payload:
                continue
            if len(photos) >= MAX_FIELD_WORK_MESSAGE_PHOTOS:
                raise ValueError(f"照片最多 {MAX_FIELD_WORK_MESSAGE_PHOTOS} 張")
            if len(payload) > MAX_FIELD_WORK_MESSAGE_PHOTO_BYTES:
                raise ValueError("單張照片最多 8MB")
            filename = self.filename_from_header(header_text)
            if not filename:
                continue
            content_type_value = self.content_type_from_header(header_text)
            normalized_content_type, suffix = self.normalize_field_work_message_photo_type(
                content_type_value,
                filename,
            )
            safe_original_name = Path(filename).name.replace("/", "_").replace("\\", "_")
            file_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}{suffix}"
            storage_key = f"{FIELD_WORK_MESSAGE_PHOTO_STORAGE_PREFIX}/{file_name}"
            store.save_bytes(storage_key, payload, normalized_content_type)
            saved_keys.append(storage_key)
            photos.append(
                {
                    "fileName": file_name,
                    "originalName": safe_original_name,
                    "contentType": normalized_content_type,
                    "urlPath": f"{FIELD_WORK_MESSAGE_PHOTO_URL_PREFIX}/{quote(file_name)}",
                }
            )
        return work_date, zone_name, net_house_code, message, photos, saved_keys

    def normalize_field_work_message_photo_type(self, content_type: str, filename: str) -> tuple[str, str]:
        normalized = content_type.split(";", 1)[0].strip().lower()
        suffix = Path(filename).suffix.lower()
        if normalized in FIELD_WORK_MESSAGE_PHOTO_TYPES:
            return normalized, FIELD_WORK_MESSAGE_PHOTO_TYPES[normalized]
        if suffix in FIELD_WORK_MESSAGE_PHOTO_EXTENSIONS:
            return FIELD_WORK_MESSAGE_PHOTO_EXTENSIONS[suffix], ".jpg" if suffix == ".jpeg" else suffix
        raise ValueError("只支援 JPG、PNG、WEBP、GIF、HEIC 照片")

    def multipart_name_from_header(self, header: str) -> str:
        match = re.search(r'name="([^"]+)"', header)
        return unquote(match.group(1)) if match else ""

    def content_type_from_header(self, header: str) -> str:
        match = re.search(r"Content-Type:\s*([^\r\n]+)", header, flags=re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def delete_field_work_message_photo_files(self, file_names: list[str]) -> None:
        store = storage_backend()
        for file_name in file_names:
            normalized_name = str(file_name or "")
            if normalized_name.startswith(f"{FIELD_WORK_MESSAGE_PHOTO_STORAGE_PREFIX}/"):
                relative_name = normalized_name.split("/", 1)[1]
                if Path(relative_name).name != relative_name:
                    continue
                storage_key = normalized_name
            else:
                if Path(normalized_name).name != normalized_name:
                    continue
                storage_key = f"{FIELD_WORK_MESSAGE_PHOTO_STORAGE_PREFIX}/{normalized_name}"
            store.delete(storage_key)

    def parse_import_form(self) -> tuple[Path, str]:
        content_type = self.headers.get("Content-Type", "")
        marker = "boundary="
        if marker not in content_type:
            raise ValueError("Upload must be multipart/form-data")
        boundary = content_type.split(marker, 1)[1].strip().strip('"').encode()
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        parts = body.split(b"--" + boundary)
        sales_date: str | None = None
        saved_path: Path | None = None
        for part in parts:
            if b"\r\n\r\n" not in part:
                continue
            header, _, payload = part.partition(b"\r\n\r\n")
            header_text = header.decode("utf-8", errors="ignore")
            payload = payload.rsplit(b"\r\n", 1)[0]
            if 'name="salesDate"' in header_text:
                sales_date = payload.decode("utf-8").strip()
                continue
            if 'name="salesFile"' not in header_text:
                continue
            filename = self.filename_from_header(header_text)
            if not filename.lower().endswith(".xlsx"):
                raise ValueError("Only .xlsx files are supported")
            UPLOADS.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = Path(filename).name.replace("/", "_").replace("\\", "_")
            path = UPLOADS / f"{timestamp}_{safe_name}"
            path.write_bytes(payload)
            saved_path = path
        if not saved_path:
            raise ValueError("No salesFile upload found")
        if not sales_date or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", sales_date):
            raise ValueError("Please choose a valid import date")
        return saved_path, sales_date

    def filename_from_header(self, header: str) -> str:
        match = re.search(r'filename="([^"]+)"', header)
        if match:
            return unquote(match.group(1))
        return "salesdata.xlsx"

    def send_json(self, payload: dict, status: int = 200, headers: list[tuple[str, str]] | None = None) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        for name, value in headers or []:
            self.send_header(name, value)
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    bootstrap_db()
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "4174"))
    server = ThreadingHTTPServer((host, port), AppHandler)
    display_host = "0.0.0.0" if host in {"", "0.0.0.0"} else host
    print(f"Serving {DOCS} with database-backed API at http://{display_host}:{port}/")
    server.serve_forever()


if __name__ == "__main__":
    main()
